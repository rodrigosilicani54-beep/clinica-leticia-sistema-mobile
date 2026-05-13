# =======================
# 1. IMPORTS
# =======================
import os
import sys
import base64
import hmac
from db import get_connection
import threading
from flask import Flask, request, jsonify, send_file, send_from_directory, session
from flask_cors import CORS
from psycopg2.extras import Json
from werkzeug.security import check_password_hash, generate_password_hash
from db import get_connection
import unicodedata
from pathlib import Path
import json
import io
import socket
from datetime import datetime, date
from time import monotonic

try:
    import webview
except ImportError:
    webview = None

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("⚠️ openpyxl não instalado - exportação Excel desabilitada")

print("APP.PY CORRETO FOI CARREGADO")

# =======================
# CONFIGURAÇÃO DE CAMINHOS
# =======================
# Caminho seguro para backup/exportação de arquivos
from pathlib import Path

#  Pasta base do usuário (funciona em qualquer PC)
USER_BASE_DIR = Path.home() / "AppData" / "Local" / "ClinicaLeticiaSegretti"

#  Subpastas
BACKUP_DIR = USER_BASE_DIR / "backup"
REPORT_DIR = USER_BASE_DIR / "relatorios"

#  Garantir que existem
BACKUP_DIR.mkdir(parents=True, exist_ok=True)
REPORT_DIR.mkdir(parents=True, exist_ok=True)

def get_backup_filepath(filename):
    """
    Retorna o caminho seguro para salvar um arquivo no diretório de backup
    
    Args:
        filename (str): Nome do arquivo a ser salvo
        
    Returns:
        str: Caminho completo e seguro para o arquivo
    """
    return str(BACKUP_DIR / filename)

print(f" Diretório de backup configurado: {BACKUP_DIR}")

def get_base_path():
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = get_base_path()
APP_HOST = os.environ.get('CLINICA_FLASK_HOST', '0.0.0.0')
APP_PORT = int(os.environ.get('CLINICA_FLASK_PORT') or os.environ.get('PORT') or '5000')
AUDIT_ACCESS_TTL_SECONDS = int(os.environ.get('AUDITORIA_TTL_SECONDS') or '900')


def get_runtime_config_paths():
    paths = []
    explicit_path = os.environ.get('CLINICA_CONFIG') or os.environ.get('CLINICA_DB_CONFIG') or os.environ.get('DB_CONFIG_FILE')
    if explicit_path:
        paths.append(Path(explicit_path))

    local_app_data = os.environ.get('LOCALAPPDATA')
    if local_app_data:
        paths.append(Path(local_app_data) / 'ClinicaLeticiaSegretti' / 'db_config.json')

    if getattr(sys, 'frozen', False):
        paths.append(Path(sys.executable).resolve().parent / 'db_config.local.json')
        bundle_dir = getattr(sys, '_MEIPASS', None)
        if bundle_dir:
            paths.append(Path(bundle_dir) / 'db_config.local.json')

    paths.append(Path(BASE_DIR) / 'db_config.local.json')
    paths.append(Path(__file__).resolve().with_name('db_config.local.json'))
    return paths


def get_runtime_config_value(env_name, *config_names, default=None):
    env_value = os.environ.get(env_name)
    if env_value not in (None, ''):
        return env_value

    seen = set()
    for path in get_runtime_config_paths():
        try:
            resolved = str(path.resolve())
            if resolved in seen or not path.is_file():
                continue
            seen.add(resolved)
            with path.open('r', encoding='utf-8') as f:
                config = json.load(f)
            if not isinstance(config, dict):
                continue
            for name in config_names:
                value = config.get(name)
                if value not in (None, ''):
                    return value
        except Exception:
            continue
    return default


def get_auditoria_password():
    value = get_runtime_config_value(
        'AUDITORIA_SENHA',
        'AUDITORIA_SENHA',
        'auditoria_senha',
        'audit_password',
        'auditPassword'
    )
    return str(value or '').strip()


def get_current_timestamp_seconds():
    return int(datetime.utcnow().timestamp())

# =======================
# FLASK API
# =======================
app = Flask(__name__, static_folder=os.path.join(BASE_DIR, 'static'), static_url_path='/static')
CORS(
    app,
    supports_credentials=True,
    resources={r"/api/*": {"origins": [f"http://127.0.0.1:{APP_PORT}", f"http://localhost:{APP_PORT}", "null"]}}
)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'sistema-secret-key')
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=False
)

def get_file_path(*parts):
    return os.path.join(BASE_DIR, *parts)

@app.route('/')
@app.route('/index.html')
def home():
    index_path = get_file_path('index.html')
    if os.path.isfile(index_path):
        return send_file(index_path)
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/mobile')
@app.route('/mobile/')
@app.route('/mobile.html')
def mobile_home():
    mobile_path = get_file_path('mobile.html')
    if os.path.isfile(mobile_path):
        return send_file(mobile_path)
    return jsonify({'error': 'Mobile app file does not exist.'}), 404

@app.route('/<path:requested_path>')
def serve_any_file(requested_path):
    # Serve any bundled frontend files (CSS/JS/html) when available.
    candidate = get_file_path(requested_path)
    if os.path.isfile(candidate):
        return send_file(candidate)

    candidate_static = get_file_path('static', requested_path)
    if os.path.isfile(candidate_static):
        return send_file(candidate_static)

    if requested_path in ('', 'index.html'):
        return home()

    return jsonify({'error': 'File does not exist.'}), 404

# Normalize level values to technical identifiers used by the app
def normalize_level(level):
    if not level:
        return 'viewer'
    s = str(level).strip().lower()
    mapping = {
        'administrador principal': 'admin',
        'administrador': 'admin',
        'admin': 'admin',
        'editor': 'editor',
        'edit': 'editor',
        'visualizador': 'viewer',
        'viewer': 'viewer',
        'visual': 'viewer'
    }
    return mapping.get(s, s)


def normalize_date_for_db(value):
    if not value:
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        # Accept dd/mm/yyyy or yyyy-mm-dd
        if '/' in value:
            parts = value.split('/')
            if len(parts) == 3:
                day, month, year = parts
                try:
                    return datetime(int(year), int(month), int(day)).date()
                except Exception:
                    pass
        try:
            return datetime.fromisoformat(value).date()
        except Exception:
            pass
        try:
            return datetime.strptime(value, '%Y/%m/%d').date()
        except Exception:
            pass
    elif isinstance(value, datetime):
        return value.date()
    return value


def time_to_minutes(value):
    try:
        if hasattr(value, 'hour') and hasattr(value, 'minute'):
            return int(value.hour) * 60 + int(value.minute)
        parts = str(value).strip().split(':')
        if len(parts) < 2:
            return None
        hours, minutes = parts[0], parts[1]
        return int(hours) * 60 + int(minutes)
    except Exception:
        return None


def times_overlap(start_a, end_a, start_b, end_b):
    start_a_min = time_to_minutes(start_a)
    end_a_min = time_to_minutes(end_a)
    start_b_min = time_to_minutes(start_b)
    end_b_min = time_to_minutes(end_b)
    if None in (start_a_min, end_a_min, start_b_min, end_b_min):
        return False
    return start_a_min < end_b_min and end_a_min > start_b_min


PASSWORD_HASH_PREFIXES = ('scrypt:', 'pbkdf2:', 'argon2:', 'bcrypt:', '$2a$', '$2b$', '$2y$')
VALID_APPOINTMENT_STATUSES = {
    'agendado',
    'pre_atendimento',
    'confirmado',
    'chegou',
    'em_atendimento',
    'em_analise',
    'online',
    'finalizado',
    'cancelado_profissional',
    'cancelado_paciente',
    'faltou'
}
VALID_WAITLIST_STATUSES = {
    'aguardando',
    'em_contato',
    'encaixado',
    'cancelado'
}
WAITLIST_STATUS_ALIASES = {
    'aguardando': 'aguardando',
    'espera': 'aguardando',
    'na fila': 'aguardando',
    'em_contato': 'em_contato',
    'em contato': 'em_contato',
    'contato': 'em_contato',
    'encaixado': 'encaixado',
    'encaixe': 'encaixado',
    'agendado': 'encaixado',
    'cancelado': 'cancelado',
    'cancelada': 'cancelado'
}
VALID_WAITLIST_PRIORITIES = {
    'baixa',
    'normal',
    'alta',
    'urgente'
}
WAITLIST_PRIORITY_ALIASES = {
    'baixa': 'baixa',
    'baixo': 'baixa',
    'normal': 'normal',
    'media': 'normal',
    'medio': 'normal',
    'alta': 'alta',
    'alto': 'alta',
    'urgente': 'urgente',
    'urgencia': 'urgente'
}
PATIENT_ROOM_CONFLICT_IGNORED_STATUSES = {
    'cancelado_profissional',
    'cancelado_paciente',
    'faltou',
    'nao_compareceu',
    'online'
}
APPOINTMENT_STATUS_ALIASES = {
    'agendado': 'agendado',
    'agendada': 'agendado',
    'agendados': 'agendado',
    'marcado': 'agendado',
    'marcada': 'agendado',
    'pre_atendimento': 'pre_atendimento',
    'pre atendimento': 'pre_atendimento',
    'pre-atendimento': 'pre_atendimento',
    'pré atendimento': 'pre_atendimento',
    'confirmado': 'confirmado',
    'confirmada': 'confirmado',
    'confirmou': 'confirmado',
    'chegou': 'chegou',
    'presente': 'chegou',
    'em_atendimento': 'em_atendimento',
    'em atendimento': 'em_atendimento',
    'atendimento': 'em_atendimento',
    'em_analise': 'em_analise',
    'em analise': 'em_analise',
    'em análise': 'em_analise',
    'analise': 'em_analise',
    'análise': 'em_analise',
    'online': 'online',
    'on-line': 'online',
    'remoto': 'online',
    'remota': 'online',
    'finalizado': 'finalizado',
    'finalizada': 'finalizado',
    'concluido': 'finalizado',
    'concluida': 'finalizado',
    'cancelado_profissional': 'cancelado_profissional',
    'cancelado profissional': 'cancelado_profissional',
    'cancelado pelo profissional': 'cancelado_profissional',
    'cancelado_paciente': 'cancelado_paciente',
    'cancelado paciente': 'cancelado_paciente',
    'cancelado pelo paciente': 'cancelado_paciente',
    'nao_compareceu': 'faltou',
    'não_compareceu': 'faltou',
    'nao compareceu': 'faltou',
    'não compareceu': 'faltou',
    'falta': 'faltou',
    'faltou': 'faltou'
}


def is_password_hash(value):
    text = str(value or '')
    return text.startswith(PASSWORD_HASH_PREFIXES)


def hash_password(value):
    return generate_password_hash(str(value or ''))


def verify_password_value(stored_password, provided_password):
    stored = str(stored_password or '')
    provided = str(provided_password or '')
    if not stored or not provided:
        return False
    if is_password_hash(stored):
        try:
            return check_password_hash(stored, provided)
        except Exception:
            return False
    return stored == provided


def upgrade_plain_password_if_needed(cur, username, stored_password, provided_password):
    if is_password_hash(stored_password):
        return False
    hashed = hash_password(provided_password)
    cur.execute(
        'UPDATE usuarios SET password = %s WHERE lower(username) = %s',
        (hashed, str(username or '').lower())
    )
    return True


def sanitize_user_for_client(username, name, level, profissional_id=None, preferences=None):
    normalized_preferences = normalize_user_preferences(preferences, level)
    effective_permissions = get_effective_user_permissions(level, normalized_preferences)
    user = {
        'username': username,
        'name': name or username,
        'level': normalize_level(level),
        'preferences': normalized_preferences,
        'effectivePermissions': effective_permissions,
        'permissions': effective_permissions
    }
    if profissional_id is not None:
        user['professionalId'] = profissional_id
        user['profissional_id'] = profissional_id
    return user


def normalize_appointment_status(value, default='agendado'):
    if value is None:
        return default
    text = str(value).strip()
    if not text:
        return default
    key = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
    key = key.strip().lower().replace('-', '_')
    key = ' '.join(key.replace('_', ' ').split())
    normalized = APPOINTMENT_STATUS_ALIASES.get(key) or APPOINTMENT_STATUS_ALIASES.get(key.replace(' ', '_'))
    if normalized:
        return normalized
    return text.strip().lower() if text.strip().lower() in VALID_APPOINTMENT_STATUSES else None


def normalize_lookup_key(value):
    text = str(value or '').strip()
    if not text:
        return ''
    key = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
    key = key.strip().lower().replace('-', ' ')
    return ' '.join(key.replace('_', ' ').split())


def normalize_waitlist_status(value, default='aguardando'):
    if value is None:
        return default
    key = normalize_lookup_key(value)
    if not key:
        return default
    normalized = WAITLIST_STATUS_ALIASES.get(key) or WAITLIST_STATUS_ALIASES.get(key.replace(' ', '_'))
    if normalized:
        return normalized
    return key.replace(' ', '_') if key.replace(' ', '_') in VALID_WAITLIST_STATUSES else None


def normalize_waitlist_priority(value, default='normal'):
    if value is None:
        return default
    key = normalize_lookup_key(value)
    if not key:
        return default
    normalized = WAITLIST_PRIORITY_ALIASES.get(key) or WAITLIST_PRIORITY_ALIASES.get(key.replace(' ', '_'))
    if normalized:
        return normalized
    return key.replace(' ', '_') if key.replace(' ', '_') in VALID_WAITLIST_PRIORITIES else None


def authenticate_credentials(username, password):
    username = (username or '').strip()
    password = password or ''
    if not username or not password:
        return None, (jsonify({'success': False, 'error': 'username and password required'}), 400)

    conn = None
    cur = None
    try:
        conn = get_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                'SELECT username, name, level, password, is_active, profissional_id FROM usuarios WHERE lower(username) = %s',
                (username.lower(),)
            )
            row = cur.fetchone()
            has_profissional_id = True
        except Exception:
            conn.rollback()
            cur.execute(
                'SELECT username, name, level, password, is_active FROM usuarios WHERE lower(username) = %s',
                (username.lower(),)
            )
            row = cur.fetchone()
            has_profissional_id = False

        if not row:
            return None, (jsonify({'success': False, 'error': 'Usuario nao encontrado'}), 403)

        db_username, db_name, db_level, db_password, db_active = row[:5]
        profissional_id = row[5] if has_profissional_id else None
        if db_active is False:
            return None, (jsonify({'success': False, 'error': 'Usuario inativo'}), 403)
        if not verify_password_value(db_password, password):
            return None, (jsonify({'success': False, 'error': 'Credenciais invalidas'}), 401)

        upgrade_plain_password_if_needed(cur, db_username, db_password, password)
        preferences = get_user_preferences(cur, db_username, db_level)
        conn.commit()

        user = sanitize_user_for_client(db_username, db_name, db_level, profissional_id, preferences)
        return user, None
    except Exception as e:
        print('Erro ao autenticar credenciais:', e)
        try:
            if conn:
                conn.rollback()
        except Exception:
            pass
        return None, (jsonify({'success': False, 'error': 'Erro interno ao verificar credenciais'}), 500)
    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass


def get_authenticated_user():
    sess_user = None
    try:
        sess_user = session.get('current_user')
    except Exception:
        sess_user = None

    if sess_user and isinstance(sess_user, dict) and sess_user.get('username'):
        username = str(sess_user.get('username') or '').strip()
        cache_key = ('session', username.lower())
        cached_user = get_auth_user_cache(cache_key)
        if cached_user:
            return cached_user, None

        conn = None
        cur = None
        try:
            conn = get_connection()
            cur = conn.cursor()
            try:
                cur.execute(
                    'SELECT username, name, level, is_active, profissional_id FROM usuarios WHERE lower(username) = %s',
                    (username.lower(),)
                )
                row = cur.fetchone()
                has_profissional_id = True
            except Exception:
                conn.rollback()
                cur.execute(
                    'SELECT username, name, level, is_active FROM usuarios WHERE lower(username) = %s',
                    (username.lower(),)
                )
                row = cur.fetchone()
                has_profissional_id = False
        except Exception as e:
            print('Erro ao buscar usuario autenticado via sessao:', e)
            return None, (jsonify({'success': False, 'error': 'Erro interno ao verificar sessao'}), 500)
        finally:
            try:
                if cur:
                    cur.close()
                if conn:
                    conn.close()
            except Exception:
                pass

        if not row:
            return None, (jsonify({'success': False, 'error': 'Usuario da sessao nao encontrado'}), 403)

        db_username, db_name, db_level, db_active = row[:4]
        profissional_id = row[4] if has_profissional_id else None
        if db_active is False:
            return None, (jsonify({'success': False, 'error': 'Usuario inativo'}), 403)

        conn = None
        cur = None
        try:
            conn = get_connection()
            cur = conn.cursor()
            preferences = get_user_preferences(cur, db_username, db_level)
            conn.commit()
        except Exception as e:
            print('Erro ao buscar preferencias do usuario autenticado:', e)
            preferences = normalize_user_preferences({}, db_level)
        finally:
            try:
                if cur:
                    cur.close()
                if conn:
                    conn.close()
            except Exception:
                pass

        authenticated = sanitize_user_for_client(db_username, db_name, db_level, profissional_id, preferences)
        set_auth_user_cache(cache_key, authenticated)
        return authenticated, None

    auth_header = request.headers.get('Authorization', '')
    if not auth_header or not auth_header.startswith('Bearer '):
        return None, (jsonify({'success': False, 'error': 'Autenticacao necessaria'}), 401)

    credentials = auth_header.replace('Bearer ', '')
    if ':' not in credentials:
        return None, (jsonify({'success': False, 'error': 'Formato de autenticacao invalido'}), 401)

    username, password = credentials.split(':', 1)
    user, auth_error = authenticate_credentials(username, password)
    if auth_error:
        return None, auth_error

    cache_key = ('header', str(user.get('username') or '').lower())
    set_auth_user_cache(cache_key, user)
    return user, None


def require_authenticated():
    _user, auth_error = get_authenticated_user()
    return auth_error


def require_roles(*allowed_roles):
    user, auth_error = get_authenticated_user()
    if auth_error:
        return auth_error
    allowed = {normalize_level(role) for role in allowed_roles}
    if normalize_level(user.get('level')) not in allowed:
        return jsonify({'success': False, 'error': 'Acesso negado'}), 403
    return None


# Helper to require admin-level credentials on sensitive endpoints
def require_admin():
    return require_roles('admin')


def require_editor_or_admin():
    return require_roles('admin', 'editor')


def ensure_agendamento_lock_columns(cur, appointment_cols):
    if 'cancelado_por_username' not in appointment_cols:
        cur.execute("ALTER TABLE agendamentos ADD COLUMN IF NOT EXISTS cancelado_por_username VARCHAR(255)")
        appointment_cols.add('cancelado_por_username')


AGENDAMENTO_RECURRENCE_SCHEMA_READY = False
AGENDAMENTO_RECURRENCE_SCHEMA_LOCK = threading.Lock()


def ensure_agendamento_recurrence_columns(cur, appointment_cols):
    global AGENDAMENTO_RECURRENCE_SCHEMA_READY

    recurrence_cols = {'recorrencia_grupo_id', 'recorrencia_indice', 'recorrencia_total'}
    if AGENDAMENTO_RECURRENCE_SCHEMA_READY:
        appointment_cols.update(recurrence_cols)
        return

    with AGENDAMENTO_RECURRENCE_SCHEMA_LOCK:
        if AGENDAMENTO_RECURRENCE_SCHEMA_READY:
            appointment_cols.update(recurrence_cols)
            return

    if 'recorrencia_grupo_id' not in appointment_cols:
        cur.execute("ALTER TABLE agendamentos ADD COLUMN IF NOT EXISTS recorrencia_grupo_id VARCHAR(80)")
        appointment_cols.add('recorrencia_grupo_id')
    if 'recorrencia_indice' not in appointment_cols:
        cur.execute("ALTER TABLE agendamentos ADD COLUMN IF NOT EXISTS recorrencia_indice INTEGER")
        appointment_cols.add('recorrencia_indice')
    if 'recorrencia_total' not in appointment_cols:
        cur.execute("ALTER TABLE agendamentos ADD COLUMN IF NOT EXISTS recorrencia_total INTEGER")
        appointment_cols.add('recorrencia_total')
    cur.execute(
        "CREATE INDEX IF NOT EXISTS idx_agendamentos_recorrencia_grupo "
        "ON agendamentos (recorrencia_grupo_id)"
    )
    cur.connection.commit()
    AGENDAMENTO_RECURRENCE_SCHEMA_READY = True


DEFAULT_SALAS = [
    ('AMARELA', '#fff200'),
    ('AZUL', '#00aeef'),
    ('AZUL BEBE', '#5bc0de'),
    ('AZUL TURQUESA', '#39d5c8'),
    ('BEGE', '#f5f5dc'),
    ('BRANCA', '#ffffff'),
    ('CINZA', '#808080'),
    ('COLORIDA - IS', '#1d1b8f'),
    ('COPA 6\u00ba', '#ffffcc'),
    ('COPA 7\u00ba', '#ffffcc'),
    ('DOURADA', '#f5d45a'),
    ('LARANJA', '#ffc000'),
    ('LIL\u00c1S', '#d9c2f0'),
    ('MARROM', '#8b5a2b'),
    ('NEON', '#ff7f27'),
    ('PRATEADA', '#c0c0c0'),
    ('PRETA', '#000000'),
    ('ROSA', '#ff99ff'),
    ('ROSA PINK', '#ff5bc8'),
    ('ROS\u00c9', '#f7d8c8'),
    ('ROXA', '#a259d9'),
    ('SALA DE BRINQUEDOS', '#8aa100'),
    ('SALA DE DISCUSS\u00c3O', '#ffffff'),
    ('Sala de Reuni\u00e3o', '#5b5b5b'),
    ('Sala Let\u00edcia', '#ffffff'),
    ('Sala do 3\u00baandar', '#ffffff'),
    ('SALA ONLINE 6\u00ba andar', '#ffffff'),
    ('Sala Online Neuro', '#ffffff'),
    ('Sala Online PsicoPM MT', '#ffffff'),
    ('Sala Online Reuni\u00e3o', '#808080'),
    ('Sala Reuni\u00e3o ABA', '#808080'),
    ('Sala Reuni\u00e3o Fono', '#808080'),
    ('Sala Reuni\u00e3o MT e PM', '#808080'),
    ('Sala Reuni\u00e3o Neuro', '#808080'),
    ('Sala Sup T.O', '#555555'),
    ('Sala Supervis\u00e3o Online', '#808080'),
    ('VERDE', '#30d830'),
    ('VERDE \u00c1GUA', '#00ffd5'),
    ('VERDE LIM\u00c3O', '#55ff33'),
    ('VERMELHA', '#ff1a1a'),
    ('VINHO', '#c00000')
]

REMOVED_SALAS = [
    'SALA DE CONVEN\u00c7\u00d5ES 17',
    'SALA DE CONVENCOES 17',
    'SALA DE CONVENCAO',
    'SALA DE CONVEN\u00c7\u00c3O'
]

SALAS_SCHEMA_READY = False
SALAS_SCHEMA_LOCK = threading.Lock()
AGENDAMENTO_LINKS_SCHEMA_READY = False
AGENDAMENTO_LINKS_SCHEMA_LOCK = threading.Lock()


def normalize_room_id(value):
    if value in (None, '', 'null', 'None'):
        return None
    try:
        return int(value)
    except Exception:
        return None


def ensure_salas_schema(cur, appointment_cols=None):
    global SALAS_SCHEMA_READY

    if SALAS_SCHEMA_READY:
        if appointment_cols is None:
            appointment_cols = get_table_columns_cached(cur, 'agendamentos')
        appointment_cols.add('sala_id')
        return appointment_cols

    with SALAS_SCHEMA_LOCK:
        if SALAS_SCHEMA_READY:
            if appointment_cols is None:
                appointment_cols = get_table_columns_cached(cur, 'agendamentos')
            appointment_cols.add('sala_id')
            return appointment_cols

        cur.execute("""
            CREATE TABLE IF NOT EXISTS salas (
                id SERIAL PRIMARY KEY,
                nome VARCHAR(120) NOT NULL,
                cor VARCHAR(20),
                ativo BOOLEAN DEFAULT TRUE,
                criado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("ALTER TABLE salas ADD COLUMN IF NOT EXISTS cor VARCHAR(20)")
        cur.execute("ALTER TABLE salas ADD COLUMN IF NOT EXISTS ativo BOOLEAN DEFAULT TRUE")
        cur.execute("ALTER TABLE salas ADD COLUMN IF NOT EXISTS criado_em TIMESTAMP DEFAULT NOW()")

        for nome, cor in DEFAULT_SALAS:
            cur.execute(
                """
                INSERT INTO salas (nome, cor, ativo)
                SELECT %s, %s, TRUE
                WHERE NOT EXISTS (
                    SELECT 1 FROM salas WHERE lower(nome) = lower(%s)
                )
                """,
                (nome, cor, nome)
            )

        for nome in REMOVED_SALAS:
            cur.execute("UPDATE salas SET ativo = FALSE WHERE lower(nome) = lower(%s)", (nome,))

        cur.execute("CREATE INDEX IF NOT EXISTS idx_salas_nome_lower ON salas (lower(nome))")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_salas_ativo_nome ON salas (ativo, nome)")

        if appointment_cols is None:
            appointment_cols = get_table_columns_cached(cur, 'agendamentos')
        if 'sala_id' not in appointment_cols:
            cur.execute("ALTER TABLE agendamentos ADD COLUMN IF NOT EXISTS sala_id INTEGER")
            appointment_cols.add('sala_id')

        cur.connection.commit()
        SALAS_SCHEMA_READY = True
        return appointment_cols


def is_truthy_env(value):
    return str(value or '').strip().lower() in ('1', 'true', 'yes', 'sim', 'on')


def ensure_agendamento_link_schema(cur, appointment_cols=None):
    global AGENDAMENTO_LINKS_SCHEMA_READY

    appointment_cols = ensure_salas_schema(cur, appointment_cols)

    if AGENDAMENTO_LINKS_SCHEMA_READY and {'profissional_id', 'paciente_id'}.issubset(appointment_cols):
        return appointment_cols

    with AGENDAMENTO_LINKS_SCHEMA_LOCK:
        if not {'profissional_id', 'paciente_id'}.issubset(appointment_cols):
            if 'profissional_id' not in appointment_cols:
                cur.execute("ALTER TABLE agendamentos ADD COLUMN IF NOT EXISTS profissional_id INTEGER")
                appointment_cols.add('profissional_id')
            if 'paciente_id' not in appointment_cols:
                cur.execute("ALTER TABLE agendamentos ADD COLUMN IF NOT EXISTS paciente_id INTEGER")
                appointment_cols.add('paciente_id')

        cur.execute("CREATE INDEX IF NOT EXISTS idx_agendamentos_profissional_id_data_hora ON agendamentos (profissional_id, data, hora_inicio)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_agendamentos_paciente_id_data_hora ON agendamentos (paciente_id, data, hora_inicio)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_agendamentos_sala_id_data_hora ON agendamentos (sala_id, data, hora_inicio)")

        AGENDAMENTO_LINKS_SCHEMA_READY = True
        return appointment_cols


def add_constraint_if_missing(cur, constraint_name, statement):
    cur.execute(
        """
        SELECT 1
        FROM pg_constraint
        WHERE conname = %s
        LIMIT 1
        """,
        (constraint_name,)
    )
    if not cur.fetchone():
        cur.execute(statement)


def migrate_existing_agendamento_links(cur, delete_invalid=False, enforce_constraints=False):
    try:
        appointment_cols = ensure_agendamento_link_schema(cur, get_table_columns_cached(cur, 'agendamentos'))
        ensure_agendamento_auditoria_table(cur)

        cur.execute(
            """
            UPDATE agendamentos a
            SET profissional_id = p.id
            FROM profissionais p
            WHERE a.profissional_id IS NULL
              AND (
                    a.profissional::text = p.id::text
                    OR lower(trim(a.profissional::text)) = lower(trim(p.nome::text))
              )
            """
        )
        cur.execute(
            """
            UPDATE agendamentos a
            SET paciente_id = p.id
            FROM pacientes p
            WHERE a.paciente_id IS NULL
              AND (
                    a.paciente::text = p.id::text
                    OR lower(trim(a.paciente::text)) = lower(trim(p.nome::text))
              )
            """
        )

        if delete_invalid:
            cur.execute(
                """
                CREATE TEMP TABLE invalid_agendamentos_link AS
                SELECT a.id
                FROM agendamentos a
                LEFT JOIN profissionais pr ON pr.id = a.profissional_id
                LEFT JOIN pacientes pa ON pa.id = a.paciente_id
                LEFT JOIN salas s ON s.id = a.sala_id
                WHERE pr.id IS NULL OR pa.id IS NULL OR s.id IS NULL
                """
            )
            cur.execute(
                """
                DELETE FROM agendamento_auditoria
                WHERE agendamento_id IN (SELECT id FROM invalid_agendamentos_link)
                """
            )
            cur.execute(
                """
                DELETE FROM agendamentos
                WHERE id IN (SELECT id FROM invalid_agendamentos_link)
                """
            )
            cur.execute("DROP TABLE IF EXISTS invalid_agendamentos_link")

        cur.execute(
            """
            UPDATE agendamentos
            SET profissional = profissional_id::text
            WHERE profissional_id IS NOT NULL
              AND profissional::text IS DISTINCT FROM profissional_id::text
            """
        )
        cur.execute(
            """
            UPDATE agendamentos a
            SET paciente = p.nome
            FROM pacientes p
            WHERE p.id = a.paciente_id
              AND a.paciente IS DISTINCT FROM p.nome
            """
        )

        if enforce_constraints:
            cur.execute(
                """
                SELECT COUNT(*)
                FROM agendamentos a
                LEFT JOIN profissionais pr ON pr.id = a.profissional_id
                LEFT JOIN pacientes pa ON pa.id = a.paciente_id
                LEFT JOIN salas s ON s.id = a.sala_id
                WHERE a.profissional_id IS NULL
                   OR a.paciente_id IS NULL
                   OR a.sala_id IS NULL
                   OR pr.id IS NULL
                   OR pa.id IS NULL
                   OR s.id IS NULL
                """
            )
            invalid_count = cur.fetchone()[0] or 0
            if invalid_count == 0:
                cur.execute("ALTER TABLE agendamentos ALTER COLUMN profissional_id SET NOT NULL")
                cur.execute("ALTER TABLE agendamentos ALTER COLUMN paciente_id SET NOT NULL")
                cur.execute("ALTER TABLE agendamentos ALTER COLUMN sala_id SET NOT NULL")
                add_constraint_if_missing(
                    cur,
                    'fk_agendamentos_profissional_id',
                    """
                    ALTER TABLE agendamentos
                    ADD CONSTRAINT fk_agendamentos_profissional_id
                    FOREIGN KEY (profissional_id) REFERENCES profissionais(id)
                    """
                )
                add_constraint_if_missing(
                    cur,
                    'fk_agendamentos_paciente_id',
                    """
                    ALTER TABLE agendamentos
                    ADD CONSTRAINT fk_agendamentos_paciente_id
                    FOREIGN KEY (paciente_id) REFERENCES pacientes(id)
                    """
                )
                add_constraint_if_missing(
                    cur,
                    'fk_agendamentos_sala_id',
                    """
                    ALTER TABLE agendamentos
                    ADD CONSTRAINT fk_agendamentos_sala_id
                    FOREIGN KEY (sala_id) REFERENCES salas(id)
                    """
                )
            else:
                print(f'Vinculo obrigatorio de agendamentos ainda nao aplicado: {invalid_count} registro(s) invalido(s)')

        appointment_cols.update({'profissional_id', 'paciente_id', 'sala_id'})
        return appointment_cols
    except Exception as e:
        print('Erro ao migrar vinculos de agendamentos:', e)
        raise


def normalize_optional_int(value):
    if value in (None, '', 'null', 'None'):
        return None
    try:
        return int(value)
    except Exception:
        return None


def resolve_professional_reference(cur, value=None, profissional_id=None, require_active=True):
    raw_id = normalize_optional_int(profissional_id)
    raw_text = str(value or profissional_id or '').strip()
    professional_cols = get_table_columns_cached(cur, 'profissionais')
    active_select = 'ativo' if 'ativo' in professional_cols else 'TRUE AS ativo'
    if raw_id is None and raw_text:
        raw_id = normalize_optional_int(raw_text)
    if raw_id is not None:
        cur.execute(f"SELECT id, nome, {active_select} FROM profissionais WHERE id = %s LIMIT 1", (raw_id,))
    else:
        if not raw_text:
            return None, 'Selecione um profissional cadastrado.'
        cur.execute(
            f"""
            SELECT id, nome, {active_select}
            FROM profissionais
            WHERE lower(trim(nome::text)) = lower(trim(%s))
            ORDER BY CASE WHEN {('ativo' if 'ativo' in professional_cols else 'TRUE')} IS DISTINCT FROM FALSE THEN 0 ELSE 1 END, id
            LIMIT 1
            """,
            (raw_text,)
        )
    row = cur.fetchone()
    if not row:
        return None, 'Profissional informado nao existe no cadastro.'
    if require_active and row[2] is False:
        return None, 'Profissional informado esta inativo.'
    return {'id': row[0], 'nome': row[1], 'ativo': row[2]}, None


def resolve_patient_reference(cur, value=None, paciente_id=None, require_active=True):
    raw_id = normalize_optional_int(paciente_id)
    raw_text = str(value or paciente_id or '').strip()
    patient_cols = get_table_columns_cached(cur, 'pacientes')
    active_select = 'ativo' if 'ativo' in patient_cols else 'TRUE AS ativo'
    if raw_id is None and raw_text:
        raw_id = normalize_optional_int(raw_text)
    if raw_id is not None:
        cur.execute(f"SELECT id, nome, {active_select} FROM pacientes WHERE id = %s LIMIT 1", (raw_id,))
    else:
        if not raw_text:
            return None, 'Selecione um paciente cadastrado.'
        cur.execute(
            f"""
            SELECT id, nome, {active_select}
            FROM pacientes
            WHERE lower(trim(nome::text)) = lower(trim(%s))
            ORDER BY CASE WHEN {('ativo' if 'ativo' in patient_cols else 'TRUE')} IS DISTINCT FROM FALSE THEN 0 ELSE 1 END, id
            LIMIT 1
            """,
            (raw_text,)
        )
    row = cur.fetchone()
    if not row:
        return None, 'Paciente informado nao existe no cadastro.'
    if require_active and row[2] is False:
        return None, 'Paciente informado esta inativo.'
    return {'id': row[0], 'nome': row[1], 'ativo': row[2]}, None


def resolve_room_reference(cur, sala_id, require_active=True):
    room_id = normalize_room_id(sala_id)
    if room_id is None:
        return None, 'Selecione uma sala cadastrada.'
    cur.execute("SELECT id, nome, ativo FROM salas WHERE id = %s LIMIT 1", (room_id,))
    row = cur.fetchone()
    if not row:
        return None, 'Sala informada nao existe no cadastro.'
    if require_active and row[2] is False:
        return None, 'Sala informada esta inativa.'
    return {'id': row[0], 'nome': row[1], 'ativo': row[2]}, None


def values_same_date(first_value, second_value):
    first_date = normalize_date_for_db(first_value)
    second_date = normalize_date_for_db(second_value)
    if not first_date or not second_date:
        return False
    return str(first_date) == str(second_date)


def format_conflict_time(value):
    if value in (None, ''):
        return ''
    if hasattr(value, 'strftime'):
        return value.strftime('%H:%M')
    parts = str(value).strip().split(':')
    if len(parts) >= 2:
        return f'{parts[0]}:{parts[1]}'
    return str(value)


def build_patient_room_conflict_error(conflict):
    if not conflict:
        return 'Paciente ja possui agendamento no mesmo horario em outra sala.'
    appointment_id = conflict.get('id')
    room_label = conflict.get('sala_nome') or f"Sala {conflict.get('sala_id') or 'nao informada'}"
    start_time = format_conflict_time(conflict.get('hora_inicio'))
    end_time = format_conflict_time(conflict.get('hora_fim')) or start_time
    patient_name = conflict.get('paciente') or 'Paciente'
    id_label = f"ID {appointment_id}, " if appointment_id else ''
    return (
        f"{patient_name} ja possui agendamento no mesmo horario em outra sala "
        f"({id_label}{room_label}, {start_time}-{end_time})."
    )


def find_pending_patient_room_conflict(pending_items, paciente_id, sala_id, data_field, hora_inicio, hora_fim):
    patient_id = normalize_optional_int(paciente_id)
    room_id = normalize_room_id(sala_id)
    end_time = hora_fim or hora_inicio
    if patient_id is None or room_id is None or not data_field or not hora_inicio:
        return None
    if time_to_minutes(hora_inicio) is None or time_to_minutes(end_time) is None:
        return None

    for item in pending_items or []:
        item_patient_id = normalize_optional_int(item.get('paciente_id'))
        item_room_id = normalize_room_id(item.get('sala_id'))
        if item_patient_id != patient_id:
            continue
        if item_room_id == room_id:
            continue
        if not values_same_date(item.get('data'), data_field):
            continue
        if times_overlap(hora_inicio, end_time, item.get('hora_inicio'), item.get('hora_fim') or item.get('hora_inicio')):
            return {
                'id': item.get('id'),
                'paciente': item.get('paciente'),
                'sala_id': item.get('sala_id'),
                'sala_nome': item.get('sala_nome'),
                'data': item.get('data'),
                'hora_inicio': item.get('hora_inicio'),
                'hora_fim': item.get('hora_fim')
            }
    return None


def find_patient_room_conflict(
    cur,
    paciente_id,
    sala_id,
    data_field,
    hora_inicio,
    hora_fim=None,
    exclude_agendamento_id=None,
    appointment_cols=None
):
    patient_id = normalize_optional_int(paciente_id)
    room_id = normalize_room_id(sala_id)
    end_time = hora_fim or hora_inicio
    if patient_id is None or room_id is None or not data_field or not hora_inicio:
        return None
    if time_to_minutes(hora_inicio) is None or time_to_minutes(end_time) is None:
        return None

    if appointment_cols is None:
        appointment_cols = get_table_columns_cached(cur, 'agendamentos')

    params = [patient_id, data_field, end_time, hora_inicio, room_id]
    status_filter = ''
    if 'status' in appointment_cols:
        ignored_statuses = sorted(PATIENT_ROOM_CONFLICT_IGNORED_STATUSES)
        placeholders = ','.join(['%s'] * len(ignored_statuses))
        status_filter = f"AND lower(COALESCE(a.status::text, 'agendado')) NOT IN ({placeholders})"
        params.extend(ignored_statuses)

    exclude_filter = ''
    if exclude_agendamento_id:
        exclude_filter = 'AND a.id <> %s'
        params.append(exclude_agendamento_id)

    cur.execute(
        f"""
        SELECT a.id, a.paciente, a.profissional, a.data, a.hora_inicio, a.hora_fim,
               a.sala_id, s.nome AS sala_nome
        FROM agendamentos a
        LEFT JOIN salas s ON s.id = a.sala_id
        WHERE a.paciente_id = %s
          AND a.data::date = %s::date
          AND NULLIF(a.hora_inicio::text, '')::time < %s::time
          AND COALESCE(NULLIF(a.hora_fim::text, '')::time, NULLIF(a.hora_inicio::text, '')::time) > %s::time
          AND COALESCE(a.sala_id, -1) <> %s
          {status_filter}
          {exclude_filter}
        ORDER BY a.hora_inicio ASC, a.id ASC
        LIMIT 1
        """,
        tuple(params)
    )
    row = cur.fetchone()
    if not row:
        return None
    return {
        'id': row[0],
        'paciente': row[1],
        'profissional': row[2],
        'data': row[3],
        'hora_inicio': row[4],
        'hora_fim': row[5],
        'sala_id': row[6],
        'sala_nome': row[7]
    }


def get_room_name_by_id(cur, sala_id):
    if sala_id is None:
        return None
    cur.execute("SELECT nome FROM salas WHERE id = %s LIMIT 1", (sala_id,))
    row = cur.fetchone()
    return row[0] if row else None


def normalize_recurrence_scope(value, default='single'):
    raw = str(value or '').strip().lower()
    aliases = {
        'single': 'single',
        'only': 'single',
        'este': 'single',
        'so_este': 'single',
        'somente_este': 'single',
        'weekday': 'weekday',
        'same_weekday': 'weekday',
        'dia_semana': 'weekday',
        'mesmo_dia_semana': 'weekday',
        'all': 'all',
        'series': 'all',
        'serie': 'all',
        'todos': 'all',
        'todas': 'all'
    }
    return aliases.get(raw, default)


TABLE_COLUMNS_CACHE = {}
TABLE_COLUMNS_CACHE_LOCK = threading.Lock()
AGENDAMENTOS_LIST_CACHE = {}
AGENDAMENTOS_LIST_CACHE_LOCK = threading.Lock()
AGENDAMENTOS_LIST_CACHE_TTL_SECONDS = float(os.environ.get('AGENDAMENTOS_LIST_CACHE_TTL_SECONDS', '20'))
AUTH_USER_CACHE = {}
AUTH_USER_CACHE_LOCK = threading.Lock()
AUTH_USER_CACHE_TTL_SECONDS = float(os.environ.get('AUTH_USER_CACHE_TTL_SECONDS', '30'))
REMARQUE_SCHEMA_READY = False
REMARQUE_SCHEMA_LOCK = threading.Lock()
REMARQUE_LIST_CACHE = {}
REMARQUE_LIST_CACHE_LOCK = threading.Lock()
REMARQUE_LIST_CACHE_TTL_SECONDS = float(os.environ.get('REMARQUE_LIST_CACHE_TTL_SECONDS', '15'))
APP_CONFIG_SCHEMA_READY = False
APP_CONFIG_SCHEMA_LOCK = threading.Lock()
APP_CONFIG_CACHE = {}
APP_CONFIG_CACHE_LOCK = threading.Lock()
APP_CONFIG_CACHE_TTL_SECONDS = float(os.environ.get('APP_CONFIG_CACHE_TTL_SECONDS', '30'))
WAITLIST_SCHEMA_READY = False
WAITLIST_SCHEMA_LOCK = threading.Lock()
USER_PREFERENCES_SCHEMA_READY = False
USER_PREFERENCES_SCHEMA_LOCK = threading.Lock()

USER_PERMISSION_KEYS = (
    'canView',
    'canViewPatients',
    'canCreate',
    'canCreateProfessional',
    'canCreatePatient',
    'canEdit',
    'canEditProfessionals',
    'canEditPatients',
    'canDelete',
    'canExport',
    'canExportReport',
    'canImport',
    'canBulkEdit',
    'canBulkCancel',
    'canManageProfessionals',
    'canManageUsers',
    'canViewAudit'
)

USER_PERMISSION_DEFAULTS = {
    'admin': {
        'canView': True,
        'canViewPatients': True,
        'canCreate': True,
        'canCreateProfessional': True,
        'canCreatePatient': True,
        'canEdit': True,
        'canEditProfessionals': True,
        'canEditPatients': True,
        'canDelete': True,
        'canExport': True,
        'canExportReport': True,
        'canImport': True,
        'canBulkEdit': True,
        'canBulkCancel': True,
        'canManageProfessionals': True,
        'canManageUsers': True,
        'canViewAudit': True
    },
    'editor': {
        'canView': True,
        'canViewPatients': True,
        'canCreate': True,
        'canCreateProfessional': False,
        'canCreatePatient': False,
        'canEdit': True,
        'canEditProfessionals': False,
        'canEditPatients': False,
        'canDelete': False,
        'canExport': True,
        'canExportReport': True,
        'canImport': True,
        'canBulkEdit': True,
        'canBulkCancel': False,
        'canManageProfessionals': True,
        'canManageUsers': False,
        'canViewAudit': False
    },
    'viewer': {
        'canView': True,
        'canViewPatients': False,
        'canCreate': False,
        'canCreateProfessional': False,
        'canCreatePatient': False,
        'canEdit': False,
        'canEditProfessionals': False,
        'canEditPatients': False,
        'canDelete': False,
        'canExport': True,
        'canExportReport': False,
        'canImport': False,
        'canBulkEdit': False,
        'canBulkCancel': False,
        'canManageProfessionals': False,
        'canManageUsers': False,
        'canViewAudit': False
    }
}

DEFAULT_ACTION_CENTER_FAVORITES = ('weekly', 'daily', 'schedule', 'waitlist')


def get_table_columns_cached(cur, table_name, refresh=False):
    key = str(table_name).lower()
    with TABLE_COLUMNS_CACHE_LOCK:
        if not refresh and key in TABLE_COLUMNS_CACHE:
            return TABLE_COLUMNS_CACHE[key]

    cur.execute(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_name = %s
          AND table_schema NOT IN ('information_schema', 'pg_catalog')
        """,
        (table_name,)
    )
    columns = {row[0].lower() for row in cur.fetchall()}
    with TABLE_COLUMNS_CACHE_LOCK:
        TABLE_COLUMNS_CACHE[key] = columns
    return columns


def ensure_table_updated_timestamp(cur, table_name, columns=None):
    allowed_tables = {'agendamentos', 'profissionais', 'pacientes'}
    if table_name not in allowed_tables:
        raise ValueError(f'Tabela sem suporte para atualizado_em: {table_name}')

    columns = columns or get_table_columns_cached(cur, table_name)
    if 'atualizado_em' not in columns:
        cur.execute(f"ALTER TABLE {table_name} ADD COLUMN IF NOT EXISTS atualizado_em TIMESTAMP DEFAULT NOW()")
        columns.add('atualizado_em')
        with TABLE_COLUMNS_CACHE_LOCK:
            TABLE_COLUMNS_CACHE[str(table_name).lower()] = columns
    return columns


def get_agendamentos_list_cache(cache_key):
    if AGENDAMENTOS_LIST_CACHE_TTL_SECONDS <= 0:
        return None
    now = monotonic()
    with AGENDAMENTOS_LIST_CACHE_LOCK:
        cached = AGENDAMENTOS_LIST_CACHE.get(cache_key)
        if not cached:
            return None
        cached_at, payload = cached
        if now - cached_at <= AGENDAMENTOS_LIST_CACHE_TTL_SECONDS:
            return payload
        AGENDAMENTOS_LIST_CACHE.pop(cache_key, None)
    return None


def set_agendamentos_list_cache(cache_key, payload):
    if AGENDAMENTOS_LIST_CACHE_TTL_SECONDS <= 0:
        return
    with AGENDAMENTOS_LIST_CACHE_LOCK:
        AGENDAMENTOS_LIST_CACHE[cache_key] = (monotonic(), payload)


def invalidate_agendamentos_list_cache():
    with AGENDAMENTOS_LIST_CACHE_LOCK:
        AGENDAMENTOS_LIST_CACHE.clear()


def get_ttl_cache(cache, lock, key, ttl_seconds):
    if ttl_seconds <= 0:
        return None
    now = monotonic()
    with lock:
        cached = cache.get(key)
        if not cached:
            return None
        cached_at, payload = cached
        if now - cached_at <= ttl_seconds:
            return payload
        cache.pop(key, None)
    return None


def set_ttl_cache(cache, lock, key, payload, ttl_seconds):
    if ttl_seconds <= 0:
        return
    with lock:
        cache[key] = (monotonic(), payload)


def get_auth_user_cache(cache_key):
    cached = get_ttl_cache(AUTH_USER_CACHE, AUTH_USER_CACHE_LOCK, cache_key, AUTH_USER_CACHE_TTL_SECONDS)
    return dict(cached) if cached else None


def set_auth_user_cache(cache_key, user):
    set_ttl_cache(AUTH_USER_CACHE, AUTH_USER_CACHE_LOCK, cache_key, dict(user), AUTH_USER_CACHE_TTL_SECONDS)


def invalidate_auth_user_cache(username=None):
    username_key = str(username or '').strip().lower()
    with AUTH_USER_CACHE_LOCK:
        if not username_key:
            AUTH_USER_CACHE.clear()
            return
        for cache_key in list(AUTH_USER_CACHE.keys()):
            if isinstance(cache_key, tuple) and len(cache_key) > 1 and cache_key[1] == username_key:
                AUTH_USER_CACHE.pop(cache_key, None)


def get_remarque_list_cache(cache_key):
    cached = get_ttl_cache(REMARQUE_LIST_CACHE, REMARQUE_LIST_CACHE_LOCK, cache_key, REMARQUE_LIST_CACHE_TTL_SECONDS)
    return dict(cached) if cached else None


def set_remarque_list_cache(cache_key, payload):
    set_ttl_cache(REMARQUE_LIST_CACHE, REMARQUE_LIST_CACHE_LOCK, cache_key, dict(payload), REMARQUE_LIST_CACHE_TTL_SECONDS)


def invalidate_remarque_list_cache():
    with REMARQUE_LIST_CACHE_LOCK:
        REMARQUE_LIST_CACHE.clear()


def invalidate_app_config_cache():
    with APP_CONFIG_CACHE_LOCK:
        APP_CONFIG_CACHE.clear()


def clear_runtime_caches():
    cleared = {}
    with TABLE_COLUMNS_CACHE_LOCK:
        cleared['table_columns'] = len(TABLE_COLUMNS_CACHE)
        TABLE_COLUMNS_CACHE.clear()
    with AGENDAMENTOS_LIST_CACHE_LOCK:
        cleared['agendamentos'] = len(AGENDAMENTOS_LIST_CACHE)
        AGENDAMENTOS_LIST_CACHE.clear()
    with AUTH_USER_CACHE_LOCK:
        cleared['auth_users'] = len(AUTH_USER_CACHE)
        AUTH_USER_CACHE.clear()
    with REMARQUE_LIST_CACHE_LOCK:
        cleared['remarques'] = len(REMARQUE_LIST_CACHE)
        REMARQUE_LIST_CACHE.clear()
    with APP_CONFIG_CACHE_LOCK:
        cleared['configuracoes'] = len(APP_CONFIG_CACHE)
        APP_CONFIG_CACHE.clear()
    return cleared


def ensure_user_preferences_table(cur):
    global USER_PREFERENCES_SCHEMA_READY
    if USER_PREFERENCES_SCHEMA_READY:
        return
    with USER_PREFERENCES_SCHEMA_LOCK:
        if USER_PREFERENCES_SCHEMA_READY:
            return
        cur.execute("""
            CREATE TABLE IF NOT EXISTS usuario_preferencias (
                username VARCHAR(255) PRIMARY KEY,
                preferencias JSONB NOT NULL DEFAULT '{}'::jsonb,
                atualizado_por VARCHAR(255),
                atualizado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("ALTER TABLE usuario_preferencias ADD COLUMN IF NOT EXISTS preferencias JSONB NOT NULL DEFAULT '{}'::jsonb")
        cur.execute("ALTER TABLE usuario_preferencias ADD COLUMN IF NOT EXISTS atualizado_por VARCHAR(255)")
        cur.execute("ALTER TABLE usuario_preferencias ADD COLUMN IF NOT EXISTS atualizado_em TIMESTAMP DEFAULT NOW()")
        cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_usuario_preferencias_username ON usuario_preferencias (username)")
        USER_PREFERENCES_SCHEMA_READY = True


def get_base_user_permissions(level):
    normalized_level = normalize_level(level)
    return dict(USER_PERMISSION_DEFAULTS.get(normalized_level) or USER_PERMISSION_DEFAULTS['viewer'])


def normalize_user_preferences(preferences=None, level=None):
    if isinstance(preferences, str):
        try:
            preferences = json.loads(preferences)
        except Exception:
            preferences = {}
    if not isinstance(preferences, dict):
        preferences = {}

    raw_permissions = preferences.get('permissions')
    if raw_permissions is None:
        raw_permissions = preferences.get('permissoes')
    if not isinstance(raw_permissions, dict):
        raw_permissions = {}

    base_permissions = get_base_user_permissions(level)
    normalized_permissions = {}
    for key in USER_PERMISSION_KEYS:
        base_value = bool(base_permissions.get(key, False))
        raw_value = raw_permissions[key] if key in raw_permissions else base_value
        normalized_permissions[key] = base_value and bool(raw_value)

    ui_payload = preferences.get('ui')
    if ui_payload is None:
        ui_payload = preferences.get('interface')
    if not isinstance(ui_payload, dict):
        ui_payload = {}

    action_center_payload = ui_payload.get('actionCenter')
    if action_center_payload is None:
        action_center_payload = ui_payload.get('action_center')
    if not isinstance(action_center_payload, dict):
        action_center_payload = {}

    if 'favorites' in action_center_payload:
        raw_favorites = action_center_payload.get('favorites')
    elif 'favoritos' in action_center_payload:
        raw_favorites = action_center_payload.get('favoritos')
    else:
        raw_favorites = DEFAULT_ACTION_CENTER_FAVORITES
    if not isinstance(raw_favorites, list):
        raw_favorites = DEFAULT_ACTION_CENTER_FAVORITES

    favorites = []
    for item in raw_favorites:
        text = str(item or '').strip()
        if text and text not in favorites:
            favorites.append(text)

    return {
        'permissions': normalized_permissions,
        'ui': {
            'actionCenter': {
                'favorites': favorites[:24]
            }
        }
    }


def get_effective_user_permissions(level, preferences=None):
    normalized = normalize_user_preferences(preferences, level)
    return dict(normalized.get('permissions') or {})


def user_has_effective_permission(user, permission_key):
    if not user or not permission_key:
        return False
    permissions_payload = user.get('effectivePermissions') or user.get('permissions')
    if isinstance(permissions_payload, dict) and permission_key in permissions_payload:
        return bool(permissions_payload.get(permission_key))
    return bool(get_base_user_permissions(user.get('level')).get(permission_key, False))


def require_admin_permission(permission_key=None):
    user, auth_error = get_authenticated_user()
    if auth_error:
        return None, auth_error
    if normalize_level(user.get('level')) != 'admin':
        return None, (jsonify({'success': False, 'error': 'Acesso negado'}), 403)
    if permission_key and not user_has_effective_permission(user, permission_key):
        return None, (jsonify({'success': False, 'error': 'Permissao personalizada negada'}), 403)
    return user, None


def get_user_preferences(cur, username, level=None):
    ensure_user_preferences_table(cur)
    username_key = str(username or '').strip().lower()
    if not username_key:
        return normalize_user_preferences({}, level)
    cur.execute(
        "SELECT preferencias FROM usuario_preferencias WHERE lower(username) = %s LIMIT 1",
        (username_key,)
    )
    row = cur.fetchone()
    return normalize_user_preferences(row[0] if row else {}, level)


def set_user_preferences(cur, username, preferences, updated_by=None, level=None):
    ensure_user_preferences_table(cur)
    username_key = str(username or '').strip().lower()
    normalized = normalize_user_preferences(preferences, level)
    cur.execute("""
        INSERT INTO usuario_preferencias (username, preferencias, atualizado_por, atualizado_em)
        VALUES (%s, %s, %s, NOW())
        ON CONFLICT (username) DO UPDATE
        SET preferencias = EXCLUDED.preferencias,
            atualizado_por = EXCLUDED.atualizado_por,
            atualizado_em = NOW()
    """, (username_key, Json(normalized), updated_by))
    invalidate_auth_user_cache(username_key)
    return normalized


def ensure_app_config_table(cur):
    global APP_CONFIG_SCHEMA_READY
    if APP_CONFIG_SCHEMA_READY:
        return
    with APP_CONFIG_SCHEMA_LOCK:
        if APP_CONFIG_SCHEMA_READY:
            return
        cur.execute("""
            CREATE TABLE IF NOT EXISTS sistema_configuracoes (
                chave VARCHAR(120) PRIMARY KEY,
                valor TEXT,
                atualizado_por VARCHAR(255),
                atualizado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("ALTER TABLE sistema_configuracoes ADD COLUMN IF NOT EXISTS atualizado_por VARCHAR(255)")
        cur.execute("ALTER TABLE sistema_configuracoes ADD COLUMN IF NOT EXISTS atualizado_em TIMESTAMP DEFAULT NOW()")
        APP_CONFIG_SCHEMA_READY = True


def get_app_config_value(cur, key, default_value=None, use_cache=True):
    if use_cache:
        cached = get_ttl_cache(APP_CONFIG_CACHE, APP_CONFIG_CACHE_LOCK, key, APP_CONFIG_CACHE_TTL_SECONDS)
        if cached is not None:
            return cached
    ensure_app_config_table(cur)
    cur.execute("SELECT valor FROM sistema_configuracoes WHERE chave = %s LIMIT 1", (key,))
    row = cur.fetchone()
    value = row[0] if row else default_value
    set_ttl_cache(APP_CONFIG_CACHE, APP_CONFIG_CACHE_LOCK, key, value, APP_CONFIG_CACHE_TTL_SECONDS)
    return value


def set_app_config_value(cur, key, value, updated_by=None):
    ensure_app_config_table(cur)
    cur.execute("""
        INSERT INTO sistema_configuracoes (chave, valor, atualizado_por, atualizado_em)
        VALUES (%s, %s, %s, NOW())
        ON CONFLICT (chave) DO UPDATE
        SET valor = EXCLUDED.valor,
            atualizado_por = EXCLUDED.atualizado_por,
            atualizado_em = NOW()
    """, (key, value, updated_by))
    invalidate_app_config_cache()


def parse_bool_config(value, default=True):
    if value is None:
        return default
    text = str(value).strip().lower()
    if text in ('1', 'true', 'yes', 'sim', 's', 'on', 'ativo', 'enabled'):
        return True
    if text in ('0', 'false', 'no', 'nao', 'não', 'n', 'off', 'inativo', 'disabled'):
        return False
    return default


def get_remarque_requests_enabled(cur):
    return parse_bool_config(get_app_config_value(cur, 'remarque_solicitacoes_ativas', 'true', use_cache=False), True)


def ensure_professional_extra_columns(cur, professional_cols):
    if 'preferencia' not in professional_cols:
        cur.execute("ALTER TABLE profissionais ADD COLUMN IF NOT EXISTS preferencia TEXT")
        professional_cols.add('preferencia')
    if 'contato_emergencia' not in professional_cols:
        cur.execute("ALTER TABLE profissionais ADD COLUMN IF NOT EXISTS contato_emergencia VARCHAR(255)")
        professional_cols.add('contato_emergencia')


def ensure_remarque_table(cur):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS remarque_solicitacoes (
            id SERIAL PRIMARY KEY,
            agendamento_id INTEGER NOT NULL,
            profissional_id VARCHAR(255),
            original_data DATE,
            original_hora_inicio VARCHAR(10),
            original_hora_fim VARCHAR(10),
            nova_data DATE NOT NULL,
            nova_hora_inicio VARCHAR(10) NOT NULL,
            nova_hora_fim VARCHAR(10) NOT NULL,
            inverter_horarios BOOLEAN DEFAULT FALSE,
            conflito_agendamento_id INTEGER,
            observacao TEXT,
            status VARCHAR(30) DEFAULT 'pendente',
            solicitado_por VARCHAR(255),
            solicitado_por_username VARCHAR(255),
            solicitado_em TIMESTAMP DEFAULT NOW(),
            autorizado_por VARCHAR(255),
            autorizado_em TIMESTAMP,
            rejeitado_por VARCHAR(255),
            rejeitado_em TIMESTAMP
        )
    """)
    cur.execute("ALTER TABLE remarque_solicitacoes ADD COLUMN IF NOT EXISTS conflito_nova_data DATE")
    cur.execute("ALTER TABLE remarque_solicitacoes ADD COLUMN IF NOT EXISTS conflito_nova_hora_inicio VARCHAR(10)")
    cur.execute("ALTER TABLE remarque_solicitacoes ADD COLUMN IF NOT EXISTS conflito_nova_hora_fim VARCHAR(10)")
    cur.execute("ALTER TABLE remarque_solicitacoes ADD COLUMN IF NOT EXISTS conflito_realocacoes JSONB")
    cur.execute("ALTER TABLE remarque_solicitacoes ADD COLUMN IF NOT EXISTS motivo_reprovacao TEXT")
    cur.execute("ALTER TABLE remarque_solicitacoes ADD COLUMN IF NOT EXISTS decidido_por_setor VARCHAR(80)")


def ensure_remarque_table_cached(cur, force=False):
    global REMARQUE_SCHEMA_READY
    if not force:
        with REMARQUE_SCHEMA_LOCK:
            if REMARQUE_SCHEMA_READY:
                return

    with REMARQUE_SCHEMA_LOCK:
        if not force and REMARQUE_SCHEMA_READY:
            return
        ensure_remarque_table(cur)
        REMARQUE_SCHEMA_READY = True


def ensure_agendamento_auditoria_table(cur):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS agendamento_auditoria (
            id SERIAL PRIMARY KEY,
            agendamento_id INTEGER NOT NULL,
            acao VARCHAR(80) NOT NULL,
            status_anterior VARCHAR(50),
            status_novo VARCHAR(50),
            usuario_nome VARCHAR(255),
            usuario_username VARCHAR(255),
            detalhes TEXT,
            criado_em TIMESTAMP DEFAULT NOW()
        )
    """)


def ensure_lista_espera_table(cur, force=False):
    global WAITLIST_SCHEMA_READY
    if not force:
        with WAITLIST_SCHEMA_LOCK:
            if WAITLIST_SCHEMA_READY:
                return

    with WAITLIST_SCHEMA_LOCK:
        if not force and WAITLIST_SCHEMA_READY:
            return
        cur.execute("""
            CREATE TABLE IF NOT EXISTS lista_espera (
                id SERIAL PRIMARY KEY,
                paciente_id INTEGER NOT NULL,
                paciente_nome VARCHAR(255) NOT NULL,
                profissional_id INTEGER,
                sala_id INTEGER,
                tipo_atendimento VARCHAR(80),
                prioridade VARCHAR(20) DEFAULT 'normal',
                status VARCHAR(30) DEFAULT 'aguardando',
                preferencia_dias TEXT,
                preferencia_horarios TEXT,
                observacao TEXT,
                criado_por_nome VARCHAR(255),
                criado_por_username VARCHAR(255),
                encaixado_agendamento_id INTEGER,
                criado_em TIMESTAMP DEFAULT NOW(),
                atualizado_em TIMESTAMP DEFAULT NOW(),
                encaixado_em TIMESTAMP
            )
        """)
        cur.execute("ALTER TABLE lista_espera ADD COLUMN IF NOT EXISTS paciente_id INTEGER")
        cur.execute("ALTER TABLE lista_espera ADD COLUMN IF NOT EXISTS paciente_nome VARCHAR(255)")
        cur.execute("ALTER TABLE lista_espera ADD COLUMN IF NOT EXISTS profissional_id INTEGER")
        cur.execute("ALTER TABLE lista_espera ADD COLUMN IF NOT EXISTS sala_id INTEGER")
        cur.execute("ALTER TABLE lista_espera ADD COLUMN IF NOT EXISTS tipo_atendimento VARCHAR(80)")
        cur.execute("ALTER TABLE lista_espera ADD COLUMN IF NOT EXISTS prioridade VARCHAR(20) DEFAULT 'normal'")
        cur.execute("ALTER TABLE lista_espera ADD COLUMN IF NOT EXISTS status VARCHAR(30) DEFAULT 'aguardando'")
        cur.execute("ALTER TABLE lista_espera ADD COLUMN IF NOT EXISTS preferencia_dias TEXT")
        cur.execute("ALTER TABLE lista_espera ADD COLUMN IF NOT EXISTS preferencia_horarios TEXT")
        cur.execute("ALTER TABLE lista_espera ADD COLUMN IF NOT EXISTS observacao TEXT")
        cur.execute("ALTER TABLE lista_espera ADD COLUMN IF NOT EXISTS criado_por_nome VARCHAR(255)")
        cur.execute("ALTER TABLE lista_espera ADD COLUMN IF NOT EXISTS criado_por_username VARCHAR(255)")
        cur.execute("ALTER TABLE lista_espera ADD COLUMN IF NOT EXISTS encaixado_agendamento_id INTEGER")
        cur.execute("ALTER TABLE lista_espera ADD COLUMN IF NOT EXISTS criado_em TIMESTAMP DEFAULT NOW()")
        cur.execute("ALTER TABLE lista_espera ADD COLUMN IF NOT EXISTS atualizado_em TIMESTAMP DEFAULT NOW()")
        cur.execute("ALTER TABLE lista_espera ADD COLUMN IF NOT EXISTS encaixado_em TIMESTAMP")
        cur.execute("UPDATE lista_espera SET prioridade = 'normal' WHERE prioridade IS NULL OR trim(prioridade::text) = ''")
        cur.execute("UPDATE lista_espera SET status = 'aguardando' WHERE status IS NULL OR trim(status::text) = ''")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_lista_espera_status_prioridade ON lista_espera (status, prioridade, criado_em)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_lista_espera_paciente ON lista_espera (paciente_id)")
        WAITLIST_SCHEMA_READY = True


def user_can_manage_waitlist(cur, user):
    if not user:
        return False

    username = user.get('username')
    header = normalize_name_py(f"{user.get('level', '')} {user.get('name', '')} {username or ''}").upper()
    if normalize_level(user.get('level')) in ('admin', 'editor') or 'ADMINISTRADOR' in header:
        return True
    if any(marker in header for marker in ('ATAC', 'RECEP', 'RECEPCAO', 'CEO')):
        return True

    try:
        user_cols = get_table_columns_cached(cur, 'usuarios')
        if 'profissional_id' in user_cols:
            cur.execute("""
                SELECT u.level, u.name, u.profissional_id, p.especialidade
                FROM usuarios u
                LEFT JOIN profissionais p ON CAST(p.id AS TEXT) = CAST(u.profissional_id AS TEXT)
                WHERE lower(u.username) = %s
                LIMIT 1
            """, ((username or '').lower(),))
        else:
            cur.execute("""
                SELECT level, name, NULL AS profissional_id, NULL AS especialidade
                FROM usuarios
                WHERE lower(username) = %s
                LIMIT 1
            """, ((username or '').lower(),))
        row = cur.fetchone()
    except Exception:
        row = None

    if not row:
        return False

    level, name, _profissional_id, especialidade = row
    text = normalize_name_py(f"{level or ''} {name or ''} {especialidade or ''}").upper()
    return normalize_level(level) in ('admin', 'editor') or 'ADMINISTRADOR' in text or any(
        marker in text for marker in ('ATAC', 'RECEP', 'RECEPCAO', 'CEO')
    )


def user_can_view_patients(cur, user):
    if not user:
        return False
    if normalize_level(user.get('level')) in ('admin', 'editor'):
        return True
    return user_can_manage_waitlist(cur, user)


def serialize_waitlist_item(row, keys):
    item = dict(zip(keys, row))
    if item.get('paciente_nome_cadastro'):
        item['paciente_nome'] = item.get('paciente_nome_cadastro')
    item.pop('paciente_nome_cadastro', None)
    for dt_key in ('criado_em', 'atualizado_em', 'encaixado_em'):
        if item.get(dt_key) is not None and hasattr(item[dt_key], 'isoformat'):
            item[dt_key] = item[dt_key].isoformat()
    item['status'] = normalize_waitlist_status(item.get('status'), default='aguardando')
    item['prioridade'] = normalize_waitlist_priority(item.get('prioridade'), default='normal')
    return item


WAITLIST_SELECT_FIELDS = [
    'le.id',
    'le.paciente_id',
    'le.paciente_nome',
    'le.profissional_id',
    'le.sala_id',
    'le.tipo_atendimento',
    'le.prioridade',
    'le.status',
    'le.preferencia_dias',
    'le.preferencia_horarios',
    'le.observacao',
    'le.criado_por_nome',
    'le.criado_por_username',
    'le.encaixado_agendamento_id',
    'le.criado_em',
    'le.atualizado_em',
    'le.encaixado_em',
    'p.nome AS paciente_nome_cadastro',
    'pr.nome AS profissional_nome',
    's.nome AS sala_nome'
]
WAITLIST_SELECT_KEYS = [
    'id',
    'paciente_id',
    'paciente_nome',
    'profissional_id',
    'sala_id',
    'tipo_atendimento',
    'prioridade',
    'status',
    'preferencia_dias',
    'preferencia_horarios',
    'observacao',
    'criado_por_nome',
    'criado_por_username',
    'encaixado_agendamento_id',
    'criado_em',
    'atualizado_em',
    'encaixado_em',
    'paciente_nome_cadastro',
    'profissional_nome',
    'sala_nome'
]


def fetch_waitlist_item_by_id(cur, item_id):
    cur.execute(f"""
        SELECT {', '.join(WAITLIST_SELECT_FIELDS)}
        FROM lista_espera le
        LEFT JOIN pacientes p ON p.id = le.paciente_id
        LEFT JOIN profissionais pr ON pr.id = le.profissional_id
        LEFT JOIN salas s ON s.id = le.sala_id
        WHERE le.id = %s
        LIMIT 1
    """, (item_id,))
    row = cur.fetchone()
    return serialize_waitlist_item(row, WAITLIST_SELECT_KEYS) if row else None


def serialize_audit_value(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def normalize_audit_payload(value):
    if value is None:
        return None
    if isinstance(value, str):
        try:
            return json.loads(value)
        except Exception:
            return value
    try:
        return json.loads(json.dumps(value, ensure_ascii=False, default=str))
    except Exception:
        return str(value)


def ensure_usuarios_audit_columns(cur):
    user_cols = get_table_columns_cached(cur, 'usuarios')
    changed = False
    if 'ultimo_login_em' not in user_cols:
        cur.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS ultimo_login_em TIMESTAMP")
        user_cols.add('ultimo_login_em')
        changed = True
    if 'ultimo_login_ip' not in user_cols:
        cur.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS ultimo_login_ip VARCHAR(80)")
        user_cols.add('ultimo_login_ip')
        changed = True
    if 'ultimo_login_user_agent' not in user_cols:
        cur.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS ultimo_login_user_agent TEXT")
        user_cols.add('ultimo_login_user_agent')
        changed = True
    if changed:
        with TABLE_COLUMNS_CACHE_LOCK:
            TABLE_COLUMNS_CACHE['usuarios'] = user_cols
    return user_cols


def ensure_audit_logs_table(cur):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS audit_logs (
            id SERIAL PRIMARY KEY,
            acao VARCHAR(120) NOT NULL,
            entidade_tipo VARCHAR(80) NOT NULL,
            entidade_id VARCHAR(120),
            entidade_rotulo TEXT,
            usuario_nome VARCHAR(255),
            usuario_username VARCHAR(255),
            dados_antes JSONB,
            dados_depois JSONB,
            detalhes JSONB,
            ip VARCHAR(80),
            user_agent TEXT,
            criado_em TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.execute("ALTER TABLE audit_logs ADD COLUMN IF NOT EXISTS entidade_rotulo TEXT")
    cur.execute("ALTER TABLE audit_logs ADD COLUMN IF NOT EXISTS dados_antes JSONB")
    cur.execute("ALTER TABLE audit_logs ADD COLUMN IF NOT EXISTS dados_depois JSONB")
    cur.execute("ALTER TABLE audit_logs ADD COLUMN IF NOT EXISTS detalhes JSONB")
    cur.execute("ALTER TABLE audit_logs ADD COLUMN IF NOT EXISTS ip VARCHAR(80)")
    cur.execute("ALTER TABLE audit_logs ADD COLUMN IF NOT EXISTS user_agent TEXT")


def get_request_audit_context():
    try:
        forwarded = request.headers.get('X-Forwarded-For', '')
        ip = forwarded.split(',')[0].strip() if forwarded else request.remote_addr
        return {
            'ip': ip,
            'user_agent': request.headers.get('User-Agent')
        }
    except Exception:
        return {'ip': None, 'user_agent': None}


def build_audit_changes(before_data, after_data, fields=None, exclude_fields=None):
    before_data = before_data or {}
    after_data = after_data or {}
    exclude_fields = set(exclude_fields or [])
    if fields is None:
        fields = sorted((set(before_data.keys()) | set(after_data.keys())) - exclude_fields)
    changes = {}
    for field_name in fields:
        if field_name in exclude_fields:
            continue
        old_value = serialize_audit_value(before_data.get(field_name))
        new_value = serialize_audit_value(after_data.get(field_name))
        if str(old_value or '') != str(new_value or ''):
            changes[field_name] = {
                'antes': old_value,
                'depois': new_value
            }
    return changes


def insert_audit_log(
    cur,
    acao,
    entidade_tipo,
    entidade_id=None,
    entidade_rotulo=None,
    actor=None,
    dados_antes=None,
    dados_depois=None,
    detalhes=None,
    ip=None,
    user_agent=None
):
    ensure_audit_logs_table(cur)
    actor = actor or {'name': 'Sistema', 'username': None}
    request_context = get_request_audit_context()
    cur.execute("""
        INSERT INTO audit_logs
            (acao, entidade_tipo, entidade_id, entidade_rotulo, usuario_nome, usuario_username,
             dados_antes, dados_depois, detalhes, ip, user_agent)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        acao,
        entidade_tipo,
        str(entidade_id) if entidade_id is not None else None,
        entidade_rotulo,
        actor.get('name') or actor.get('username') or 'Sistema',
        actor.get('username'),
        Json(normalize_audit_payload(dados_antes)) if dados_antes is not None else None,
        Json(normalize_audit_payload(dados_depois)) if dados_depois is not None else None,
        Json(normalize_audit_payload(detalhes)) if detalhes is not None else None,
        ip if ip is not None else request_context.get('ip'),
        user_agent if user_agent is not None else request_context.get('user_agent')
    ))


def record_audit_log_standalone(
    acao,
    entidade_tipo,
    entidade_id=None,
    entidade_rotulo=None,
    actor=None,
    dados_antes=None,
    dados_depois=None,
    detalhes=None
):
    conn = None
    cur = None
    try:
        conn = get_connection()
        cur = conn.cursor()
        insert_audit_log(
            cur,
            acao,
            entidade_tipo,
            entidade_id=entidade_id,
            entidade_rotulo=entidade_rotulo,
            actor=actor,
            dados_antes=dados_antes,
            dados_depois=dados_depois,
            detalhes=detalhes
        )
        conn.commit()
    except Exception as e:
        print('Erro ao gravar auditoria avulsa:', e)
        try:
            if conn:
                conn.rollback()
        except Exception:
            pass
    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass


def build_agendamento_audit_payload(acao, detalhes, status_anterior=None, status_novo=None):
    parsed_details = normalize_audit_payload(detalhes)
    dados_antes = None
    dados_depois = None
    if isinstance(parsed_details, dict):
        if str(acao or '').startswith('excluido'):
            dados_antes = parsed_details
        elif acao == 'criado':
            dados_depois = parsed_details
        elif isinstance(parsed_details.get('alteracoes'), dict):
            dados_antes = {
                field: change.get('antes')
                for field, change in parsed_details['alteracoes'].items()
                if isinstance(change, dict)
            }
            dados_depois = {
                field: change.get('depois')
                for field, change in parsed_details['alteracoes'].items()
                if isinstance(change, dict)
            }
    general_details = {
        'status_anterior': status_anterior,
        'status_novo': status_novo,
        'detalhes': parsed_details
    }
    return dados_antes, dados_depois, general_details


def build_entity_label(*parts):
    clean_parts = [str(part).strip() for part in parts if part not in (None, '') and str(part).strip()]
    return ' - '.join(clean_parts) if clean_parts else None


def build_audit_user(data=None, authenticated_user=None, fallback_name=None):
    data = data or {}
    if authenticated_user:
        return {
            'name': authenticated_user.get('name') or authenticated_user.get('username') or fallback_name or 'Sistema',
            'username': authenticated_user.get('username')
        }
    return {
        'name': (
            data.get('usuario_nome') or
            data.get('ultima_acao') or
            data.get('created_by') or
            data.get('criado_por') or
            fallback_name or
            'Sistema'
        ),
        'username': data.get('usuario_username') or data.get('username')
    }


def insert_agendamento_audit(cur, agendamento_id, acao, user=None, status_anterior=None, status_novo=None, detalhes=None):
    ensure_agendamento_auditoria_table(cur)
    user = user or {'name': 'Sistema', 'username': None}
    original_detalhes = detalhes
    if detalhes is not None and not isinstance(detalhes, str):
        detalhes = json.dumps(detalhes, ensure_ascii=False, default=str)
    cur.execute("""
        INSERT INTO agendamento_auditoria
            (agendamento_id, acao, status_anterior, status_novo, usuario_nome, usuario_username, detalhes)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (
        int(agendamento_id),
        acao,
        status_anterior,
        status_novo,
        user.get('name') or 'Sistema',
        user.get('username'),
        detalhes
    ))
    try:
        parsed_details = normalize_audit_payload(original_detalhes if original_detalhes is not None else detalhes)
        dados_antes, dados_depois, general_details = build_agendamento_audit_payload(
            acao,
            parsed_details,
            status_anterior=status_anterior,
            status_novo=status_novo
        )
        entity_label = None
        if isinstance(parsed_details, dict):
            entity_label = build_entity_label(
                parsed_details.get('paciente'),
                parsed_details.get('data'),
                parsed_details.get('hora_inicio')
            )
        insert_audit_log(
            cur,
            acao,
            'agendamento',
            entidade_id=agendamento_id,
            entidade_rotulo=entity_label,
            actor=user,
            dados_antes=dados_antes,
            dados_depois=dados_depois,
            detalhes=general_details
        )
    except Exception as e:
        print('Erro ao espelhar auditoria de agendamento no log geral:', e)


def user_can_authorize_remarque(cur, user):
    if not user:
        return False

    username = user.get('username')
    level_or_name = f"{user.get('level', '')} {user.get('name', '')} {username or ''}".upper()
    if normalize_level(user.get('level')) == 'admin' or 'ADMINISTRADOR' in level_or_name:
        return True
    if 'CEO' in level_or_name:
        return True

    try:
        user_cols = get_table_columns_cached(cur, 'usuarios')
        if 'profissional_id' in user_cols:
            cur.execute("""
                SELECT u.level, u.name, u.profissional_id, p.especialidade
                FROM usuarios u
                LEFT JOIN profissionais p ON CAST(p.id AS TEXT) = CAST(u.profissional_id AS TEXT)
                WHERE lower(u.username) = %s
                LIMIT 1
            """, ((username or '').lower(),))
        else:
            cur.execute("""
                SELECT level, name, NULL AS profissional_id, NULL AS especialidade
                FROM usuarios
                WHERE lower(username) = %s
                LIMIT 1
            """, ((username or '').lower(),))
        row = cur.fetchone()
    except Exception:
        row = None

    if not row:
        return False

    level, name, profissional_id, especialidade = row
    text = f"{level or ''} {name or ''} {especialidade or ''}".upper()
    return normalize_level(level) == 'admin' or 'ADMINISTRADOR' in text or 'CEO' in text or 'ATAC' in text or 'FINANCEIRO' in text


def user_can_manage_remarque_config(cur, user):
    if not user:
        return False

    username = user.get('username')
    level_or_name = f"{user.get('level', '')} {user.get('name', '')} {username or ''}".upper()
    if 'CEO' in level_or_name or 'ATAC' in level_or_name or 'FINANCEIRO' in level_or_name:
        return True

    try:
        user_cols = get_table_columns_cached(cur, 'usuarios')
        if 'profissional_id' in user_cols:
            cur.execute("""
                SELECT u.level, u.name, u.profissional_id, p.especialidade
                FROM usuarios u
                LEFT JOIN profissionais p ON CAST(p.id AS TEXT) = CAST(u.profissional_id AS TEXT)
                WHERE lower(u.username) = %s
                LIMIT 1
            """, ((username or '').lower(),))
        else:
            cur.execute("""
                SELECT level, name, NULL AS profissional_id, NULL AS especialidade
                FROM usuarios
                WHERE lower(username) = %s
                LIMIT 1
            """, ((username or '').lower(),))
        row = cur.fetchone()
    except Exception:
        row = None

    if not row:
        return False

    level, name, profissional_id, especialidade = row
    text = f"{level or ''} {name or ''} {especialidade or ''}".upper()
    return 'CEO' in text or 'ATAC' in text or 'FINANCEIRO' in text


PROFESSIONAL_STATUS_UPDATE_ALLOWED = {'finalizado', 'cancelado_profissional', 'em_analise', 'online'}


def user_matches_appointment_professional(user_profissional_id, current_data):
    if user_profissional_id in (None, '') or not current_data:
        return False
    appointment_profissional_id = current_data.get('profissional_id')
    appointment_profissional_text = current_data.get('profissional')
    user_profissional_text = str(user_profissional_id)
    return (
        (appointment_profissional_id not in (None, '') and str(appointment_profissional_id) == user_profissional_text)
        or (appointment_profissional_text not in (None, '') and str(appointment_profissional_text) == user_profissional_text)
    )


def user_can_update_appointment_status(cur, user, status=None, current_data=None):
    if not user:
        return False

    normalized_status = normalize_appointment_status(status, default=None) if status is not None else None
    username = user.get('username')
    level_or_name = f"{user.get('level', '')} {user.get('name', '')} {username or ''}"
    normalized_header = normalize_name_py(level_or_name).upper()
    if normalize_level(user.get('level')) == 'admin' or 'ADMINISTRADOR' in normalized_header:
        return True
    if any(marker in normalized_header for marker in ('ATAC', 'RECEP', 'RECEPCAO', 'CEO')):
        return True
    if (
        normalize_level(user.get('level')) == 'viewer'
        and normalized_status in PROFESSIONAL_STATUS_UPDATE_ALLOWED
        and user_matches_appointment_professional(user.get('professionalId') or user.get('profissional_id'), current_data)
    ):
        return True

    try:
        user_cols = get_table_columns_cached(cur, 'usuarios')
        if 'profissional_id' in user_cols:
            cur.execute("""
                SELECT u.level, u.name, u.profissional_id, p.especialidade
                FROM usuarios u
                LEFT JOIN profissionais p ON CAST(p.id AS TEXT) = CAST(u.profissional_id AS TEXT)
                WHERE lower(u.username) = %s
                LIMIT 1
            """, ((username or '').lower(),))
        else:
            cur.execute("""
                SELECT level, name, NULL AS profissional_id, NULL AS especialidade
                FROM usuarios
                WHERE lower(username) = %s
                LIMIT 1
            """, ((username or '').lower(),))
        row = cur.fetchone()
    except Exception:
        row = None

    if not row:
        return False

    level, name, profissional_id, especialidade = row
    text = normalize_name_py(f"{level or ''} {name or ''} {especialidade or ''}").upper()
    has_full_status_access = normalize_level(level) == 'admin' or 'ADMINISTRADOR' in text or any(
        marker in text for marker in ('ATAC', 'RECEP', 'RECEPCAO', 'CEO')
    )
    if has_full_status_access:
        return True
    return (
        normalize_level(level) == 'viewer'
        and normalized_status in PROFESSIONAL_STATUS_UPDATE_ALLOWED
        and user_matches_appointment_professional(profissional_id, current_data)
    )


def get_remarque_authorizer_sector(cur, user):
    if not user:
        return 'SETOR'

    username = user.get('username')
    level_or_name = f"{user.get('level', '')} {user.get('name', '')} {username or ''}".upper()
    if 'CEO' in level_or_name:
        return 'CEO'

    try:
        user_cols = get_table_columns_cached(cur, 'usuarios')
        if 'profissional_id' in user_cols:
            cur.execute("""
                SELECT u.level, u.name, p.especialidade
                FROM usuarios u
                LEFT JOIN profissionais p ON CAST(p.id AS TEXT) = CAST(u.profissional_id AS TEXT)
                WHERE lower(u.username) = %s
                LIMIT 1
            """, ((username or '').lower(),))
        else:
            cur.execute("""
                SELECT level, name, NULL AS especialidade
                FROM usuarios
                WHERE lower(username) = %s
                LIMIT 1
            """, ((username or '').lower(),))
        row = cur.fetchone()
    except Exception:
        row = None

    text = level_or_name
    if row:
        text = f"{row[0] or ''} {row[1] or ''} {row[2] or ''} {level_or_name}".upper()
    if 'ATAC' in text:
        return 'ATAC'
    if 'FINANCEIRO' in text:
        return 'FINANCEIRO'
    if 'CEO' in text:
        return 'CEO'
    if normalize_level(user.get('level')) == 'admin' or 'ADMINISTRADOR' in text:
        return 'ADMINISTRADOR'
    return 'SETOR'


def normalize_existing_agendamento_statuses(cur):
    try:
        appointment_cols = get_table_columns_cached(cur, 'agendamentos')
        if 'status' not in appointment_cols:
            return
        cur.execute("SELECT DISTINCT status FROM agendamentos WHERE status IS NOT NULL")
        for (status_value,) in cur.fetchall():
            normalized = normalize_appointment_status(status_value, default=None)
            if normalized and normalized != status_value:
                cur.execute(
                    "UPDATE agendamentos SET status = %s WHERE status = %s",
                    (normalized, status_value)
                )
        cur.execute("UPDATE agendamentos SET status = 'agendado' WHERE status IS NULL OR trim(status::text) = ''")
    except Exception as e:
        print('Erro ao normalizar status de agendamentos:', e)


def migrate_plain_user_passwords(cur):
    try:
        cur.execute("SELECT username, password FROM usuarios")
        for username, stored_password in cur.fetchall():
            if stored_password and not is_password_hash(stored_password):
                cur.execute(
                    "UPDATE usuarios SET password = %s WHERE lower(username) = %s",
                    (hash_password(stored_password), str(username or '').lower())
                )
    except Exception as e:
        print('Erro ao migrar senhas de usuarios para hash:', e)


def create_performance_indexes(cur):
    index_statements = [
        "CREATE INDEX IF NOT EXISTS idx_usuarios_username_lower ON usuarios (lower(username))",
        "CREATE INDEX IF NOT EXISTS idx_profissionais_nome_lower ON profissionais (lower(nome))",
        "CREATE INDEX IF NOT EXISTS idx_agendamentos_data_criado_em ON agendamentos (data, criado_em DESC)",
        "CREATE INDEX IF NOT EXISTS idx_agendamentos_prof_data_hora ON agendamentos (profissional, data, hora_inicio)",
        "CREATE INDEX IF NOT EXISTS idx_agendamentos_profissional_id_data_hora ON agendamentos (profissional_id, data, hora_inicio)",
        "CREATE INDEX IF NOT EXISTS idx_agendamentos_paciente_id_data_hora ON agendamentos (paciente_id, data, hora_inicio)",
        "CREATE INDEX IF NOT EXISTS idx_agendamentos_sala_data_hora ON agendamentos (sala_id, data, hora_inicio)",
        "CREATE INDEX IF NOT EXISTS idx_agendamentos_criado_em ON agendamentos (criado_em DESC)",
        "CREATE INDEX IF NOT EXISTS idx_remarque_agendamento_status ON remarque_solicitacoes (agendamento_id, status)",
        "CREATE INDEX IF NOT EXISTS idx_remarque_solicitante_status ON remarque_solicitacoes (solicitado_por_username, status)",
        "CREATE INDEX IF NOT EXISTS idx_remarque_status_solicitado_em ON remarque_solicitacoes (status, solicitado_em DESC)",
        "CREATE INDEX IF NOT EXISTS idx_remarque_agendamento_status_autorizado ON remarque_solicitacoes (agendamento_id, status, autorizado_em DESC)",
        "CREATE INDEX IF NOT EXISTS idx_agendamento_auditoria_agendamento_em ON agendamento_auditoria (agendamento_id, criado_em DESC)",
        "CREATE INDEX IF NOT EXISTS idx_audit_logs_criado_em ON audit_logs (criado_em DESC)",
        "CREATE INDEX IF NOT EXISTS idx_audit_logs_entidade ON audit_logs (entidade_tipo, entidade_id, criado_em DESC)",
        "CREATE INDEX IF NOT EXISTS idx_audit_logs_usuario ON audit_logs (usuario_username, criado_em DESC)",
        "CREATE INDEX IF NOT EXISTS idx_audit_logs_acao ON audit_logs (acao, criado_em DESC)",
        "CREATE INDEX IF NOT EXISTS idx_lista_espera_status_prioridade ON lista_espera (status, prioridade, criado_em)",
        "CREATE INDEX IF NOT EXISTS idx_lista_espera_paciente ON lista_espera (paciente_id)"
    ]
    for statement in index_statements:
        cur.execute(statement)


def ensure_performance_indexes_background():
    conn = None
    cur = None
    try:
        conn = get_connection()
        cur = conn.cursor()
        ensure_remarque_table_cached(cur, force=True)
        ensure_lista_espera_table(cur, force=True)
        ensure_agendamento_auditoria_table(cur)
        ensure_audit_logs_table(cur)
        ensure_usuarios_audit_columns(cur)
        ensure_user_preferences_table(cur)
        appointment_cols = ensure_salas_schema(cur)
        appointment_cols = migrate_existing_agendamento_links(
            cur,
            delete_invalid=is_truthy_env(os.environ.get('DELETE_INVALID_TEST_APPOINTMENTS', '1')),
            enforce_constraints=is_truthy_env(os.environ.get('ENFORCE_APPOINTMENT_LINK_CONSTRAINTS', '1'))
        )
        ensure_agendamento_recurrence_columns(cur, appointment_cols)
        normalize_existing_agendamento_statuses(cur)
        migrate_plain_user_passwords(cur)
        create_performance_indexes(cur)
        conn.commit()
    except Exception as e:
        print('Erro ao criar indices de performance:', e)
        try:
            if conn:
                conn.rollback()
        except Exception:
            pass
    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass


@app.route("/api/teste")
def teste():
    return jsonify({"status": "API funcionando"})


def fetch_salas_rows(cur, include_inactive=False):
    where_sql = '' if include_inactive else 'WHERE ativo IS TRUE'
    cur.execute(f"""
        SELECT id, nome, cor, ativo
        FROM salas
        {where_sql}
        ORDER BY nome
    """)
    return cur.fetchall()


@app.route("/api/salas", methods=["GET"])
def listar_salas():
    auth_check = require_authenticated()
    if auth_check:
        return auth_check

    try:
        conn = get_connection()
        cur = conn.cursor()

        include_inactive = str(request.args.get('include_inactive') or '').lower() in ('1', 'true', 'yes')
        try:
            rows = fetch_salas_rows(cur, include_inactive)
        except Exception:
            conn.rollback()
            ensure_salas_schema(cur)
            conn.commit()
            rows = fetch_salas_rows(cur, include_inactive)

        cur.close()
        conn.close()

        salas = [
            {
                'id': row[0],
                'nome': row[1],
                'cor': row[2],
                'ativo': row[3]
            }
            for row in rows
        ]
        return jsonify({'success': True, 'salas': salas})
    except Exception as e:
        print('Erro ao listar salas:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500

# =======================
# USUÁRIOS
# =======================
@app.route("/api/usuarios", methods=["POST"])
def criar_usuario():
    authenticated_user, err = require_admin_permission('canManageUsers')
    if err:
        return err

    data = request.json or {}

    print('Payload criar_usuario:', data)

    username = data.get("username") or data.get('email')
    name = data.get("name") or data.get('nome')
    password = data.get('password') or data.get('senha')
    level = data.get('level') or data.get('nivel') or 'viewer'
    level = normalize_level(level)
    notes = data.get("notes", "")
    profissional_id = data.get('profissional_id') or data.get('professional_id') or data.get('professionalId') or data.get('profissionalId')
    preferences = data.get('preferences') if 'preferences' in data else data.get('preferencias')
    current_user = (authenticated_user or {}).get('username') or "admin"

    if not username:
        return jsonify({"success": False, "error": "username is required"}), 400
    if not password:
        return jsonify({"success": False, "error": "password is required"}), 400

    try:
        conn = get_connection()
        cur = conn.cursor()

        columns = [
            'username',
            'password',
            'level',
            'name',
            'notes',
            'created_by',
            'is_active'
        ]
        values = [
            username,
            hash_password(password),
            level,
            name,
            notes,
            current_user,
            True
        ]

        if profissional_id not in (None, ''):
            columns.insert(5, 'profissional_id')
            values.insert(5, profissional_id)

        placeholders = ', '.join(['%s'] * len(values))
        columns_sql = ',\n                '.join(columns)

        cur.execute(f"""
            INSERT INTO usuarios (
                {columns_sql},
                created_at
            )
            VALUES ({placeholders}, NOW())
        """, tuple(values))

        saved_preferences = set_user_preferences(cur, username, preferences or {}, current_user, level)

        insert_audit_log(
            cur,
            'usuario_criado',
            'usuario',
            entidade_id=username,
            entidade_rotulo=name or username,
            actor=authenticated_user,
            dados_depois={
                'username': username,
                'name': name,
                'level': level,
                'profissional_id': profissional_id,
                'is_active': True,
                'preferences': saved_preferences
            }
        )

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({"success": True})

    except Exception as e:
        print("ERRO AO INSERIR USUÁRIO:", e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({"success": False, "error": str(e)}), 500



# Endpoint para listar usuários (útil para verificação/sincronização)
@app.route("/api/usuarios", methods=["GET"])
def listar_usuarios():
    _authenticated_user, auth_check = require_admin_permission('canManageUsers')
    if auth_check:
        return auth_check

    try:
        conn = get_connection()
        cur = conn.cursor()
        user_cols = ensure_usuarios_audit_columns(cur)
        ensure_user_preferences_table(cur)
        conn.commit()

        select_fields = ['id', 'name', 'username', 'level', 'created_at']
        if 'profissional_id' in user_cols:
            select_fields.append('profissional_id')
        if 'ultimo_login_em' in user_cols:
            select_fields.append('ultimo_login_em')
        if 'ultimo_login_ip' in user_cols:
            select_fields.append('ultimo_login_ip')
        if 'is_active' in user_cols:
            select_fields.append('is_active')

        cur.execute(
            f"SELECT {', '.join(select_fields)} FROM usuarios ORDER BY created_at DESC NULLS LAST LIMIT 100"
        )
        rows = cur.fetchall()
        column_names = [desc[0] for desc in cur.description]

        users_list = []
        for r in rows:
            item = dict(zip(column_names, r))
            last_login = item.get('ultimo_login_em')
            user_obj = {
                'id': item.get('id'),
                'name': item.get('name'),
                'username': item.get('username'),
                'level': normalize_level(item.get('level')),
                'created_at': item.get('created_at').isoformat() if item.get('created_at') else None,
                'last_login': last_login.isoformat() if hasattr(last_login, 'isoformat') else last_login,
                'ultimo_login_em': last_login.isoformat() if hasattr(last_login, 'isoformat') else last_login,
                'ultimo_login_ip': item.get('ultimo_login_ip'),
                'is_active': item.get('is_active', True),
                'isActive': item.get('is_active', True)
            }
            if 'profissional_id' in item:
                user_obj['professionalId'] = item.get('profissional_id')
                user_obj['profissional_id'] = item.get('profissional_id')
            user_preferences = get_user_preferences(cur, item.get('username'), user_obj.get('level'))
            user_obj['preferences'] = user_preferences
            user_obj['effectivePermissions'] = get_effective_user_permissions(user_obj.get('level'), user_preferences)
            user_obj['permissions'] = user_obj['effectivePermissions']
            users_list.append(user_obj)

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({"success": True, "users": users_list})
    except Exception as e:
        print('Erro ao listar usuários:', e)
        return jsonify({"success": False, "error": str(e)})


# Autenticação: valida usuário e senha no banco
@app.route('/api/authenticate', methods=['POST'])
def authenticate():
    data = request.json or {}
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''

    user, auth_error = authenticate_credentials(username, password)
    if auth_error:
        if username:
            record_audit_log_standalone(
                'login_falha',
                'usuario',
                entidade_id=username,
                entidade_rotulo=username,
                actor={'name': username, 'username': username},
                detalhes={'motivo': 'credenciais_invalidas_ou_usuario_inativo'}
            )
        return auth_error

    conn = None
    cur = None
    try:
        conn = get_connection()
        cur = conn.cursor()
        ensure_usuarios_audit_columns(cur)
        context = get_request_audit_context()
        cur.execute("""
            UPDATE usuarios
            SET ultimo_login_em = NOW(),
                ultimo_login_ip = %s,
                ultimo_login_user_agent = %s
            WHERE lower(username) = %s
            RETURNING ultimo_login_em, ultimo_login_ip
        """, (context.get('ip'), context.get('user_agent'), username.lower()))
        login_row = cur.fetchone()
        if login_row:
            user['lastLogin'] = login_row[0].isoformat() if hasattr(login_row[0], 'isoformat') else login_row[0]
            user['ultimo_login_em'] = user['lastLogin']
            user['ultimo_login_ip'] = login_row[1]
        insert_audit_log(
            cur,
            'login_sucesso',
            'usuario',
            entidade_id=user.get('username'),
            entidade_rotulo=user.get('name') or user.get('username'),
            actor=user,
            detalhes={'origem': 'login'}
        )
        conn.commit()
    except Exception as e:
        print('Erro ao atualizar ultimo login/auditoria:', e)
        try:
            if conn:
                conn.rollback()
        except Exception:
            pass
    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass

    session['current_user'] = {
        'username': user['username'],
        'level': user['level'],
        'name': user.get('name')
    }
    return jsonify({'success': True, 'user': user})


# Atualizar usuário (por username/email)
@app.route('/api/me', methods=['GET'])
def authenticated_user_info():
    user, auth_error = get_authenticated_user()
    if auth_error:
        return auth_error
    return jsonify({'success': True, 'user': user})


@app.route('/api/me/preferences', methods=['PUT'])
def update_authenticated_user_preferences():
    user, auth_error = get_authenticated_user()
    if auth_error:
        return auth_error

    data = request.json or {}
    ui_payload = data.get('ui') if 'ui' in data else data.get('interface')
    if not isinstance(ui_payload, dict):
        return jsonify({'success': False, 'error': 'Preferencias visuais invalidas'}), 400

    conn = None
    cur = None
    try:
        username = user.get('username')
        level = user.get('level')
        conn = get_connection()
        cur = conn.cursor()
        current_preferences = get_user_preferences(cur, username, level)
        next_preferences = {
            'permissions': current_preferences.get('permissions') or get_base_user_permissions(level),
            'ui': ui_payload
        }
        saved_preferences = set_user_preferences(
            cur,
            username,
            next_preferences,
            username,
            level
        )
        conn.commit()
        cur.close()
        conn.close()

        updated_user = sanitize_user_for_client(
            user.get('username'),
            user.get('name'),
            user.get('level'),
            user.get('professionalId') or user.get('profissional_id'),
            saved_preferences
        )
        return jsonify({
            'success': True,
            'preferences': saved_preferences,
            'effectivePermissions': get_effective_user_permissions(level, saved_preferences),
            'user': updated_user
        })
    except Exception as e:
        print('Erro ao atualizar preferencias visuais do usuario:', e)
        try:
            if conn:
                conn.rollback()
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/logout', methods=['POST'])
def logout_user():
    try:
        current_user = session.get('current_user')
        if current_user and current_user.get('username'):
            record_audit_log_standalone(
                'logout',
                'usuario',
                entidade_id=current_user.get('username'),
                entidade_rotulo=current_user.get('name') or current_user.get('username'),
                actor=current_user,
                detalhes={'origem': 'logout'}
            )
    except Exception:
        pass
    session.pop('current_user', None)
    session.pop('auditoria_unlocked_until', None)
    return jsonify({'success': True})


@app.route('/api/cache/clear', methods=['POST'])
def clear_cache():
    auth_check = require_authenticated()
    if auth_check:
        return auth_check
    return jsonify({'success': True, 'cleared': clear_runtime_caches()})


def serialize_audit_log_row(row, column_names):
    item = dict(zip(column_names, row))
    for key in ('dados_antes', 'dados_depois', 'detalhes'):
        item[key] = normalize_audit_payload(item.get(key))
    if item.get('criado_em') is not None and hasattr(item['criado_em'], 'isoformat'):
        item['criado_em'] = item['criado_em'].isoformat()
    return item


def get_auditoria_unlocked_until():
    try:
        return int(session.get('auditoria_unlocked_until') or 0)
    except Exception:
        return 0


def is_auditoria_unlocked():
    return get_auditoria_unlocked_until() > get_current_timestamp_seconds()


def require_auditoria_password_unlocked():
    if is_auditoria_unlocked():
        return None
    return jsonify({
        'success': False,
        'error': 'Senha da auditoria necessaria',
        'requires_audit_password': True
    }), 403


@app.route('/api/auditoria/desbloquear', methods=['POST'])
def desbloquear_auditoria():
    authenticated_user, auth_error = get_authenticated_user()
    if auth_error:
        return auth_error
    if normalize_level(authenticated_user.get('level')) != 'admin':
        return jsonify({'success': False, 'error': 'Acesso negado'}), 403

    audit_password = get_auditoria_password()
    if not audit_password:
        return jsonify({
            'success': False,
            'error': 'Senha de auditoria nao configurada no servidor',
            'audit_password_configured': False
        }), 500

    data = request.json or {}
    provided_password = str(data.get('senha') or data.get('password') or '').strip()
    if not provided_password or not hmac.compare_digest(provided_password, audit_password):
        session.pop('auditoria_unlocked_until', None)
        session.modified = True
        return jsonify({'success': False, 'error': 'Senha da auditoria incorreta'}), 403

    now = get_current_timestamp_seconds()
    unlocked_until = now + AUDIT_ACCESS_TTL_SECONDS
    session['auditoria_unlocked_until'] = unlocked_until
    session.modified = True
    return jsonify({
        'success': True,
        'unlocked': True,
        'expires_in': AUDIT_ACCESS_TTL_SECONDS,
        'expires_at': unlocked_until
    })


@app.route('/api/auditoria/status', methods=['GET'])
def status_auditoria():
    admin_check = require_admin()
    if admin_check:
        return admin_check
    unlocked_until = get_auditoria_unlocked_until()
    now = get_current_timestamp_seconds()
    return jsonify({
        'success': True,
        'unlocked': unlocked_until > now,
        'expires_in': max(0, unlocked_until - now),
        'audit_password_configured': bool(get_auditoria_password())
    })


def query_audit_logs(force_deleted=False):
    auth_check = require_admin()
    if auth_check:
        return auth_check
    audit_lock = require_auditoria_password_unlocked()
    if audit_lock:
        return audit_lock

    try:
        limit = int(request.args.get('limit') or 100)
    except Exception:
        limit = 100
    limit = max(1, min(limit, 500))

    where = []
    params = []

    if force_deleted or str(request.args.get('deleted') or request.args.get('excluidos') or '').lower() in ('1', 'true', 'yes', 'sim', 's'):
        where.append("acao ILIKE %s")
        params.append('%exclu%')

    action = (request.args.get('acao') or request.args.get('action') or '').strip()
    if action:
        if action == 'editado':
            where.append("(acao ILIKE %s OR acao ILIKE %s)")
            params.extend(['%editado%', '%atualizado%'])
        elif action in ('criado', 'atualizado', 'status_alterado'):
            where.append("acao ILIKE %s")
            params.append(f"%{action}%")
        else:
            where.append("lower(acao) = lower(%s)")
            params.append(action)

    entity_type = (request.args.get('entidade_tipo') or request.args.get('entity_type') or request.args.get('modulo') or '').strip()
    if entity_type:
        where.append("lower(entidade_tipo) = lower(%s)")
        params.append(entity_type)

    entity_id = (request.args.get('entidade_id') or request.args.get('entity_id') or '').strip()
    if entity_id:
        where.append("entidade_id = %s")
        params.append(entity_id)

    user_filter = (request.args.get('usuario') or request.args.get('user') or '').strip()
    if user_filter:
        where.append("(lower(COALESCE(usuario_username, '')) = lower(%s) OR usuario_nome ILIKE %s)")
        params.extend([user_filter, f"%{user_filter}%"])

    start_date = normalize_date_for_db(request.args.get('inicio') or request.args.get('start') or request.args.get('from'))
    if start_date:
        where.append("criado_em >= %s")
        params.append(start_date)

    end_date = normalize_date_for_db(request.args.get('fim') or request.args.get('end') or request.args.get('to'))
    if end_date:
        where.append("criado_em < (%s::date + INTERVAL '1 day')")
        params.append(end_date)

    search = (request.args.get('q') or request.args.get('busca') or '').strip()
    if search:
        pattern = f"%{search}%"
        where.append("""(
            entidade_rotulo ILIKE %s OR entidade_id ILIKE %s OR usuario_nome ILIKE %s OR
            usuario_username ILIKE %s OR acao ILIKE %s OR entidade_tipo ILIKE %s OR
            COALESCE(detalhes::text, '') ILIKE %s OR COALESCE(dados_antes::text, '') ILIKE %s OR
            COALESCE(dados_depois::text, '') ILIKE %s
        )""")
        params.extend([pattern] * 9)

    where_sql = f"WHERE {' AND '.join(where)}" if where else ''

    conn = None
    cur = None
    try:
        conn = get_connection()
        cur = conn.cursor()
        ensure_audit_logs_table(cur)
        conn.commit()
        cur.execute(f"""
            SELECT id, acao, entidade_tipo, entidade_id, entidade_rotulo,
                   usuario_nome, usuario_username, dados_antes, dados_depois,
                   detalhes, ip, user_agent, criado_em
            FROM audit_logs
            {where_sql}
            ORDER BY criado_em DESC, id DESC
            LIMIT %s
        """, tuple(params + [limit]))
        rows = cur.fetchall()
        column_names = [desc[0] for desc in cur.description]
        logs = [serialize_audit_log_row(row, column_names) for row in rows]
        cur.close()
        conn.close()
        return jsonify({'success': True, 'logs': logs, 'limit': limit})
    except Exception as e:
        print('Erro ao listar auditoria geral:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/auditoria', methods=['GET'])
def listar_auditoria_geral():
    return query_audit_logs(force_deleted=False)


@app.route('/api/auditoria/excluidos', methods=['GET'])
def listar_auditoria_excluidos():
    return query_audit_logs(force_deleted=True)


@app.route("/api/usuarios/<username>", methods=["PUT"])
def atualizar_usuario(username):
    data = request.json or {}
    authenticated_user, auth_error = get_authenticated_user()
    if auth_error:
        return auth_error

    is_admin = normalize_level(authenticated_user.get('level')) == 'admin'
    is_self = str(authenticated_user.get('username') or '').lower() == str(username or '').lower()
    if not is_admin and not is_self:
        return jsonify({'success': False, 'error': 'Acesso negado'}), 403
    if is_admin and not is_self and not user_has_effective_permission(authenticated_user, 'canManageUsers'):
        return jsonify({'success': False, 'error': 'Permissao personalizada negada'}), 403

    name = data.get('name')
    password = data.get('password')
    level = data.get('level')
    profissional_id = data.get('profissional_id') or data.get('professional_id') or data.get('professionalId') or data.get('profissionalId')
    is_active = data.get('isActive', None)
    preferences_provided = 'preferences' in data or 'preferencias' in data
    preferences = data.get('preferences') if 'preferences' in data else data.get('preferencias')

    if not is_admin and (level is not None or profissional_id is not None or is_active is not None or preferences_provided):
        return jsonify({'success': False, 'error': 'Apenas administradores podem alterar permissões de usuário'}), 403
    if is_admin and preferences_provided and not user_has_effective_permission(authenticated_user, 'canManageUsers'):
        return jsonify({'success': False, 'error': 'Permissao personalizada negada'}), 403

    try:
        conn = get_connection()
        cur = conn.cursor()
        user_cols = ensure_usuarios_audit_columns(cur)
        ensure_user_preferences_table(cur)
        select_fields = ['username', 'name', 'level', 'is_active']
        if 'profissional_id' in user_cols:
            select_fields.append('profissional_id')
        cur.execute(
            f"SELECT {', '.join(select_fields)} FROM usuarios WHERE lower(username) = %s LIMIT 1",
            (str(username or '').lower(),)
        )
        before_row = cur.fetchone()
        if not before_row:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Usuario nao encontrado'}), 404
        before_data = dict(zip(select_fields, before_row))
        before_preferences = get_user_preferences(cur, before_data.get('username'), before_data.get('level'))
        before_data['preferences'] = before_preferences

        # Build update dynamically
        fields = []
        values = []
        if name is not None:
            fields.append('name = %s')
            values.append(name)
        if password:
            fields.append('password = %s')
            values.append(hash_password(password))
        if level is not None:
            level = normalize_level(level)
            fields.append('level = %s')
            values.append(level)
        if profissional_id is not None:
            fields.append('profissional_id = %s')
            values.append(profissional_id)
        if is_active is not None:
            fields.append('is_active = %s')
            values.append(is_active)

        if len(fields) == 0 and not preferences_provided:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Nada para atualizar'})

        if fields:
            values.append(str(username or '').lower())
            sql = f"UPDATE usuarios SET {', '.join(fields)} WHERE lower(username) = %s"
            cur.execute(sql, tuple(values))

        cur.execute(
            f"SELECT {', '.join(select_fields)} FROM usuarios WHERE lower(username) = %s LIMIT 1",
            (str(username or '').lower(),)
        )
        after_row = cur.fetchone()
        after_data = dict(zip(select_fields, after_row)) if after_row else {}
        if preferences_provided:
            if is_self and normalize_level(after_data.get('level') or level or before_data.get('level')) == 'admin':
                preferences = normalize_user_preferences(preferences or {}, after_data.get('level') or level or before_data.get('level'))
                preferences['permissions']['canManageUsers'] = True
            after_preferences = set_user_preferences(
                cur,
                after_data.get('username') or username,
                preferences or {},
                authenticated_user.get('username'),
                after_data.get('level') or level or before_data.get('level')
            )
        else:
            after_preferences = get_user_preferences(cur, after_data.get('username') or username, after_data.get('level') or level)
        after_data['preferences'] = after_preferences
        changes = build_audit_changes(before_data, after_data)
        if password:
            changes['senha'] = {'antes': '[protegida]', 'depois': 'alterada'}
        if changes:
            action = 'usuario_atualizado'
            if 'is_active' in changes and len(changes) == 1:
                action = 'usuario_reativado' if after_data.get('is_active') else 'usuario_inativado'
            insert_audit_log(
                cur,
                action,
                'usuario',
                entidade_id=after_data.get('username') or username,
                entidade_rotulo=after_data.get('name') or after_data.get('username') or username,
                actor=authenticated_user,
                dados_antes=before_data,
                dados_depois=after_data,
                detalhes={'alteracoes': changes}
            )
        conn.commit()
        cur.close()
        conn.close()
        invalidate_auth_user_cache(username)

        return jsonify({'success': True, 'preferences': after_preferences, 'effectivePermissions': get_effective_user_permissions(after_data.get('level'), after_preferences)})
    except Exception as e:
        print('Erro ao atualizar usuário:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)})


# Deletar usuário por username/email
@app.route('/api/usuarios/<username>', methods=['DELETE'])
def deletar_usuario(username):
    # Require admin credentials to delete a user
    authenticated_user, admin_check = require_admin_permission('canManageUsers')
    if admin_check:
        return admin_check

    try:
        conn = get_connection()
        cur = conn.cursor()
        user_cols = ensure_usuarios_audit_columns(cur)
        select_fields = ['username', 'name', 'level', 'is_active', 'created_at']
        if 'profissional_id' in user_cols:
            select_fields.append('profissional_id')
        if 'ultimo_login_em' in user_cols:
            select_fields.append('ultimo_login_em')
        cur.execute(
            f"SELECT {', '.join(select_fields)} FROM usuarios WHERE lower(username) = %s LIMIT 1",
            (str(username or '').lower(),)
        )
        before_row = cur.fetchone()
        before_data = dict(zip(select_fields, before_row)) if before_row else None
        if before_data:
            before_data['preferences'] = get_user_preferences(cur, before_data.get('username'), before_data.get('level'))
        ensure_user_preferences_table(cur)
        cur.execute("DELETE FROM usuario_preferencias WHERE lower(username) = %s", (str(username or '').lower(),))
        cur.execute("DELETE FROM usuarios WHERE lower(username) = %s", (str(username or '').lower(),))
        if before_data:
            insert_audit_log(
                cur,
                'usuario_excluido',
                'usuario',
                entidade_id=before_data.get('username') or username,
                entidade_rotulo=before_data.get('name') or before_data.get('username') or username,
                actor=authenticated_user,
                dados_antes=before_data
            )
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        print('Erro ao deletar usuário:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)})

# Helper: normalize name (strip accents, lower, collapse spaces)
def normalize_name_py(name):
    if not name:
        return ''
    s = str(name)
    s = unicodedata.normalize('NFD', s)
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower().strip()
    s = ' '.join(s.split())
    return s

# Helper: perform batch insert for professionals — returns mapping tempId -> serverId and created list
def do_batch_profissionais(items):
    normalized_to_temp = {}
    names_set = set()
    temp_lookup = {}
    for it in items:
        temp = it.get('tempId') or it.get('id')
        nome = it.get('nome') or it.get('name') or ''
        key = normalize_name_py(nome)
        temp_lookup[str(temp)] = nome
        if key:
            normalized_to_temp[key] = normalized_to_temp.get(key, []) + [str(temp)]
            names_set.add(nome.lower())

    mapping = {}
    created = []

    conn = get_connection()
    cur = conn.cursor()

    try:
        # Try to use unaccent() if available, else fallback to lower(nome)
        existing = {}
        if names_set:
            try:
                placeholders = ','.join(['%s'] * len(names_set))
                cur.execute(f"SELECT id, nome FROM profissionais WHERE lower(unaccent(nome)) IN ({placeholders})", tuple([n for n in names_set]))
                rows = cur.fetchall()
                for r in rows:
                    existing[normalize_name_py(r[1])] = r[0]
            except Exception:
                placeholders = ','.join(['%s'] * len(names_set))
                cur.execute(f"SELECT id, nome FROM profissionais WHERE lower(nome) IN ({placeholders})", tuple([n for n in names_set]))
                rows = cur.fetchall()
                for r in rows:
                    existing[normalize_name_py(r[1])] = r[0]

        # Determine which items to insert
        to_insert = []
        for it in items:
            temp = str(it.get('tempId') or it.get('id'))
            nome = it.get('nome') or it.get('name') or ''
            key = normalize_name_py(nome)
            if key in existing:
                mapping[temp] = str(existing[key])
            else:
                to_insert.append((temp, nome, it.get('especialidade') or it.get('specialty'), it.get('ativo') if 'ativo' in it else True))

        # Bulk insert new professionals
        if to_insert:
            values = []
            placeholders = []
            for t in to_insert:
                placeholders.append('(%%s, %%s, %%s, NOW())')
                values.extend([t[1], t[2], t[3]])
            sql = 'INSERT INTO profissionais (nome, especialidade, ativo, criado_em) VALUES ' + ','.join(placeholders) + ' RETURNING id, nome'
            cur.execute(sql, tuple(values))
            rows = cur.fetchall()
            conn.commit()
            for r in rows:
                nid = r[0]
                nname = r[1]
                existing[normalize_name_py(nname)] = nid

            # map temps
            for t in to_insert:
                temp = str(t[0])
                key = normalize_name_py(t[1])
                if key in existing:
                    mapping[temp] = str(existing[key])
                    created.append({'id': existing[key], 'nome': t[1]})

        return {'mapping': mapping, 'created': created}
    finally:
        try:
            cur.close()
            conn.close()
        except:
            pass

# Helper: perform batch insert for appointments — returns created rows
def do_batch_agendamentos(items):
    created = []
    conn = get_connection()
    cur = conn.cursor()
    try:
        appointment_cols = ensure_agendamento_link_schema(cur, get_table_columns_cached(cur, 'agendamentos'))
        to_insert = []
        errors = []
        for idx, a in enumerate(items, start=1):
            profissional_input = a.get('profissional') or a.get('professional') or a.get('professionalId')
            profissional_id_input = a.get('profissional_id') or a.get('professional_id') or a.get('profissionalId')
            paciente_input = a.get('paciente') or a.get('patient') or a.get('clientName') or ''
            paciente_id_input = a.get('paciente_id') or a.get('patient_id') or a.get('patientId') or a.get('clientId')
            tipo = a.get('tipo_atendimento') or a.get('tipo') or a.get('type') or ''
            data_field = a.get('data') or a.get('date')
            hora_inicio = a.get('hora_inicio') or a.get('hora') or a.get('time')
            hora_fim = a.get('hora_fim') or a.get('endTime') or hora_inicio
            sala_id = normalize_room_id(a.get('sala_id') or a.get('salaId') or a.get('roomId') or a.get('sala'))

            item_errors = []
            if not data_field:
                item_errors.append('data')
            if not hora_inicio:
                item_errors.append('hora_inicio')

            professional, professional_error = resolve_professional_reference(cur, profissional_input, profissional_id_input)
            patient, patient_error = resolve_patient_reference(cur, paciente_input, paciente_id_input)
            room, room_error = resolve_room_reference(cur, sala_id)
            for error in (professional_error, patient_error, room_error):
                if error:
                    item_errors.append(error)

            if item_errors:
                errors.append({'index': idx, 'errors': item_errors})
                continue

            conflict = find_patient_room_conflict(
                cur,
                patient['id'],
                room['id'],
                data_field,
                hora_inicio,
                hora_fim,
                appointment_cols=appointment_cols
            ) or find_pending_patient_room_conflict(
                to_insert,
                patient['id'],
                room['id'],
                data_field,
                hora_inicio,
                hora_fim
            )
            if conflict:
                errors.append({'index': idx, 'errors': [build_patient_room_conflict_error(conflict)]})
                continue

            to_insert.append({
                'profissional': str(professional['id']),
                'profissional_id': professional['id'],
                'paciente': patient['nome'],
                'paciente_id': patient['id'],
                'tipo_atendimento': tipo,
                'data': data_field,
                'hora_inicio': hora_inicio,
                'hora_fim': hora_fim,
                'sala_id': room['id'],
                'sala_nome': room['nome']
            })

        if errors:
            raise ValueError(f"Agendamentos sem vinculo obrigatorio: {errors[:5]}")

        if to_insert:
            insert_columns = [
                'profissional', 'profissional_id', 'paciente', 'paciente_id',
                'tipo_atendimento', 'data', 'hora_inicio', 'hora_fim', 'sala_id', 'criado_em'
            ]
            placeholders = []
            values = []
            for item in to_insert:
                placeholders.append('(%%s,%%s,%%s,%%s,%%s,%%s,%%s,%%s,%%s,NOW())')
                values.extend([
                    item['profissional'],
                    item['profissional_id'],
                    item['paciente'],
                    item['paciente_id'],
                    item['tipo_atendimento'],
                    item['data'],
                    item['hora_inicio'],
                    item['hora_fim'],
                    item['sala_id']
                ])
            sql = (
                f"INSERT INTO agendamentos ({', '.join(insert_columns)}) VALUES "
                + ','.join(placeholders)
                + " RETURNING id, profissional, profissional_id, paciente, paciente_id, tipo_atendimento, data, hora_inicio, hora_fim, sala_id, criado_em"
            )

            cur.execute(sql, tuple(values))
            rows = cur.fetchall()
            for r in rows:
                created.append({
                    'id': r[0],
                    'profissional': r[1],
                    'profissional_id': r[2],
                    'paciente': r[3],
                    'paciente_id': r[4],
                    'tipo_atendimento': r[5],
                    'data': r[6].isoformat() if hasattr(r[6], 'isoformat') else r[6],
                    'hora_inicio': r[7].isoformat() if hasattr(r[7], 'isoformat') else r[7],
                    'hora_fim': r[8].isoformat() if hasattr(r[8], 'isoformat') else r[8],
                    'sala_id': r[9],
                    'criado_em': r[10].isoformat() if r[10] else None
                })

        conn.commit()
        invalidate_agendamentos_list_cache()
    finally:
        try:
            cur.close()
            conn.close()
        except:
            pass

    return created

@app.route("/api/profissionais", methods=["POST"])
def criar_profissional():
    err = require_admin()
    if err:
        return err

    data = request.json
    nome = data.get("nome") or data.get('name')
    especialidade = data.get("especialidade") or data.get('specialty')
    especialidades = data.get('especialidades') or data.get('specialties')
    if not especialidade and isinstance(especialidades, list):
        especialidade = '; '.join(str(item).strip() for item in especialidades if str(item).strip())
    telefone = data.get('telefone') or data.get('phone')
    data_nascimento = data.get('data_nascimento') or data.get('birthdate')
    if data_nascimento:
        try:
            data_nascimento = datetime.fromisoformat(data_nascimento).date()
        except Exception:
            pass
    email = data.get('email')
    numero_conselho = data.get('numero_conselho') or data.get('conselho') or data.get('council_number')
    preferencia = data.get('preferencia') if 'preferencia' in data else (data.get('preference') or data.get('profPreference'))
    contato_emergencia = data.get('contato_emergencia') if 'contato_emergencia' in data else (data.get('emergency_contact') or data.get('emergencyContact'))
    ativo = data.get('ativo') if 'ativo' in data else True

    if not nome:
        return jsonify({"success": False, "error": "nome is required"}), 400

    try:
        conn = get_connection()
        cur = conn.cursor()

        professional_cols = ensure_table_updated_timestamp(cur, 'profissionais', get_table_columns_cached(cur, 'profissionais'))
        ensure_professional_extra_columns(cur, professional_cols)

        insert_columns = ['nome', 'especialidade', 'ativo']
        values = [nome, especialidade, ativo]
        if 'telefone' in professional_cols:
            insert_columns.append('telefone')
            values.append(telefone)
        elif 'phone' in professional_cols:
            insert_columns.append('phone')
            values.append(telefone)
        if 'data_nascimento' in professional_cols:
            insert_columns.append('data_nascimento')
            values.append(data_nascimento)
        elif 'birthdate' in professional_cols:
            insert_columns.append('birthdate')
            values.append(data_nascimento)
        if 'email' in professional_cols:
            insert_columns.append('email')
            values.append(email)
        if 'numero_conselho' in professional_cols:
            insert_columns.append('numero_conselho')
            values.append(numero_conselho)
        elif 'conselho' in professional_cols:
            insert_columns.append('conselho')
            values.append(numero_conselho)
        elif 'council_number' in professional_cols:
            insert_columns.append('council_number')
            values.append(numero_conselho)
        if 'preferencia' in professional_cols:
            insert_columns.append('preferencia')
            values.append(preferencia)
        if 'contato_emergencia' in professional_cols:
            insert_columns.append('contato_emergencia')
            values.append(contato_emergencia)

        placeholders = ', '.join(['%s'] * len(values))
        cur.execute(f"""
            INSERT INTO profissionais ({', '.join(insert_columns)}, criado_em)
            VALUES ({placeholders}, NOW())
            RETURNING id, ativo, criado_em
        """, tuple(values))

        row = cur.fetchone()
        authenticated_user, _auth_error = get_authenticated_user()
        insert_audit_log(
            cur,
            'profissional_criado',
            'profissional',
            entidade_id=row[0],
            entidade_rotulo=nome,
            actor=authenticated_user,
            dados_depois={
                'id': row[0],
                'nome': nome,
                'especialidade': especialidade,
                'telefone': telefone,
                'data_nascimento': serialize_audit_value(data_nascimento),
                'email': email,
                'numero_conselho': numero_conselho,
                'preferencia': preferencia,
                'contato_emergencia': contato_emergencia,
                'ativo': row[1]
            }
        )
        conn.commit()
        cur.close()
        conn.close()

        created = {
            'id': row[0],
            'nome': nome,
            'especialidade': especialidade,
            'telefone': telefone,
            'data_nascimento': data_nascimento.isoformat() if hasattr(data_nascimento, 'isoformat') else data_nascimento,
            'email': email,
            'numero_conselho': numero_conselho,
            'preferencia': preferencia,
            'contato_emergencia': contato_emergencia,
            'ativo': row[1],
            'criado_em': row[2].isoformat() if row[2] else None
        }

        return jsonify({"success": True, "profissional": created})

    except Exception as e:
        print("Erro ao criar profissional:", e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/profissionais", methods=["GET"])
def listar_profissionais():
    auth_check = require_authenticated()
    if auth_check:
        return auth_check

    try:
        conn = get_connection()
        cur = conn.cursor()
        professional_cols = ensure_table_updated_timestamp(cur, 'profissionais', get_table_columns_cached(cur, 'profissionais'))
        ensure_professional_extra_columns(cur, professional_cols)
        conn.commit()

        select_fields = ['id', 'nome', 'especialidade', 'ativo']
        if 'telefone' in professional_cols:
            select_fields.append('telefone')
        elif 'phone' in professional_cols:
            select_fields.append('phone')
        if 'data_nascimento' in professional_cols:
            select_fields.append('data_nascimento')
        elif 'birthdate' in professional_cols:
            select_fields.append('birthdate')
        if 'email' in professional_cols:
            select_fields.append('email')
        if 'numero_conselho' in professional_cols:
            select_fields.append('numero_conselho')
        elif 'conselho' in professional_cols:
            select_fields.append('conselho')
        elif 'council_number' in professional_cols:
            select_fields.append('council_number')
        if 'preferencia' in professional_cols:
            select_fields.append('preferencia')
        if 'contato_emergencia' in professional_cols:
            select_fields.append('contato_emergencia')
        select_fields.append('criado_em')

        query_where = []
        params = []
        active_filter = request.args.get('active') or request.args.get('ativo')
        if active_filter is not None and 'ativo' in professional_cols:
            active_value = str(active_filter).strip().lower() in ('1', 'true', 'sim', 's', 'yes')
            query_where.append('ativo = %s')
            params.append(active_value)
        try:
            limit = int(request.args.get('limit') or 500)
        except Exception:
            limit = 500
        limit = max(1, min(limit, 1000))
        params.append(limit)
        where_sql = f" WHERE {' AND '.join(query_where)}" if query_where else ''

        cur.execute(
            f"SELECT {', '.join(select_fields)} FROM profissionais{where_sql} ORDER BY criado_em DESC NULLS LAST LIMIT %s",
            tuple(params)
        )
        rows = cur.fetchall()
        column_names = [desc[0] for desc in cur.description]

        profissionais = []
        for r in rows:
            profissional = dict(zip(column_names, r))
            if 'birthdate' in profissional and 'data_nascimento' not in profissional:
                profissional['data_nascimento'] = profissional.pop('birthdate')
            if 'data_nascimento' in profissional and hasattr(profissional['data_nascimento'], 'isoformat'):
                profissional['data_nascimento'] = profissional['data_nascimento'].isoformat()
            if 'criado_em' in profissional and hasattr(profissional['criado_em'], 'isoformat'):
                profissional['criado_em'] = profissional['criado_em'].isoformat()
            profissionais.append(profissional)

        cur.close()
        conn.close()

        return jsonify({'success': True, 'profissionais': profissionais})

    except Exception as e:
        print('Erro ao listar profissionais:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/pacientes', methods=['POST'])
def criar_paciente():
    err = require_admin()
    if err:
        return err

    data = request.json or {}
    nome = data.get('nome') or data.get('name') or data.get('patientName') or data.get('nome_paciente')
    data_nascimento = data.get('data_nascimento') or data.get('birthdate') or data.get('dataNascimento')
    endereco = data.get('endereco') or data.get('address')
    telefone = data.get('telefone') or data.get('phone')
    nome_mae = data.get('nome_mae') or data.get('motherName') or data.get('nomeMae')
    nome_pai = data.get('nome_pai') or data.get('fatherName') or data.get('nomePai')
    convenio = data.get('convenio') or data.get('insurance')

    if not nome:
        return jsonify({'success': False, 'error': 'nome is required'}), 400

    if data_nascimento:
        try:
            data_nascimento = datetime.fromisoformat(data_nascimento).date()
        except Exception:
            pass

    try:
        conn = get_connection()
        cur = conn.cursor()
        patient_cols = ensure_table_updated_timestamp(cur, 'pacientes', get_table_columns_cached(cur, 'pacientes'))

        insert_columns = ['nome', 'data_nascimento', 'endereco', 'nome_mae', 'nome_pai', 'convenio', 'telefone']
        values = [nome, data_nascimento, endereco, nome_mae, nome_pai, convenio, telefone]
        if 'ativo' in patient_cols:
            insert_columns.append('ativo')
            values.append(True)

        placeholders = ', '.join(['%s'] * len(values))
        cur.execute(f"""
            INSERT INTO pacientes ({', '.join(insert_columns)}, criado_em)
            VALUES ({placeholders}, NOW())
            RETURNING id, criado_em
        """, tuple(values))

        row = cur.fetchone()
        authenticated_user, _auth_error = get_authenticated_user()
        insert_audit_log(
            cur,
            'paciente_criado',
            'paciente',
            entidade_id=row[0],
            entidade_rotulo=nome,
            actor=authenticated_user,
            dados_depois={
                'id': row[0],
                'nome': nome,
                'telefone': telefone,
                'data_nascimento': serialize_audit_value(data_nascimento),
                'endereco': endereco,
                'nome_mae': nome_mae,
                'nome_pai': nome_pai,
                'convenio': convenio,
                'ativo': True
            }
        )
        conn.commit()
        cur.close()
        conn.close()

        created = {
            'id': row[0],
            'nome': nome,
            'telefone': telefone,
            'data_nascimento': data_nascimento,
            'endereco': endereco,
            'nome_mae': nome_mae,
            'nome_pai': nome_pai,
            'convenio': convenio,
            'criado_em': row[1].isoformat() if row[1] else None
        }

        return jsonify({'success': True, 'paciente': created})

    except Exception as e:
        print('Erro ao criar paciente:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/pacientes', methods=['GET'])
def listar_pacientes():
    authenticated_user, auth_error = get_authenticated_user()
    if auth_error:
        return auth_error

    try:
        conn = get_connection()
        cur = conn.cursor()
        if not user_can_view_patients(cur, authenticated_user):
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Sem permissao para consultar pacientes'}), 403

        patient_cols = ensure_table_updated_timestamp(cur, 'pacientes', get_table_columns_cached(cur, 'pacientes'))
        conn.commit()

        select_fields = ['id', 'nome', 'data_nascimento', 'endereco', 'nome_mae', 'nome_pai', 'convenio', 'telefone']
        has_ativo = 'ativo' in patient_cols
        if has_ativo:
            select_fields.append('ativo')
        select_fields.append('criado_em')

        cur.execute(f"SELECT {', '.join(select_fields)} FROM pacientes ORDER BY criado_em DESC NULLS LAST LIMIT 500")
        rows = cur.fetchall()
        pacientes = []
        for r in rows:
            pacientes.append({
                'id': r[0],
                'nome': r[1],
                'data_nascimento': r[2].isoformat() if hasattr(r[2], 'isoformat') else r[2],
                'endereco': r[3],
                'nome_mae': r[4],
                'nome_pai': r[5],
                'convenio': r[6],
                'telefone': r[7],
                'ativo': r[8] if has_ativo else True,
                'criado_em': r[9].isoformat() if has_ativo and r[9] else (r[8].isoformat() if r[8] else None)
            })
        cur.close()
        conn.close()
        return jsonify({'success': True, 'pacientes': pacientes})
    except Exception as e:
        print('Erro ao listar pacientes:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/pacientes/<int:paciente_id>', methods=['PUT'])
def atualizar_paciente(paciente_id):
    err = require_admin()
    if err:
        return err

    data = request.json or {}
    nome = data.get('nome') or data.get('name') or data.get('patientName') or data.get('nome_paciente')
    data_nascimento = data.get('data_nascimento') or data.get('birthdate') or data.get('dataNascimento')
    endereco = data.get('endereco') or data.get('address')
    telefone = data.get('telefone') or data.get('phone') or data.get('telefone_paciente') or data.get('patientPhone')
    nome_mae = data.get('nome_mae') or data.get('motherName') or data.get('nomeMae')
    nome_pai = data.get('nome_pai') or data.get('fatherName') or data.get('nomePai')
    convenio = data.get('convenio') or data.get('insurance')

    if data_nascimento:
        try:
            data_nascimento = datetime.fromisoformat(data_nascimento).date()
        except Exception:
            pass

    conn = get_connection()
    cur = conn.cursor()
    patient_cols = ensure_table_updated_timestamp(cur, 'pacientes', get_table_columns_cached(cur, 'pacientes'))
    conn.commit()
    cur.close()
    conn.close()

    phone_column_name = None
    if 'telefone' in patient_cols:
        phone_column_name = 'telefone'
    elif 'phone' in patient_cols:
        phone_column_name = 'phone'

    ativo = data.get('ativo') if 'ativo' in data else None

    fields = []
    values = []
    if nome is not None:
        fields.append('nome = %s')
        values.append(nome)
    if data_nascimento is not None:
        fields.append('data_nascimento = %s')
        values.append(data_nascimento)
    if endereco is not None:
        fields.append('endereco = %s')
        values.append(endereco)
    if telefone is not None and phone_column_name:
        fields.append(f'{phone_column_name} = %s')
        values.append(telefone)
    if nome_mae is not None:
        fields.append('nome_mae = %s')
        values.append(nome_mae)
    if nome_pai is not None:
        fields.append('nome_pai = %s')
        values.append(nome_pai)
    if convenio is not None:
        fields.append('convenio = %s')
        values.append(convenio)
    if ativo is not None and 'ativo' in patient_cols:
        fields.append('ativo = %s')
        values.append(ativo)

    if not fields:
        return jsonify({'success': False, 'error': 'Nenhum campo para atualizar'}), 400

    try:
        conn = get_connection()
        cur = conn.cursor()
        if 'atualizado_em' in patient_cols:
            fields.append('atualizado_em = NOW()')
        return_phone_field = phone_column_name or 'NULL AS telefone'
        return_ativo_field = 'ativo' if 'ativo' in patient_cols else 'TRUE AS ativo'
        audit_select_fields = [
            'id',
            'nome',
            'data_nascimento',
            'endereco',
            return_phone_field,
            'nome_mae',
            'nome_pai',
            'convenio',
            return_ativo_field
        ]
        cur.execute(f"SELECT {', '.join(audit_select_fields)} FROM pacientes WHERE id = %s", (paciente_id,))
        before_row = cur.fetchone()
        before_columns = [desc[0] for desc in cur.description]
        before_data = dict(zip(before_columns, before_row)) if before_row else None
        query = f"UPDATE pacientes SET {', '.join(fields)} WHERE id = %s RETURNING id, nome, data_nascimento, endereco, {return_phone_field}, nome_mae, nome_pai, convenio, {return_ativo_field}, criado_em"
        values.append(paciente_id)
        cur.execute(query, tuple(values))
        row = cur.fetchone()
        if not row:
            conn.commit()
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Paciente não encontrado'}), 404

        after_data = {
            'id': row[0],
            'nome': row[1],
            'data_nascimento': serialize_audit_value(row[2]),
            'endereco': row[3],
            'telefone': row[4],
            'nome_mae': row[5],
            'nome_pai': row[6],
            'convenio': row[7],
            'ativo': row[8]
        }
        if before_data:
            before_data['data_nascimento'] = serialize_audit_value(before_data.get('data_nascimento'))
            changes = build_audit_changes(before_data, after_data, exclude_fields={'id'})
            if changes:
                authenticated_user, _auth_error = get_authenticated_user()
                action = 'paciente_inativado' if set(changes.keys()) == {'ativo'} and after_data.get('ativo') is False else 'paciente_atualizado'
                insert_audit_log(
                    cur,
                    action,
                    'paciente',
                    entidade_id=paciente_id,
                    entidade_rotulo=after_data.get('nome') or (before_data or {}).get('nome'),
                    actor=authenticated_user,
                    dados_antes=before_data,
                    dados_depois=after_data,
                    detalhes={'alteracoes': changes}
                )
        conn.commit()
        cur.close()
        conn.close()

        updated = {
            'id': row[0],
            'nome': row[1],
            'data_nascimento': row[2].isoformat() if hasattr(row[2], 'isoformat') else row[2],
            'endereco': row[3],
            'telefone': row[4],
            'nome_mae': row[5],
            'nome_pai': row[6],
            'convenio': row[7],
            'ativo': row[8],
            'criado_em': row[9].isoformat() if row[9] else None
        }

        return jsonify({'success': True, 'paciente': updated})
    except Exception as e:
        print('Erro ao atualizar paciente:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/pacientes/<int:paciente_id>', methods=['DELETE'])
def excluir_paciente(paciente_id):
    err = require_admin()
    if err:
        return err
    try:
        conn = get_connection()
        cur = conn.cursor()
        patient_cols = ensure_table_updated_timestamp(cur, 'pacientes', get_table_columns_cached(cur, 'pacientes'))
        phone_field = 'telefone' if 'telefone' in patient_cols else ('phone AS telefone' if 'phone' in patient_cols else 'NULL AS telefone')
        ativo_field = 'ativo' if 'ativo' in patient_cols else 'TRUE AS ativo'
        cur.execute(
            f"SELECT id, nome, data_nascimento, endereco, {phone_field}, nome_mae, nome_pai, convenio, {ativo_field} FROM pacientes WHERE id = %s",
            (paciente_id,)
        )
        before_row = cur.fetchone()
        before_columns = [desc[0] for desc in cur.description]
        before_data = dict(zip(before_columns, before_row)) if before_row else None
        if 'ativo' in patient_cols:
            if 'atualizado_em' in patient_cols:
                cur.execute('UPDATE pacientes SET ativo = FALSE, atualizado_em = NOW() WHERE id = %s RETURNING id', (paciente_id,))
            else:
                cur.execute('UPDATE pacientes SET ativo = FALSE WHERE id = %s RETURNING id', (paciente_id,))
        else:
            cur.execute('DELETE FROM pacientes WHERE id = %s RETURNING id', (paciente_id,))

        row = cur.fetchone()
        if not row:
            conn.commit()
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Paciente não encontrado'}), 404

        if before_data:
            before_data['data_nascimento'] = serialize_audit_value(before_data.get('data_nascimento'))
            authenticated_user, _auth_error = get_authenticated_user()
            insert_audit_log(
                cur,
                'paciente_excluido',
                'paciente',
                entidade_id=row[0],
                entidade_rotulo=before_data.get('nome'),
                actor=authenticated_user,
                dados_antes=before_data,
                dados_depois={'id': row[0], 'ativo': False} if 'ativo' in patient_cols else None,
                detalhes={'modo': 'inativado' if 'ativo' in patient_cols else 'excluido'}
            )

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({'success': True, 'deleted_id': row[0]})
    except Exception as e:
        print('Erro ao excluir paciente:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/profissionais/batch', methods=['POST'])
def batch_profissionais():
    err = require_admin()
    if err:
        return err

    data = request.json or {}
    items = data.get('professionals') or data.get('profissionais') or []

    try:
        # Normalize and collect names
        normalized_to_temp = {}
        names_set = set()
        for it in items:
            nome = it.get('nome') or it.get('name') or ''
            key = normalize_name_py(nome)
            if key:
                normalized_to_temp[key] = normalized_to_temp.get(key, []) + [it.get('tempId') or it.get('id')]
                names_set.add(nome.lower())

        mapping = {}
        created = []

        conn = get_connection()
        cur = conn.cursor()

        # Find existing by lower(nome)
        if names_set:
            placeholders = ','.join(['%s'] * len(names_set))
            cur.execute(f"SELECT id, nome FROM profissionais WHERE lower(nome) IN ({placeholders})", tuple(names_set))
            rows = cur.fetchall()
            existing = {}
            for r in rows:
                existing[normalize_name_py(r[1])] = r[0]
        else:
            existing = {}

        # Determine which items to insert
        to_insert = []
        for it in items:
            temp = it.get('tempId') or it.get('id')
            nome = it.get('nome') or it.get('name') or ''
            key = normalize_name_py(nome)
            if key in existing:
                mapping[str(temp)] = str(existing[key])
            else:
                to_insert.append((temp, nome, it.get('especialidade') or it.get('specialty'), it.get('ativo') if 'ativo' in it else True))

        # Bulk insert new professionals
        if to_insert:
            values = []
            placeholders = []
            for t in to_insert:
                placeholders.append('(%%s, %%s, %%s, NOW())')
                values.extend([t[1], t[2], t[3]])
            sql = 'INSERT INTO profissionais (nome, especialidade, ativo, criado_em) VALUES ' + ','.join(placeholders) + ' RETURNING id, nome'
            cur.execute(sql, tuple(values))
            rows = cur.fetchall()
            conn.commit()
            for r in rows:
                nid = r[0]
                nname = r[1]
                existing[normalize_name_py(nname)] = nid

            # map temps
            for t in to_insert:
                temp = t[0]
                key = normalize_name_py(t[1])
                if key in existing:
                    mapping[str(temp)] = str(existing[key])
                    created.append({'id': existing[key], 'nome': t[1]})

        cur.close()
        conn.close()

        return jsonify({'success': True, 'mapping': mapping, 'created': created})

    except Exception as e:
        print('Erro batch profissionais:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/agendamentos/batch', methods=['POST'])
def batch_agendamentos():
    auth_check = require_editor_or_admin()
    if auth_check:
        return auth_check

    data = request.json or {}
    items = data.get('agendamentos') or data.get('appointments') or []
    conn = None
    cur = None

    try:
        authenticated_user = None
        try:
            if request.headers.get('Authorization') or session.get('current_user'):
                authenticated_user, _auth_error = get_authenticated_user()
        except Exception:
            authenticated_user = None

        conn = get_connection()
        cur = conn.cursor()
        appointment_cols = ensure_agendamento_link_schema(cur, get_table_columns_cached(cur, 'agendamentos'))
        ensure_agendamento_recurrence_columns(cur, appointment_cols)
        ensure_agendamento_auditoria_table(cur)

        created_by_column = None
        if 'criado_por' in appointment_cols:
            created_by_column = 'criado_por'
        elif 'created_by' in appointment_cols:
            created_by_column = 'created_by'

        has_quantidade_sessoes = 'quantidade_sessoes' in appointment_cols
        has_sala_id = 'sala_id' in appointment_cols
        has_profissional_id = 'profissional_id' in appointment_cols
        has_paciente_id = 'paciente_id' in appointment_cols
        has_recurrence_group = 'recorrencia_grupo_id' in appointment_cols
        has_recurrence_index = 'recorrencia_indice' in appointment_cols
        has_recurrence_total = 'recorrencia_total' in appointment_cols
        has_status = 'status' in appointment_cols
        batch_created_by = data.get('created_by') or data.get('criado_por') or data.get('createdBy')

        to_insert = []
        validation_errors = []
        for item_index, a in enumerate(items, start=1):
            profissional_input = a.get('profissional') or a.get('professional') or a.get('professionalId')
            profissional_id_input = a.get('profissional_id') or a.get('professional_id') or a.get('profissionalId')
            paciente_input = a.get('paciente') or a.get('patient') or a.get('clientName') or ''
            paciente_id_input = a.get('paciente_id') or a.get('patient_id') or a.get('patientId') or a.get('clientId')
            tipo = a.get('tipo_atendimento') or a.get('tipo') or a.get('type') or ''
            data_field = a.get('data') or a.get('date')
            hora_inicio = a.get('hora_inicio') or a.get('hora') or a.get('time')
            hora_fim = a.get('hora_fim') or a.get('endTime') or hora_inicio
            quantidade_sessoes = a.get('quantidade_sessoes')
            sala_id = normalize_room_id(a.get('sala_id') or a.get('salaId') or a.get('roomId') or a.get('sala'))
            created_by = a.get('created_by') or a.get('criado_por') or a.get('createdBy') or batch_created_by
            recurrence_group_id = (
                a.get('recorrencia_grupo_id') or
                a.get('recurrenceGroupId') or
                a.get('repeatGroupId') or
                a.get('recurrence_group_id')
            )
            recurrence_index = a.get('recorrencia_indice') or a.get('recurrenceIndex') or a.get('repeatIndex')
            recurrence_total = a.get('recorrencia_total') or a.get('recurrenceTotal') or a.get('repeatTotal')
            status = normalize_appointment_status(a.get('status'), default='agendado')
            if status is None:
                validation_errors.append({'index': item_index, 'error': 'Status de agendamento invalido'})
                continue
            client_temp_id = a.get('client_temp_id') or a.get('clientTempId') or a.get('tempId')

            item_errors = []
            if not data_field:
                item_errors.append('data is required')
            if not hora_inicio:
                item_errors.append('hora_inicio is required')

            professional, professional_error = resolve_professional_reference(cur, profissional_input, profissional_id_input)
            patient, patient_error = resolve_patient_reference(cur, paciente_input, paciente_id_input)
            room, room_error = resolve_room_reference(cur, sala_id)
            for error in (professional_error, patient_error, room_error):
                if error:
                    item_errors.append(error)

            if item_errors:
                validation_errors.append({'index': item_index, 'errors': item_errors})
                continue

            conflict = find_patient_room_conflict(
                cur,
                patient['id'],
                room['id'],
                data_field,
                hora_inicio,
                hora_fim,
                appointment_cols=appointment_cols
            ) or find_pending_patient_room_conflict(
                to_insert,
                patient['id'],
                room['id'],
                data_field,
                hora_inicio,
                hora_fim
            )
            if conflict:
                validation_errors.append({
                    'index': item_index,
                    'errors': [build_patient_room_conflict_error(conflict)]
                })
                continue

            item = {
                'profissional': str(professional['id']),
                'profissional_id': professional['id'],
                'paciente': patient['nome'],
                'paciente_id': patient['id'],
                'tipo_atendimento': tipo,
                'data': data_field,
                'hora_inicio': hora_inicio,
                'hora_fim': hora_fim,
                'quantidade_sessoes': quantidade_sessoes,
                'sala_id': room['id'],
                'sala_nome': room['nome'],
                'created_by': created_by,
                'recorrencia_grupo_id': recurrence_group_id,
                'recorrencia_indice': recurrence_index,
                'recorrencia_total': recurrence_total,
                'status': status,
                'client_temp_id': client_temp_id
            }
            to_insert.append(item)

        if validation_errors:
            return jsonify({
                'success': False,
                'error': 'Agendamentos precisam ter vinculo valido e nao podem colocar o mesmo paciente em salas diferentes no mesmo horario.',
                'errors': validation_errors[:20]
            }), 400

        created = []
        if to_insert:
            columns = ['profissional', 'paciente', 'tipo_atendimento', 'data', 'hora_inicio', 'hora_fim']
            if has_profissional_id:
                columns.insert(1, 'profissional_id')
            if has_paciente_id:
                columns.insert(3 if has_profissional_id else 2, 'paciente_id')
            if has_quantidade_sessoes:
                columns.append('quantidade_sessoes')
            if has_sala_id:
                columns.append('sala_id')
            if created_by_column:
                columns.append(created_by_column)
            if has_recurrence_group:
                columns.append('recorrencia_grupo_id')
            if has_recurrence_index:
                columns.append('recorrencia_indice')
            if has_recurrence_total:
                columns.append('recorrencia_total')
            if has_status:
                columns.append('status')
            columns.append('criado_em')

            placeholders = []
            values = []
            for item in to_insert:
                row_placeholders = ['%s'] * (len(columns) - 1) + ['NOW()']
                placeholders.append(f"({','.join(row_placeholders)})")
                values.extend([
                    item['profissional'],
                ])
                if has_profissional_id:
                    values.append(item['profissional_id'])
                values.append(item['paciente'])
                if has_paciente_id:
                    values.append(item['paciente_id'])
                values.extend([
                    item['tipo_atendimento'],
                    item['data'],
                    item['hora_inicio'],
                    item['hora_fim']
                ])
                if has_quantidade_sessoes:
                    values.append(item.get('quantidade_sessoes'))
                if has_sala_id:
                    values.append(item.get('sala_id'))
                if created_by_column:
                    values.append(item.get('created_by'))
                if has_recurrence_group:
                    values.append(item.get('recorrencia_grupo_id'))
                if has_recurrence_index:
                    values.append(item.get('recorrencia_indice'))
                if has_recurrence_total:
                    values.append(item.get('recorrencia_total'))
                if has_status:
                    values.append(item.get('status'))

            returning_fields = ['id', 'profissional', 'paciente', 'tipo_atendimento', 'data', 'hora_inicio', 'hora_fim']
            if has_profissional_id:
                returning_fields.insert(2, 'profissional_id')
            if has_paciente_id:
                returning_fields.insert(4 if has_profissional_id else 3, 'paciente_id')
            if has_quantidade_sessoes:
                returning_fields.append('quantidade_sessoes')
            if has_sala_id:
                returning_fields.append('sala_id')
            if created_by_column:
                returning_fields.append(created_by_column)
            if has_recurrence_group:
                returning_fields.append('recorrencia_grupo_id')
            if has_recurrence_index:
                returning_fields.append('recorrencia_indice')
            if has_recurrence_total:
                returning_fields.append('recorrencia_total')
            if has_status:
                returning_fields.append('status')
            returning_fields.append('criado_em')

            sql = (
                f"INSERT INTO agendamentos ({', '.join(columns)}) VALUES "
                + ','.join(placeholders)
                + f" RETURNING {', '.join(returning_fields)}"
            )
            cur.execute(sql, tuple(values))
            rows = cur.fetchall()

            for item, r in zip(to_insert, rows):
                idx = 0
                row_id = r[idx]; idx += 1
                row_profissional = r[idx]; idx += 1
                row_profissional_id = item.get('profissional_id')
                if has_profissional_id:
                    row_profissional_id = r[idx]
                    idx += 1
                row_paciente = r[idx]; idx += 1
                row_paciente_id = item.get('paciente_id')
                if has_paciente_id:
                    row_paciente_id = r[idx]
                    idx += 1
                row_tipo = r[idx]; idx += 1
                row_data = r[idx]; idx += 1
                row_hora_inicio = r[idx]; idx += 1
                row_hora_fim = r[idx]; idx += 1
                row_quantidade_sessoes = None
                if has_quantidade_sessoes:
                    row_quantidade_sessoes = r[idx]
                    idx += 1
                row_sala_id = None
                if has_sala_id:
                    row_sala_id = r[idx]
                    idx += 1
                row_created_by = item.get('created_by')
                if created_by_column:
                    row_created_by = r[idx]
                    idx += 1
                row_recurrence_group_id = None
                if has_recurrence_group:
                    row_recurrence_group_id = r[idx]
                    idx += 1
                row_recurrence_index = None
                if has_recurrence_index:
                    row_recurrence_index = r[idx]
                    idx += 1
                row_recurrence_total = None
                if has_recurrence_total:
                    row_recurrence_total = r[idx]
                    idx += 1
                row_status = item.get('status') or 'agendado'
                if has_status:
                    row_status = normalize_appointment_status(r[idx], default='agendado')
                    idx += 1
                row_criado_em = r[idx]

                insert_agendamento_audit(
                    cur,
                    row_id,
                    'criado',
                    build_audit_user(item, authenticated_user, row_created_by),
                    status_anterior=None,
                    status_novo=row_status,
                    detalhes={
                        'profissional': row_profissional,
                        'profissional_id': row_profissional_id,
                        'paciente': row_paciente,
                        'paciente_id': row_paciente_id,
                        'tipo_atendimento': row_tipo,
                        'data': row_data.isoformat() if hasattr(row_data, 'isoformat') else row_data,
                        'hora_inicio': row_hora_inicio.isoformat() if hasattr(row_hora_inicio, 'isoformat') else row_hora_inicio,
                        'hora_fim': row_hora_fim.isoformat() if hasattr(row_hora_fim, 'isoformat') else row_hora_fim,
                        'quantidade_sessoes': row_quantidade_sessoes,
                        'sala_id': row_sala_id,
                        'recorrencia_grupo_id': row_recurrence_group_id,
                        'recorrencia_indice': row_recurrence_index,
                        'recorrencia_total': row_recurrence_total,
                        'status': row_status
                    }
                )

                created_item = {
                    'id': row_id,
                    'profissional': row_profissional,
                    'profissional_id': row_profissional_id,
                    'paciente': row_paciente,
                    'paciente_id': row_paciente_id,
                    'tipo_atendimento': row_tipo,
                    'data': row_data.isoformat() if hasattr(row_data, 'isoformat') else row_data,
                    'hora_inicio': row_hora_inicio.isoformat() if hasattr(row_hora_inicio, 'isoformat') else row_hora_inicio,
                    'hora_fim': row_hora_fim.isoformat() if hasattr(row_hora_fim, 'isoformat') else row_hora_fim,
                    'sala_id': row_sala_id,
                    'created_by': row_created_by,
                    'status': row_status,
                    'criado_em': row_criado_em.isoformat() if row_criado_em else None,
                    'client_temp_id': item.get('client_temp_id')
                }
                if has_quantidade_sessoes:
                    created_item['quantidade_sessoes'] = row_quantidade_sessoes
                if has_recurrence_group:
                    created_item['recorrencia_grupo_id'] = row_recurrence_group_id
                if has_recurrence_index:
                    created_item['recorrencia_indice'] = row_recurrence_index
                if has_recurrence_total:
                    created_item['recorrencia_total'] = row_recurrence_total
                created.append(created_item)

            conn.commit()
            invalidate_agendamentos_list_cache()

        cur.close()
        conn.close()

        return jsonify({'success': True, 'agendamentos': created})

    except Exception as e:
        print('Erro batch agendamentos:', e)
        try:
            if conn:
                conn.rollback()
        except:
            pass
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


# Create helpful index to speed up normalized-name lookups (uses unaccent if available)
@app.route('/api/create_indexes', methods=['POST'])
def create_indexes():
    auth_check = require_admin()
    if auth_check:
        return auth_check

    try:
        conn = get_connection()
        cur = conn.cursor()
        ensure_remarque_table_cached(cur)
        ensure_lista_espera_table(cur)
        ensure_agendamento_link_schema(cur)
        ensure_agendamento_auditoria_table(cur)
        ensure_audit_logs_table(cur)
        ensure_usuarios_audit_columns(cur)
        try:
            cur.execute("CREATE EXTENSION IF NOT EXISTS unaccent")
            conn.commit()
        except Exception:
            # ignore if cannot create extension
            conn.rollback()
        try:
            cur.execute("CREATE INDEX IF NOT EXISTS idx_profissionais_nome_unaccent_lower ON profissionais (lower(unaccent(nome)))")
            conn.commit()
        except Exception:
            # fallback to lower(nome)
            conn.rollback()
            cur.execute("CREATE INDEX IF NOT EXISTS idx_profissionais_nome_lower ON profissionais (lower(nome))")
            conn.commit()
        create_performance_indexes(cur)
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'success': True, 'message': 'Indexes created or already exist'})
    except Exception as e:
        print('Erro creating indexes:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sync', methods=['POST'])
def full_sync():
    auth_check = require_editor_or_admin()
    if auth_check:
        return auth_check

    data = request.json or {}
    professionals_payload = data.get('professionals', [])
    appointments_payload = data.get('appointments', [])

    try:
        conn = get_connection()
        cur = conn.cursor()
        appointment_cols = ensure_agendamento_link_schema(cur, get_table_columns_cached(cur, 'agendamentos'))

        # Prepare existing professionals map by normalized name
        names_set = set()
        for p in professionals_payload:
            name = (p.get('name') or p.get('nome') or '').strip()
            if name:
                names_set.add(name.lower())

        existing = {}
        if names_set:
            try:
                placeholders = ','.join(['%s'] * len(names_set))
                cur.execute(f"SELECT id, nome, especialidade, ativo FROM profissionais WHERE lower(unaccent(nome)) IN ({placeholders})", tuple(names_set))
            except Exception:
                placeholders = ','.join(['%s'] * len(names_set))
                cur.execute(f"SELECT id, nome, especialidade, ativo FROM profissionais WHERE lower(nome) IN ({placeholders})", tuple(names_set))
            rows = cur.fetchall()
            for r in rows:
                existing[normalize_name_py(r[1])] = {'id': r[0], 'nome': r[1], 'especialidade': r[2], 'ativo': r[3]}

        mapping = {}
        created = []
        updated = []

        # Upsert professionals (create if not exists, update if changed)
        for p in professionals_payload:
            temp = str(p.get('id') or p.get('tempId') or '')
            name = (p.get('name') or p.get('nome') or '').strip()
            specialty = p.get('specialty') or p.get('especialidade') or ''
            ativo = p.get('active') if 'active' in p else (p.get('ativo') if 'ativo' in p else True)
            if not name:
                continue
            key = normalize_name_py(name)
            if key in existing:
                sid = existing[key]['id']
                mapping[temp] = str(sid)

                changes = {}
                if specialty and (existing[key].get('especialidade') or '') != specialty:
                    changes['especialidade'] = specialty
                if existing[key].get('ativo') != ativo:
                    changes['ativo'] = ativo

                if changes:
                    set_clause = ', '.join([f"{k} = %s" for k in changes.keys()])
                    params = tuple(list(changes.values()) + [sid])
                    cur.execute(f"UPDATE profissionais SET {set_clause} WHERE id = %s", params)
                    updated.append({'id': sid, 'changes': changes})
            else:
                cur.execute("INSERT INTO profissionais (nome, especialidade, ativo, criado_em) VALUES (%s,%s,%s,NOW()) RETURNING id, nome", (name, specialty, ativo))
                row = cur.fetchone()
                conn.commit()
                sid = row[0]
                mapping[temp] = str(sid)
                created.append({'id': sid, 'nome': row[1]})
                existing[key] = {'id': sid, 'nome': row[1], 'especialidade': specialty, 'ativo': ativo}

        # Insert appointments if they don't already exist
        created_appointments = []
        appointment_errors = []
        for a in appointments_payload:
            raw_prof = a.get('professionalId') or a.get('profissional') or a.get('professional') or a.get('professionalId')
            prof_id = mapping.get(str(raw_prof)) or str(raw_prof)
            paciente = a.get('clientName') or a.get('paciente') or ''
            paciente_id = a.get('paciente_id') or a.get('patient_id') or a.get('patientId') or a.get('clientId')
            sala_id = normalize_room_id(a.get('sala_id') or a.get('salaId') or a.get('roomId') or a.get('sala'))
            data_field = a.get('date') or a.get('data')
            hora_inicio = a.get('time') or a.get('hora_inicio') or a.get('hora')
            hora_fim = a.get('hora_fim') or a.get('endTime') or hora_inicio

            if not data_field or not hora_inicio:
                appointment_errors.append({'appointment': a, 'error': 'data e hora_inicio sao obrigatorios'})
                continue

            professional_ref, professional_error = resolve_professional_reference(cur, prof_id)
            patient_ref, patient_error = resolve_patient_reference(cur, paciente, paciente_id)
            room_ref, room_error = resolve_room_reference(cur, sala_id)
            if professional_error or patient_error or room_error:
                appointment_errors.append({
                    'appointment': a,
                    'error': professional_error or patient_error or room_error
                })
                continue

            prof_id = str(professional_ref['id'])
            paciente = patient_ref['nome']
            paciente_id = patient_ref['id']
            sala_id = room_ref['id']

            # skip duplicates by exact match on profissional, paciente, data, hora_inicio
            cur.execute('SELECT id FROM agendamentos WHERE profissional_id=%s AND paciente_id=%s AND data=%s AND hora_inicio=%s LIMIT 1', (professional_ref['id'], paciente_id, data_field, hora_inicio))
            if cur.fetchone():
                continue

            conflict = find_patient_room_conflict(
                cur,
                paciente_id,
                sala_id,
                data_field,
                hora_inicio,
                hora_fim,
                appointment_cols=appointment_cols
            )
            if conflict:
                appointment_errors.append({
                    'appointment': a,
                    'error': build_patient_room_conflict_error(conflict)
                })
                continue

            cur.execute(
                """
                INSERT INTO agendamentos (
                    profissional, profissional_id, paciente, paciente_id,
                    tipo_atendimento, data, hora_inicio, hora_fim, sala_id, criado_em
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
                RETURNING id
                """,
                (
                    prof_id, professional_ref['id'], paciente, paciente_id,
                    a.get('type') or a.get('tipo') or a.get('tipo_atendimento') or '',
                    data_field, hora_inicio, hora_fim, sala_id
                )
            )
            row = cur.fetchone()
            conn.commit()
            created_appointments.append({'id': row[0], 'profissional': prof_id, 'profissional_id': professional_ref['id'], 'paciente': paciente, 'paciente_id': paciente_id, 'sala_id': sala_id})

        cur.close()
        conn.close()

        return jsonify({'success': True, 'mapping': mapping, 'created_professionals': created, 'updated_professionals': updated, 'created_appointments': created_appointments, 'appointment_errors': appointment_errors[:20]})
    except Exception as e:
        print('Erro sync:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/import/upload', methods=['POST'])
def import_upload():
    err = require_admin()
    if err:
        return err

    # Accepts multipart/form-data with file and optional target=server|local
    target = (request.form.get('target') or request.args.get('target') or 'server').lower()
    f = request.files.get('file')
    if not f:
        return jsonify({'success': False, 'error': 'file is required (multipart form field name=file)'}), 400

    try:
        # Determine file type by filename/mimetype
        filename = (f.filename or '').lower()
        ext = os.path.splitext(filename)[1] if filename else ''

        new_professionals = []
        new_appointments = []

        parsed_from_csv = False
        # CSV fallback: if the user uploaded a CSV file, parse it server-side without openpyxl
        if ext == '.csv' or (hasattr(f, 'mimetype') and f.mimetype and 'csv' in f.mimetype):
            try:
                import io, csv
                text = f.read().decode('utf-8', errors='replace')
                reader = csv.DictReader(io.StringIO(text))

                prof_map = {}
                for row in reader:
                    # helper to retrieve a field by multiple possible headers
                    def get_row(keys):
                        for k in keys:
                            if k in row and row[k] and str(row[k]).strip():
                                return str(row[k]).strip()
                        return ''

                    professional = get_row(['professional', 'profissional', 'profissional_name', 'profissional_nome', 'prof', 'name'])
                    patient = get_row(['paciente', 'patient', 'client', 'cliente'])
                    date = get_row(['data', 'date'])
                    time = get_row(['hora_inicio', 'hora', 'time'])
                    tipo = get_row(['tipo_atendimento', 'tipo', 'type']) or 'clinica'
                    observations = row.get('observations') or row.get('observacoes') or ''

                    if not professional:
                        # try first column as professional if no header matched
                        try:
                            first_key = next(iter(row.keys()))
                            professional = str(row[first_key]).strip()
                        except Exception:
                            professional = ''

                    if professional:
                        if professional not in prof_map:
                            pid = f"temp-csv-{len(prof_map)}"
                            prof_map[professional] = pid
                            new_professionals.append({'id': pid, 'name': professional, 'specialty': row.get('especialidade') or row.get('specialty') or 'Terapeuta ABA'})

                        if date and time and patient:
                            new_appointments.append({
                                'id': f"temp-a-{len(new_appointments)}",
                                'professionalId': prof_map[professional],
                                'date': date,
                                'time': time,
                                'clientName': patient,
                                'type': tipo,
                                'observations': observations
                            })
                parsed_from_csv = True
            except Exception as e:
                return jsonify({'success': False, 'error': f'Erro ao processar CSV: {str(e)}'}), 500
            finally:
                try:
                    f.seek(0)
                except Exception:
                    pass

        # If not CSV, require openpyxl for .xlsx files
        if not parsed_from_csv:
            try:
                from openpyxl import load_workbook
            except Exception:
                return jsonify({'success': False, 'error': 'Para arquivos .xlsx é necessário instalar openpyxl no servidor (pip install openpyxl) ou escolha "Somente nesta máquina" para processar localmente.'}), 500

            wb = load_workbook(filename=f, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                name_cell = ws['B1'].value if 'B1' in ws else None
                if not name_cell:
                    continue
                professional_name = str(name_cell).strip()
                lowname = professional_name.lower()
                if not professional_name or any(x in lowname for x in ['horario', 'horário', 'hora', 'time', 'sheet1', 'planilha1']) or len(professional_name) < 2:
                    continue

                # Read days C2..H2
                days = []
                for col in ['C','D','E','F','G','H']:
                    cell_ref = f"{col}2"
                    v = ws[cell_ref].value if cell_ref in ws else None
                    if v:
                        days.append(str(v).strip())

                # Read times B3..B16
                times = []
                for row in range(3,17):
                    cell_ref = f"B{row}"
                    v = ws[cell_ref].value if cell_ref in ws else None
                    if v:
                        t = str(v).strip()
                        if 'h' in t:
                            t = t.replace('h', ':00')
                        times.append(t)
                    else:
                        hour = 7 + (row - 3)
                        if hour <= 19:
                            times.append(f"{str(hour).zfill(2)}:00")

                dayColumns = ['C','D','E','F','G','H']
                dayNames = ['Segunda','Terça','Quarta','Quinta','Sexta','Sábado']

                # ensure professional exists in list
                prof_obj = None
                for p in new_professionals:
                    if p['name'].lower()==professional_name.lower():
                        prof_obj = p
                        break
                if not prof_obj:
                    prof_obj = {'id': f"temp-{len(new_professionals)}", 'name': professional_name, 'specialty': 'Terapeuta ABA'}
                    new_professionals.append(prof_obj)

                for colIndex in range(len(dayColumns)):
                    col = dayColumns[colIndex]
                    dayName = dayNames[colIndex]
                    for rowIndex in range(len(times)):
                        row = rowIndex + 3
                        cell_ref = f"{col}{row}"
                        v = ws[cell_ref].value if cell_ref in ws else None
                        if not v:
                            continue
                        cellValue = str(v).strip()
                        if not cellValue:
                            continue
                        # detect type by content
                        c = cellValue.lower()
                        if 'sup' in c or 'clarissa' in c or 'reinaldo' in c:
                            appointmentType = 'supervisao'
                        elif 'trein' in c or 'reun' in c:
                            appointmentType = 'treinamento'
                        elif 'almoço' in c or 'almoco' in c or 'deslocamento' in c or 'bloqueado' in c:
                            appointmentType = 'bloqueado'
                        else:
                            appointmentType = 'clinica'

                        appointmentDate = None
                        try:
                            from datetime import date, timedelta, datetime
                            today = date.today()
                            weekday_map = {'Segunda':1,'Terça':2,'Quarta':3,'Quinta':4,'Sexta':5,'Sábado':6,'Domingo':0}
                            targetDay = weekday_map.get(dayName, None)
                            if targetDay is not None:
                                diff = targetDay - today.weekday()
                                appointmentDate = today + timedelta(days=diff)
                                appointmentDate = appointmentDate.isoformat()
                        except Exception:
                            appointmentDate = None

                        if appointmentDate:
                            appointment = {
                                'id': f"temp-a-{len(new_appointments)}",
                                'professionalId': prof_obj['id'],
                                'date': appointmentDate,
                                'time': times[rowIndex],
                                'clientName': cellValue,
                                'type': appointmentType,
                                'observations': ''
                            }
                            new_appointments.append(appointment)

        # If target local, return parsed data without saving
        if target == 'local':
            return jsonify({'success': True, 'parsed': True, 'professionals': new_professionals, 'appointments': new_appointments, 'message': 'Parsed and returned (local only)'})

        # else save to server using batch helpers
        prof_items = [{'tempId': p['id'], 'nome': p['name'], 'especialidade': p.get('specialty',''), 'ativo': True} for p in new_professionals]
        prof_result = do_batch_profissionais(prof_items)
        mapping = prof_result.get('mapping', {})

        # Map appointments professional ids
        appts_to_save = []
        for a in new_appointments:
            mapped = mapping.get(str(a['professionalId'])) or a['professionalId']
            appts_to_save.append({'profissional': mapped, 'paciente': a['clientName'], 'tipo_atendimento': a['type'], 'data': a['date'], 'hora_inicio': a['time'], 'hora_fim': a['time']})

        try:
            ag_result = do_batch_agendamentos(appts_to_save)
        except ValueError as e:
            return jsonify({'success': False, 'error': str(e)}), 400

        return jsonify({'success': True, 'parsed': True, 'professionals_count': len(new_professionals), 'appointments_count': len(new_appointments), 'mapping': mapping, 'created_professionals': prof_result.get('created', []), 'created_appointments': len(ag_result)})

    except Exception as e:
        print('Erro import upload:', e)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/import/benchmark', methods=['POST'])
def import_benchmark():
    err = require_admin()
    if err:
        return err

    data = request.json or {}
    n = int(data.get('n_professionals', 100))
    m = int(data.get('appointments_per_professional', 20))

    import time
    start = time.perf_counter()

    # generate professionals
    profs = []
    for i in range(n):
        profs.append({'tempId': f't{i}', 'nome': f'Perf Test {i}', 'especialidade': 'Teste', 'ativo': True})

    mid1 = time.perf_counter()
    prof_result = do_batch_profissionais(profs)
    mid2 = time.perf_counter()

    # generate appointments
    appts = []
    for i in range(n):
        pid = prof_result.get('mapping', {}).get(f't{i}', f'Perf Test {i}')
        for j in range(m):
            appts.append({'profissional': pid, 'paciente': f'Paciente {i}-{j}', 'tipo_atendimento': 'clinica', 'data': '2025-12-25', 'hora_inicio': '09:00', 'hora_fim': '09:00'})

    mid3 = time.perf_counter()
    ag_result = do_batch_agendamentos(appts)
    end = time.perf_counter()

    return jsonify({'success': True, 'prof_inserts': len(prof_result.get('created', [])), 'appt_inserts': len(ag_result), 'times': {'generate': mid1-start, 'prof_batch': mid2-mid1, 'build_appts': mid3-mid2, 'appt_batch': end-mid3, 'total': end-start}})


@app.route('/api/profissionais/<int:prof_id>', methods=['PUT'])
def atualizar_profissional(prof_id):
    err = require_admin()
    if err:
        return err

    data = request.json
    nome = data.get('nome') or data.get('name')
    especialidade = data.get('especialidade') or data.get('specialty')
    especialidades = data.get('especialidades') or data.get('specialties')
    if especialidade is None and isinstance(especialidades, list):
        especialidade = '; '.join(str(item).strip() for item in especialidades if str(item).strip())
    telefone = data.get('telefone') or data.get('phone')
    data_nascimento = data.get('data_nascimento') or data.get('birthdate')
    if data_nascimento:
        try:
            data_nascimento = datetime.fromisoformat(data_nascimento).date()
        except Exception:
            pass
    email = data.get('email')
    numero_conselho = data.get('numero_conselho') or data.get('conselho') or data.get('council_number')
    preferencia = data.get('preferencia') if 'preferencia' in data else (data.get('preference') or data.get('profPreference'))
    contato_emergencia = data.get('contato_emergencia') if 'contato_emergencia' in data else (data.get('emergency_contact') or data.get('emergencyContact'))
    ativo = data.get('ativo') if 'ativo' in data else None

    try:
        conn = get_connection()
        cur = conn.cursor()
        professional_cols = ensure_table_updated_timestamp(cur, 'profissionais', get_table_columns_cached(cur, 'profissionais'))
        ensure_professional_extra_columns(cur, professional_cols)
        audit_select_fields = ['id', 'nome', 'especialidade', 'ativo']
        if 'telefone' in professional_cols:
            audit_select_fields.append('telefone')
        elif 'phone' in professional_cols:
            audit_select_fields.append('phone AS telefone')
        if 'data_nascimento' in professional_cols:
            audit_select_fields.append('data_nascimento')
        elif 'birthdate' in professional_cols:
            audit_select_fields.append('birthdate AS data_nascimento')
        if 'email' in professional_cols:
            audit_select_fields.append('email')
        if 'numero_conselho' in professional_cols:
            audit_select_fields.append('numero_conselho')
        elif 'conselho' in professional_cols:
            audit_select_fields.append('conselho AS numero_conselho')
        elif 'council_number' in professional_cols:
            audit_select_fields.append('council_number AS numero_conselho')
        if 'preferencia' in professional_cols:
            audit_select_fields.append('preferencia')
        if 'contato_emergencia' in professional_cols:
            audit_select_fields.append('contato_emergencia')
        cur.execute(f"SELECT {', '.join(audit_select_fields)} FROM profissionais WHERE id = %s", (prof_id,))
        before_row = cur.fetchone()
        before_columns = [desc[0] for desc in cur.description]
        before_data = dict(zip(before_columns, before_row)) if before_row else None

        fields = []
        values = []
        if nome is not None:
            fields.append('nome = %s')
            values.append(nome)
        if especialidade is not None:
            fields.append('especialidade = %s')
            values.append(especialidade)
        if telefone is not None:
            if 'telefone' in professional_cols:
                fields.append('telefone = %s')
                values.append(telefone)
            elif 'phone' in professional_cols:
                fields.append('phone = %s')
                values.append(telefone)
        if data_nascimento is not None:
            if 'data_nascimento' in professional_cols:
                fields.append('data_nascimento = %s')
                values.append(data_nascimento)
            elif 'birthdate' in professional_cols:
                fields.append('birthdate = %s')
                values.append(data_nascimento)
        if email is not None and 'email' in professional_cols:
            fields.append('email = %s')
            values.append(email)
        if numero_conselho is not None:
            if 'numero_conselho' in professional_cols:
                fields.append('numero_conselho = %s')
                values.append(numero_conselho)
            elif 'conselho' in professional_cols:
                fields.append('conselho = %s')
                values.append(numero_conselho)
            elif 'council_number' in professional_cols:
                fields.append('council_number = %s')
                values.append(numero_conselho)
        if preferencia is not None and 'preferencia' in professional_cols:
            fields.append('preferencia = %s')
            values.append(preferencia)
        if contato_emergencia is not None and 'contato_emergencia' in professional_cols:
            fields.append('contato_emergencia = %s')
            values.append(contato_emergencia)
        if ativo is not None:
            fields.append('ativo = %s')
            values.append(ativo)

        if len(fields) == 0:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Nada para atualizar'}), 400

        if 'atualizado_em' in professional_cols:
            fields.append('atualizado_em = NOW()')
        values.append(prof_id)
        sql = f"UPDATE profissionais SET {', '.join(fields)} WHERE id = %s"
        cur.execute(sql, tuple(values))

        # Return the updated record for frontend synchronization
        select_fields = ['id', 'nome', 'especialidade', 'ativo']
        if 'telefone' in professional_cols:
            select_fields.append('telefone')
        elif 'phone' in professional_cols:
            select_fields.append('phone')
        if 'data_nascimento' in professional_cols:
            select_fields.append('data_nascimento')
        elif 'birthdate' in professional_cols:
            select_fields.append('birthdate')
        if 'email' in professional_cols:
            select_fields.append('email')
        if 'numero_conselho' in professional_cols:
            select_fields.append('numero_conselho')
        elif 'conselho' in professional_cols:
            select_fields.append('conselho')
        elif 'council_number' in professional_cols:
            select_fields.append('council_number')
        if 'preferencia' in professional_cols:
            select_fields.append('preferencia')
        if 'contato_emergencia' in professional_cols:
            select_fields.append('contato_emergencia')

        cur.execute(f"SELECT {', '.join(select_fields)} FROM profissionais WHERE id = %s", (prof_id,))
        updated_row = cur.fetchone()
        if not updated_row:
            conn.commit()
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Profissional nao encontrado'}), 404
        column_names = [desc[0] for desc in cur.description]
        updated_prof = dict(zip(column_names, updated_row))
        if 'birthdate' in updated_prof and 'data_nascimento' not in updated_prof:
            updated_prof['data_nascimento'] = updated_prof.pop('birthdate')
        if 'data_nascimento' in updated_prof and hasattr(updated_prof['data_nascimento'], 'isoformat'):
            updated_prof['data_nascimento'] = updated_prof['data_nascimento'].isoformat()
        after_data = dict(updated_prof)
        if before_data:
            if 'data_nascimento' in before_data:
                before_data['data_nascimento'] = serialize_audit_value(before_data.get('data_nascimento'))
            changes = build_audit_changes(before_data, after_data, exclude_fields={'id'})
            if changes:
                authenticated_user, _auth_error = get_authenticated_user()
                insert_audit_log(
                    cur,
                    'profissional_atualizado',
                    'profissional',
                    entidade_id=prof_id,
                    entidade_rotulo=after_data.get('nome') or before_data.get('nome'),
                    actor=authenticated_user,
                    dados_antes=before_data,
                    dados_depois=after_data,
                    detalhes={'alteracoes': changes}
                )

        conn.commit()

        cur.close()
        conn.close()

        return jsonify({'success': True, 'profissional': updated_prof})

    except Exception as e:
        print('Erro ao atualizar profissional:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/profissionais/<int:prof_id>', methods=['DELETE'])
def deletar_profissional(prof_id):
    # Require admin credentials to delete a professional
    admin_check = require_admin()
    if admin_check:
        return admin_check

    try:
        authenticated_user, _auth_error = get_authenticated_user()
        conn = get_connection()
        cur = conn.cursor()
        professional_cols = get_table_columns_cached(cur, 'profissionais')
        select_fields = ['id', 'nome', 'especialidade', 'ativo']
        if 'telefone' in professional_cols:
            select_fields.append('telefone')
        elif 'phone' in professional_cols:
            select_fields.append('phone AS telefone')
        if 'data_nascimento' in professional_cols:
            select_fields.append('data_nascimento')
        elif 'birthdate' in professional_cols:
            select_fields.append('birthdate AS data_nascimento')
        if 'email' in professional_cols:
            select_fields.append('email')
        cur.execute(f"SELECT {', '.join(select_fields)} FROM profissionais WHERE id = %s", (prof_id,))
        before_row = cur.fetchone()
        before_columns = [desc[0] for desc in cur.description]
        before_data = dict(zip(before_columns, before_row)) if before_row else None
        cur.execute('DELETE FROM profissionais WHERE id = %s RETURNING id', (prof_id,))
        deleted_row = cur.fetchone()
        if before_data:
            if 'data_nascimento' in before_data:
                before_data['data_nascimento'] = serialize_audit_value(before_data.get('data_nascimento'))
            insert_audit_log(
                cur,
                'profissional_excluido',
                'profissional',
                entidade_id=prof_id,
                entidade_rotulo=before_data.get('nome'),
                actor=authenticated_user,
                dados_antes=before_data
            )
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'success': True, 'deleted_id': deleted_row[0] if deleted_row else None})
    except Exception as e:
        print('Erro ao deletar profissional:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/agendamentos', methods=['POST'])
def criar_agendamento():
    auth_check = require_editor_or_admin()
    if auth_check:
        return auth_check

    data = request.json
    profissional = data.get('profissional') or data.get('professional') or data.get('professionalId')
    profissional_id = data.get('profissional_id') or data.get('professional_id') or data.get('profissionalId')
    paciente = data.get('paciente') or data.get('patient') or data.get('clientName')
    paciente_id = data.get('paciente_id') or data.get('patient_id') or data.get('patientId') or data.get('clientId')
    tipo = data.get('tipo_atendimento') or data.get('tipo') or data.get('type')
    data_field = data.get('data') or data.get('date')
    hora_inicio = data.get('hora_inicio') or data.get('hora') or data.get('time')
    hora_fim = data.get('hora_fim') or data.get('endTime') or hora_inicio
    quantidade_sessoes = data.get('quantidade_sessoes')
    sala_id = normalize_room_id(data.get('sala_id') or data.get('salaId') or data.get('roomId') or data.get('sala'))
    created_by = data.get('criado_por') or data.get('created_by') or data.get('createdBy') or data.get('criador')
    recurrence_group_id = (
        data.get('recorrencia_grupo_id') or
        data.get('recurrenceGroupId') or
        data.get('repeatGroupId') or
        data.get('recurrence_group_id')
    )
    recurrence_index = data.get('recorrencia_indice') or data.get('recurrenceIndex') or data.get('repeatIndex')
    recurrence_total = data.get('recorrencia_total') or data.get('recurrenceTotal') or data.get('repeatTotal')
    status = normalize_appointment_status(data.get('status'), default='agendado')
    if status is None:
        return jsonify({'success': False, 'error': 'Status de agendamento invalido'}), 400

    if not profissional and not profissional_id:
        return jsonify({'success': False, 'error': 'Selecione um profissional cadastrado.'}), 400
    if not paciente and not paciente_id:
        return jsonify({'success': False, 'error': 'Selecione um paciente cadastrado.'}), 400
    if sala_id is None:
        return jsonify({'success': False, 'error': 'Selecione uma sala cadastrada.'}), 400
    if not data_field:
        return jsonify({'success': False, 'error': 'data is required'}), 400
    if not hora_inicio:
        return jsonify({'success': False, 'error': 'hora_inicio is required'}), 400

    try:
        authenticated_user = None
        try:
            if request.headers.get('Authorization') or session.get('current_user'):
                authenticated_user, _auth_error = get_authenticated_user()
        except Exception:
            authenticated_user = None

        conn = get_connection()
        cur = conn.cursor()

        appointment_cols = ensure_agendamento_link_schema(cur, get_table_columns_cached(cur, 'agendamentos'))
        ensure_agendamento_recurrence_columns(cur, appointment_cols)
        created_by_column = None
        if 'criado_por' in appointment_cols:
            created_by_column = 'criado_por'
        elif 'created_by' in appointment_cols:
            created_by_column = 'created_by'

        professional_ref, professional_error = resolve_professional_reference(cur, profissional, profissional_id)
        if professional_error:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': professional_error}), 400
        patient_ref, patient_error = resolve_patient_reference(cur, paciente, paciente_id)
        if patient_error:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': patient_error}), 400
        room_ref, room_error = resolve_room_reference(cur, sala_id)
        if room_error:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': room_error}), 400

        profissional = str(professional_ref['id'])
        profissional_id = professional_ref['id']
        paciente = patient_ref['nome']
        paciente_id = patient_ref['id']
        sala_id = room_ref['id']

        conflict = find_patient_room_conflict(
            cur,
            paciente_id,
            sala_id,
            data_field,
            hora_inicio,
            hora_fim,
            appointment_cols=appointment_cols
        )
        if conflict:
            cur.close()
            conn.close()
            return jsonify({
                'success': False,
                'error': build_patient_room_conflict_error(conflict),
                'conflict': {
                    'id': conflict.get('id'),
                    'paciente': conflict.get('paciente'),
                    'sala_id': conflict.get('sala_id'),
                    'sala_nome': conflict.get('sala_nome'),
                    'data': serialize_audit_value(conflict.get('data')),
                    'hora_inicio': format_conflict_time(conflict.get('hora_inicio')),
                    'hora_fim': format_conflict_time(conflict.get('hora_fim'))
                }
            }), 409

        columns = ['profissional', 'profissional_id', 'paciente', 'paciente_id', 'tipo_atendimento', 'data', 'hora_inicio', 'hora_fim']
        values = [profissional, profissional_id, paciente, paciente_id, tipo, data_field, hora_inicio, hora_fim]
        if quantidade_sessoes is not None:
            columns.append('quantidade_sessoes')
            values.append(quantidade_sessoes)
        columns.append('sala_id')
        values.append(sala_id)
        if created_by_column and created_by is not None:
            columns.append(created_by_column)
            values.append(created_by)
        if recurrence_group_id:
            columns.append('recorrencia_grupo_id')
            values.append(recurrence_group_id)
        if recurrence_index is not None:
            columns.append('recorrencia_indice')
            values.append(recurrence_index)
        if recurrence_total is not None:
            columns.append('recorrencia_total')
            values.append(recurrence_total)
        if 'status' in appointment_cols:
            columns.append('status')
            values.append(status)

        columns.append('criado_em')
        placeholders = ', '.join(['%s'] * len(values) + ['NOW()'])
        cur.execute(f"""
            INSERT INTO agendamentos ({', '.join(columns)})
            VALUES ({placeholders})
            RETURNING id, criado_em
        """, tuple(values))

        row = cur.fetchone()
        insert_agendamento_audit(
            cur,
            row[0],
            'criado',
            build_audit_user(data, authenticated_user, created_by),
            status_anterior=None,
            status_novo=status,
            detalhes={
                'profissional': profissional,
                'profissional_id': profissional_id,
                'paciente': paciente,
                'paciente_id': paciente_id,
                'tipo_atendimento': tipo,
                'data': data_field,
                'hora_inicio': hora_inicio,
                'hora_fim': hora_fim,
                'quantidade_sessoes': quantidade_sessoes,
                'sala_id': sala_id,
                'recorrencia_grupo_id': recurrence_group_id,
                'recorrencia_indice': recurrence_index,
                'recorrencia_total': recurrence_total,
                'status': status
            }
        )
        conn.commit()
        invalidate_agendamentos_list_cache()
        cur.close()
        conn.close()

        created = {
            'id': row[0],
            'profissional': profissional,
            'profissional_id': profissional_id,
            'paciente': paciente,
            'paciente_id': paciente_id,
            'tipo_atendimento': tipo,
            'data': data_field,
            'hora_inicio': hora_inicio,
            'hora_fim': hora_fim,
            'quantidade_sessoes': quantidade_sessoes,
            'sala_id': sala_id,
            'created_by': created_by,
            'status': status,
            'recorrencia_grupo_id': recurrence_group_id,
            'recorrencia_indice': recurrence_index,
            'recorrencia_total': recurrence_total,
            'criado_em': row[1].isoformat() if row[1] else None
        }

        return jsonify({'success': True, 'agendamento': created})

    except Exception as e:
        print('Erro ao criar agendamento:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/agendamentos', methods=['GET'])
def listar_agendamentos():
    auth_check = require_authenticated()
    if auth_check:
        return auth_check

    conn = None
    cur = None
    try:
        query_where = []
        params = []
        start_date = normalize_date_for_db(request.args.get('start_date') or request.args.get('from'))
        end_date = normalize_date_for_db(request.args.get('end_date') or request.args.get('to'))
        profissional_filter = request.args.get('profissional') or request.args.get('professional') or request.args.get('professionalId')

        if start_date:
            query_where.append('data::date >= %s')
            params.append(start_date)
        if end_date:
            query_where.append('data::date <= %s')
            params.append(end_date)
        try:
            limit = int(request.args.get('limit') or 1000)
        except Exception:
            limit = 1000
        limit = max(1, min(limit, 2000))

        force_refresh = str(request.args.get('force') or '').lower() in ('1', 'true', 'yes', 'sim')
        cache_key = (
            start_date.isoformat() if hasattr(start_date, 'isoformat') else str(start_date or ''),
            end_date.isoformat() if hasattr(end_date, 'isoformat') else str(end_date or ''),
            str(profissional_filter or ''),
            limit
        )
        if not force_refresh:
            cached_agendamentos = get_agendamentos_list_cache(cache_key)
            if cached_agendamentos is not None:
                return jsonify({'success': True, 'agendamentos': cached_agendamentos})

        conn = get_connection()
        cur = conn.cursor()

        appointment_cols = ensure_agendamento_link_schema(cur, get_table_columns_cached(cur, 'agendamentos'))
        ensure_agendamento_recurrence_columns(cur, appointment_cols)
        conn.commit()

        if profissional_filter:
            if 'profissional_id' in appointment_cols:
                query_where.append('(profissional_id::text = %s OR profissional::text = %s)')
                params.extend([str(profissional_filter), str(profissional_filter)])
            else:
                query_where.append('profissional::text = %s')
                params.append(str(profissional_filter))

        select_fields = ['id', 'profissional', 'paciente', 'tipo_atendimento', 'data', 'hora_inicio', 'hora_fim']
        if 'quantidade_sessoes' in appointment_cols:
            select_fields.append('quantidade_sessoes')
        if 'sala_id' in appointment_cols:
            select_fields.append('sala_id')
        if 'profissional_id' in appointment_cols:
            select_fields.append('profissional_id')
        if 'paciente_id' in appointment_cols:
            select_fields.append('paciente_id')
            
        created_by_column = None
        if 'criado_por' in appointment_cols:
            select_fields.append('criado_por')
            created_by_column = 'criado_por'
        elif 'created_by' in appointment_cols:
            select_fields.append('created_by')
            created_by_column = 'created_by'
        
        # Incluir campos de status
        if 'status' in appointment_cols:
            select_fields.append('status')
        if 'ultima_acao' in appointment_cols:
            select_fields.append('ultima_acao')
        if 'cancelado_por_username' in appointment_cols:
            select_fields.append('cancelado_por_username')
        if 'atualizado_em' in appointment_cols:
            select_fields.append('atualizado_em')
        if 'recorrencia_grupo_id' in appointment_cols:
            select_fields.append('recorrencia_grupo_id')
        if 'recorrencia_indice' in appointment_cols:
            select_fields.append('recorrencia_indice')
        if 'recorrencia_total' in appointment_cols:
            select_fields.append('recorrencia_total')
        
        select_fields.append('criado_em')

        params.append(limit)

        where_sql = f" WHERE {' AND '.join(query_where)}" if query_where else ''
        cur.execute(
            f"SELECT {', '.join(select_fields)} FROM agendamentos{where_sql} ORDER BY criado_em DESC NULLS LAST LIMIT %s",
            tuple(params)
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()

        agendamentos = []
        for r in rows:
            item = {
                'id': r[0],
                'profissional': r[1],
                'paciente': r[2],
                'tipo_atendimento': r[3],
                'data': r[4].isoformat() if hasattr(r[4], 'isoformat') else r[4],
                'hora_inicio': (r[5].isoformat() if hasattr(r[5], 'isoformat') else r[5]),
                'hora_fim': (r[6].isoformat() if hasattr(r[6], 'isoformat') else r[6])
            }
            idx = 7
            if 'quantidade_sessoes' in appointment_cols:
                item['quantidade_sessoes'] = r[idx]
                idx += 1
            if 'sala_id' in appointment_cols:
                item['sala_id'] = r[idx]
                idx += 1
            if 'profissional_id' in appointment_cols:
                item['profissional_id'] = r[idx]
                idx += 1
            if 'paciente_id' in appointment_cols:
                item['paciente_id'] = r[idx]
                idx += 1
            if created_by_column:
                item['created_by'] = r[idx]
                idx += 1
            
            # Novos campos de status
            if 'status' in appointment_cols:
                item['status'] = normalize_appointment_status(r[idx] if idx < len(r) else None, default='agendado')
                idx += 1
            if 'ultima_acao' in appointment_cols:
                item['ultima_acao'] = r[idx] if idx < len(r) else None
                idx += 1
            if 'cancelado_por_username' in appointment_cols:
                item['cancelado_por_username'] = r[idx] if idx < len(r) else None
                idx += 1
            if 'atualizado_em' in appointment_cols:
                item['atualizado_em'] = r[idx].isoformat() if idx < len(r) and r[idx] else None
                idx += 1
            if 'recorrencia_grupo_id' in appointment_cols:
                item['recorrencia_grupo_id'] = r[idx] if idx < len(r) else None
                idx += 1
            if 'recorrencia_indice' in appointment_cols:
                item['recorrencia_indice'] = r[idx] if idx < len(r) else None
                idx += 1
            if 'recorrencia_total' in appointment_cols:
                item['recorrencia_total'] = r[idx] if idx < len(r) else None
                idx += 1
            
            item['criado_em'] = r[idx].isoformat() if r[idx] else None
            agendamentos.append(item)

        set_agendamentos_list_cache(cache_key, agendamentos)
        return jsonify({'success': True, 'agendamentos': agendamentos})
    except Exception as e:
        print('Erro ao listar agendamentos:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/agendamentos/<int:agendamento_id>/recorrencia', methods=['GET'])
def obter_recorrencia_agendamento(agendamento_id):
    auth_check = require_authenticated()
    if auth_check:
        return auth_check

    conn = None
    cur = None
    try:
        conn = get_connection()
        cur = conn.cursor()
        appointment_cols = ensure_agendamento_link_schema(cur, get_table_columns_cached(cur, 'agendamentos'))
        ensure_agendamento_recurrence_columns(cur, appointment_cols)
        conn.commit()

        cur.execute(
            """
            SELECT id, recorrencia_grupo_id, data,
                   EXTRACT(DOW FROM data::date)::int AS weekday
            FROM agendamentos
            WHERE id = %s
            """,
            (agendamento_id,)
        )
        row = cur.fetchone()
        if not row:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Agendamento nao encontrado'}), 404

        recurrence_group_id = row[1]
        selected_weekday = row[3]
        single_scope = {'count': 1, 'ids': [agendamento_id]}
        if not recurrence_group_id:
            cur.close()
            conn.close()
            return jsonify({
                'success': True,
                'has_recurrence': False,
                'recorrencia_grupo_id': None,
                'selected_weekday': selected_weekday,
                'scopes': {
                    'single': single_scope,
                    'weekday': single_scope,
                    'all': single_scope
                }
            })

        cur.execute(
            """
            SELECT id, data, EXTRACT(DOW FROM data::date)::int AS weekday
            FROM agendamentos
            WHERE recorrencia_grupo_id = %s
            ORDER BY data ASC, hora_inicio ASC, id ASC
            """,
            (recurrence_group_id,)
        )
        rows = cur.fetchall()
        all_ids = [int(item[0]) for item in rows]
        weekday_ids = [int(item[0]) for item in rows if item[2] == selected_weekday]
        cur.close()
        conn.close()

        return jsonify({
            'success': True,
            'has_recurrence': len(all_ids) > 1,
            'recorrencia_grupo_id': recurrence_group_id,
            'selected_weekday': selected_weekday,
            'scopes': {
                'single': single_scope,
                'weekday': {
                    'count': len(weekday_ids) or 1,
                    'ids': weekday_ids or [agendamento_id]
                },
                'all': {
                    'count': len(all_ids) or 1,
                    'ids': all_ids or [agendamento_id]
                }
            }
        })
    except Exception as e:
        print('Erro ao consultar recorrencia do agendamento:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/lista-espera/opcoes', methods=['GET'])
def listar_lista_espera_opcoes():
    authenticated_user, auth_error = get_authenticated_user()
    if auth_error:
        return auth_error

    conn = None
    cur = None
    try:
        conn = get_connection()
        cur = conn.cursor()
        ensure_lista_espera_table(cur)
        if not user_can_manage_waitlist(cur, authenticated_user):
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Sem permissao para gerenciar lista de espera'}), 403

        patient_cols = get_table_columns_cached(cur, 'pacientes')
        patient_active_where = "WHERE ativo IS DISTINCT FROM FALSE" if 'ativo' in patient_cols else ''
        cur.execute(f"""
            SELECT id, nome, convenio
            FROM pacientes
            {patient_active_where}
            ORDER BY nome ASC
            LIMIT 1000
        """)
        pacientes = [
            {'id': row[0], 'nome': row[1], 'convenio': row[2]}
            for row in cur.fetchall()
        ]

        professional_cols = get_table_columns_cached(cur, 'profissionais')
        professional_active_where = "WHERE ativo IS DISTINCT FROM FALSE" if 'ativo' in professional_cols else ''
        cur.execute(f"""
            SELECT id, nome, especialidade
            FROM profissionais
            {professional_active_where}
            ORDER BY nome ASC
            LIMIT 1000
        """)
        profissionais = [
            {'id': row[0], 'nome': row[1], 'especialidade': row[2]}
            for row in cur.fetchall()
        ]

        try:
            salas_rows = fetch_salas_rows(cur, include_inactive=False)
        except Exception:
            conn.rollback()
            ensure_salas_schema(cur)
            conn.commit()
            salas_rows = fetch_salas_rows(cur, include_inactive=False)
        salas = [
            {'id': row[0], 'nome': row[1], 'cor': row[2], 'ativo': row[3]}
            for row in salas_rows
        ]

        conn.commit()
        cur.close()
        conn.close()
        return jsonify({
            'success': True,
            'pacientes': pacientes,
            'profissionais': profissionais,
            'salas': salas
        })
    except Exception as e:
        print('Erro ao listar opcoes da lista de espera:', e)
        try:
            if conn:
                conn.rollback()
        except Exception:
            pass
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/lista-espera', methods=['GET'])
def listar_lista_espera():
    authenticated_user, auth_error = get_authenticated_user()
    if auth_error:
        return auth_error

    conn = None
    cur = None
    try:
        conn = get_connection()
        cur = conn.cursor()
        ensure_lista_espera_table(cur)
        if not user_can_manage_waitlist(cur, authenticated_user):
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Sem permissao para gerenciar lista de espera'}), 403

        filters = []
        params = []
        raw_status = request.args.get('status')
        if raw_status:
            status = normalize_waitlist_status(raw_status, default=None)
            if not status:
                cur.close()
                conn.close()
                return jsonify({'success': False, 'error': 'Status de lista de espera invalido'}), 400
            filters.append('le.status = %s')
            params.append(status)

        search = str(request.args.get('q') or request.args.get('search') or '').strip()
        if search:
            filters.append("""
                (
                    lower(COALESCE(p.nome, le.paciente_nome, '')) LIKE lower(%s)
                    OR lower(COALESCE(pr.nome, '')) LIKE lower(%s)
                    OR lower(COALESCE(s.nome, '')) LIKE lower(%s)
                    OR lower(COALESCE(le.tipo_atendimento, '')) LIKE lower(%s)
                    OR lower(COALESCE(le.observacao, '')) LIKE lower(%s)
                )
            """)
            like_value = f"%{search}%"
            params.extend([like_value] * 5)

        where_sql = f"WHERE {' AND '.join(filters)}" if filters else ''
        cur.execute(f"""
            SELECT {', '.join(WAITLIST_SELECT_FIELDS)}
            FROM lista_espera le
            LEFT JOIN pacientes p ON p.id = le.paciente_id
            LEFT JOIN profissionais pr ON pr.id = le.profissional_id
            LEFT JOIN salas s ON s.id = le.sala_id
            {where_sql}
            ORDER BY
                CASE le.prioridade
                    WHEN 'urgente' THEN 0
                    WHEN 'alta' THEN 1
                    WHEN 'normal' THEN 2
                    WHEN 'baixa' THEN 3
                    ELSE 4
                END,
                le.criado_em ASC NULLS LAST,
                le.id ASC
            LIMIT 500
        """, tuple(params))
        rows = cur.fetchall()
        itens = [serialize_waitlist_item(row, WAITLIST_SELECT_KEYS) for row in rows]
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'success': True, 'itens': itens, 'lista_espera': itens})
    except Exception as e:
        print('Erro ao listar lista de espera:', e)
        try:
            if conn:
                conn.rollback()
        except Exception:
            pass
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/lista-espera', methods=['POST'])
def criar_lista_espera_item():
    authenticated_user, auth_error = get_authenticated_user()
    if auth_error:
        return auth_error

    data = request.json or {}
    conn = None
    cur = None
    try:
        conn = get_connection()
        cur = conn.cursor()
        ensure_lista_espera_table(cur)
        if not user_can_manage_waitlist(cur, authenticated_user):
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Sem permissao para gerenciar lista de espera'}), 403

        paciente_id = data.get('paciente_id') or data.get('patient_id') or data.get('patientId')
        paciente_nome = data.get('paciente_nome') or data.get('paciente') or data.get('patient') or data.get('clientName')
        patient_ref, patient_error = resolve_patient_reference(cur, paciente_nome, paciente_id)
        if patient_error:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': patient_error}), 400

        profissional_id = data.get('profissional_id') or data.get('professional_id') or data.get('professionalId')
        profissional_nome = data.get('profissional') or data.get('professional')
        professional_ref = None
        if profissional_id not in (None, '', 'null') or str(profissional_nome or '').strip():
            professional_ref, professional_error = resolve_professional_reference(cur, profissional_nome, profissional_id)
            if professional_error:
                cur.close()
                conn.close()
                return jsonify({'success': False, 'error': professional_error}), 400

        sala_id = normalize_room_id(data.get('sala_id') or data.get('salaId') or data.get('roomId'))
        room_ref = None
        if sala_id is not None:
            room_ref, room_error = resolve_room_reference(cur, sala_id)
            if room_error:
                cur.close()
                conn.close()
                return jsonify({'success': False, 'error': room_error}), 400

        prioridade = normalize_waitlist_priority(data.get('prioridade') or data.get('priority'), default='normal')
        status = normalize_waitlist_status(data.get('status'), default='aguardando')
        if not prioridade:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Prioridade invalida'}), 400
        if not status:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Status invalido'}), 400

        cur.execute("""
            INSERT INTO lista_espera (
                paciente_id, paciente_nome, profissional_id, sala_id, tipo_atendimento,
                prioridade, status, preferencia_dias, preferencia_horarios, observacao,
                criado_por_nome, criado_por_username, atualizado_em, encaixado_em
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),CASE WHEN %s = 'encaixado' THEN NOW() ELSE NULL END)
            RETURNING id
        """, (
            patient_ref['id'],
            patient_ref['nome'],
            professional_ref['id'] if professional_ref else None,
            room_ref['id'] if room_ref else None,
            data.get('tipo_atendimento') or data.get('type') or None,
            prioridade,
            status,
            data.get('preferencia_dias') or data.get('preferredDays') or None,
            data.get('preferencia_horarios') or data.get('preferredTimes') or None,
            data.get('observacao') or data.get('notes') or None,
            authenticated_user.get('name') or authenticated_user.get('username'),
            authenticated_user.get('username'),
            status
        ))
        item_id = cur.fetchone()[0]
        item = fetch_waitlist_item_by_id(cur, item_id)
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'success': True, 'item': item})
    except Exception as e:
        print('Erro ao criar item da lista de espera:', e)
        try:
            if conn:
                conn.rollback()
        except Exception:
            pass
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/lista-espera/<int:item_id>', methods=['PUT'])
def atualizar_lista_espera_item(item_id):
    authenticated_user, auth_error = get_authenticated_user()
    if auth_error:
        return auth_error

    data = request.json or {}
    conn = None
    cur = None
    try:
        conn = get_connection()
        cur = conn.cursor()
        ensure_lista_espera_table(cur)
        if not user_can_manage_waitlist(cur, authenticated_user):
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Sem permissao para gerenciar lista de espera'}), 403

        cur.execute("""
            SELECT paciente_id, paciente_nome, profissional_id, sala_id, tipo_atendimento,
                   prioridade, status, preferencia_dias, preferencia_horarios, observacao,
                   encaixado_agendamento_id
            FROM lista_espera
            WHERE id = %s
        """, (item_id,))
        current = cur.fetchone()
        if not current:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Item da lista de espera nao encontrado'}), 404

        values = {
            'paciente_id': current[0],
            'paciente_nome': current[1],
            'profissional_id': current[2],
            'sala_id': current[3],
            'tipo_atendimento': current[4],
            'prioridade': current[5],
            'status': current[6],
            'preferencia_dias': current[7],
            'preferencia_horarios': current[8],
            'observacao': current[9],
            'encaixado_agendamento_id': current[10]
        }

        patient_provided = any(key in data for key in ('paciente_id', 'patient_id', 'patientId', 'paciente_nome', 'paciente', 'patient', 'clientName'))
        if patient_provided:
            paciente_id = data.get('paciente_id') or data.get('patient_id') or data.get('patientId')
            paciente_nome = data.get('paciente_nome') or data.get('paciente') or data.get('patient') or data.get('clientName')
            patient_ref, patient_error = resolve_patient_reference(cur, paciente_nome, paciente_id)
            if patient_error:
                cur.close()
                conn.close()
                return jsonify({'success': False, 'error': patient_error}), 400
            values['paciente_id'] = patient_ref['id']
            values['paciente_nome'] = patient_ref['nome']

        professional_provided = any(key in data for key in ('profissional_id', 'professional_id', 'professionalId', 'profissional', 'professional'))
        if professional_provided:
            profissional_id = data.get('profissional_id') or data.get('professional_id') or data.get('professionalId')
            profissional_nome = data.get('profissional') or data.get('professional')
            if profissional_id in (None, '', 'null') and not str(profissional_nome or '').strip():
                values['profissional_id'] = None
            else:
                professional_ref, professional_error = resolve_professional_reference(cur, profissional_nome, profissional_id)
                if professional_error:
                    cur.close()
                    conn.close()
                    return jsonify({'success': False, 'error': professional_error}), 400
                values['profissional_id'] = professional_ref['id']

        room_provided = any(key in data for key in ('sala_id', 'salaId', 'roomId'))
        if room_provided:
            sala_id = normalize_room_id(data.get('sala_id') or data.get('salaId') or data.get('roomId'))
            if sala_id is None:
                values['sala_id'] = None
            else:
                room_ref, room_error = resolve_room_reference(cur, sala_id)
                if room_error:
                    cur.close()
                    conn.close()
                    return jsonify({'success': False, 'error': room_error}), 400
                values['sala_id'] = room_ref['id']

        if 'tipo_atendimento' in data or 'type' in data:
            values['tipo_atendimento'] = data.get('tipo_atendimento') or data.get('type') or None
        if 'preferencia_dias' in data or 'preferredDays' in data:
            values['preferencia_dias'] = data.get('preferencia_dias') or data.get('preferredDays') or None
        if 'preferencia_horarios' in data or 'preferredTimes' in data:
            values['preferencia_horarios'] = data.get('preferencia_horarios') or data.get('preferredTimes') or None
        if 'observacao' in data or 'notes' in data:
            values['observacao'] = data.get('observacao') or data.get('notes') or None
        if 'encaixado_agendamento_id' in data or 'appointmentId' in data:
            values['encaixado_agendamento_id'] = normalize_optional_int(data.get('encaixado_agendamento_id') or data.get('appointmentId'))
        if 'prioridade' in data or 'priority' in data:
            prioridade = normalize_waitlist_priority(data.get('prioridade') or data.get('priority'), default=None)
            if not prioridade:
                cur.close()
                conn.close()
                return jsonify({'success': False, 'error': 'Prioridade invalida'}), 400
            values['prioridade'] = prioridade
        if 'status' in data:
            status = normalize_waitlist_status(data.get('status'), default=None)
            if not status:
                cur.close()
                conn.close()
                return jsonify({'success': False, 'error': 'Status invalido'}), 400
            values['status'] = status

        cur.execute("""
            UPDATE lista_espera
            SET paciente_id = %s,
                paciente_nome = %s,
                profissional_id = %s,
                sala_id = %s,
                tipo_atendimento = %s,
                prioridade = %s,
                status = %s,
                preferencia_dias = %s,
                preferencia_horarios = %s,
                observacao = %s,
                encaixado_agendamento_id = %s,
                atualizado_em = NOW(),
                encaixado_em = CASE
                    WHEN %s = 'encaixado' THEN COALESCE(encaixado_em, NOW())
                    ELSE NULL
                END
            WHERE id = %s
        """, (
            values['paciente_id'],
            values['paciente_nome'],
            values['profissional_id'],
            values['sala_id'],
            values['tipo_atendimento'],
            values['prioridade'],
            values['status'],
            values['preferencia_dias'],
            values['preferencia_horarios'],
            values['observacao'],
            values['encaixado_agendamento_id'],
            values['status'],
            item_id
        ))
        item = fetch_waitlist_item_by_id(cur, item_id)
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'success': True, 'item': item})
    except Exception as e:
        print('Erro ao atualizar item da lista de espera:', e)
        try:
            if conn:
                conn.rollback()
        except Exception:
            pass
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sync-state', methods=['GET'])
def sync_state():
    auth_check = require_authenticated()
    if auth_check:
        return auth_check

    def format_sync_timestamp(value):
        return value.isoformat() if hasattr(value, 'isoformat') else value

    try:
        conn = get_connection()
        cur = conn.cursor()
        appointment_cols = get_table_columns_cached(cur, 'agendamentos')

        appointment_timestamp = 'criado_em'
        if 'atualizado_em' in appointment_cols:
            appointment_timestamp = "GREATEST(COALESCE(atualizado_em, criado_em), criado_em)"

        cur.execute(f"SELECT COUNT(*), MAX({appointment_timestamp}) FROM agendamentos")
        appointment_count, appointment_max = cur.fetchone()

        professional_cols = ensure_table_updated_timestamp(cur, 'profissionais', get_table_columns_cached(cur, 'profissionais'))
        professional_timestamp = "GREATEST(COALESCE(atualizado_em, criado_em), criado_em)"
        cur.execute(f"SELECT COUNT(*), MAX({professional_timestamp}) FROM profissionais")
        professional_count, professional_max = cur.fetchone()

        ensure_table_updated_timestamp(cur, 'pacientes', get_table_columns_cached(cur, 'pacientes'))
        patient_timestamp = "GREATEST(COALESCE(atualizado_em, criado_em), criado_em)"
        cur.execute(f"SELECT COUNT(*), MAX({patient_timestamp}) FROM pacientes")
        patient_count, patient_max = cur.fetchone()

        ensure_lista_espera_table(cur)
        cur.execute("""
            SELECT COUNT(*), MAX(GREATEST(COALESCE(atualizado_em, criado_em), criado_em))
            FROM lista_espera
        """)
        waitlist_count, waitlist_max = cur.fetchone()

        ensure_app_config_table(cur)
        cur.execute("SELECT COUNT(*), MAX(atualizado_em) FROM sistema_configuracoes")
        config_count, config_max = cur.fetchone()

        ensure_user_preferences_table(cur)
        cur.execute("SELECT COUNT(*), MAX(atualizado_em) FROM usuario_preferencias")
        user_preferences_count, user_preferences_max = cur.fetchone()

        ensure_remarque_table_cached(cur)
        cur.execute("""
            SELECT COUNT(*),
                   MAX(GREATEST(
                       COALESCE(solicitado_em, '1970-01-01'::timestamp),
                       COALESCE(autorizado_em, '1970-01-01'::timestamp),
                       COALESCE(rejeitado_em, '1970-01-01'::timestamp)
                   ))
            FROM remarque_solicitacoes
        """)
        remarque_count, remarque_max = cur.fetchone()

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({
            'success': True,
            'agendamentos': {
                'count': appointment_count or 0,
                'max_timestamp': format_sync_timestamp(appointment_max)
            },
            'profissionais': {
                'count': professional_count or 0,
                'max_timestamp': format_sync_timestamp(professional_max)
            },
            'pacientes': {
                'count': patient_count or 0,
                'max_timestamp': format_sync_timestamp(patient_max)
            },
            'lista_espera': {
                'count': waitlist_count or 0,
                'max_timestamp': format_sync_timestamp(waitlist_max)
            },
            'configuracoes': {
                'count': config_count or 0,
                'max_timestamp': format_sync_timestamp(config_max)
            },
            'usuarios': {
                'preferences_count': user_preferences_count or 0,
                'preferences_max_timestamp': format_sync_timestamp(user_preferences_max)
            },
            'remarques': {
                'count': remarque_count or 0,
                'max_timestamp': format_sync_timestamp(remarque_max)
            }
        })
    except Exception as e:
        print('Erro ao consultar estado de sincronizacao:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/agendamentos/<int:agendamento_id>/auditoria', methods=['GET'])
def listar_agendamento_auditoria(agendamento_id):
    authenticated_user, auth_error = get_authenticated_user()
    if auth_error:
        return auth_error

    try:
        conn = get_connection()
        cur = conn.cursor()
        ensure_agendamento_auditoria_table(cur)
        appointment_cols = get_table_columns_cached(cur, 'agendamentos')
        appointment_select_fields = ['id', 'profissional']
        if 'profissional_id' in appointment_cols:
            appointment_select_fields.append('profissional_id')
        cur.execute(
            f"SELECT {', '.join(appointment_select_fields)} FROM agendamentos WHERE id = %s",
            (agendamento_id,)
        )
        appointment_row = cur.fetchone()
        if not appointment_row:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Agendamento nao encontrado'}), 404

        appointment_data = dict(zip(appointment_select_fields, appointment_row))
        can_view_audit = (
            normalize_level(authenticated_user.get('level')) in ('admin', 'editor')
            or user_can_update_appointment_status(cur, authenticated_user, 'finalizado', appointment_data)
        )
        if not can_view_audit:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Sem permissao para ver o historico deste agendamento'}), 403

        audit_limit = None
        try:
            audit_limit = int(request.args.get('limit') or 0)
        except Exception:
            audit_limit = None
        if audit_limit:
            audit_limit = max(1, min(audit_limit, 50))
        audit_desc = str(request.args.get('order') or '').lower() in ('desc', 'recent', 'recentes')
        audit_order = 'DESC' if audit_desc else 'ASC'
        audit_limit_sql = ' LIMIT %s' if audit_limit else ''
        audit_params = [agendamento_id]
        if audit_limit:
            audit_params.append(audit_limit)

        cur.execute(f"""
            SELECT id, acao, status_anterior, status_novo, usuario_nome, usuario_username, detalhes, criado_em
            FROM agendamento_auditoria
            WHERE agendamento_id = %s
            ORDER BY criado_em {audit_order}, id {audit_order}
            {audit_limit_sql}
        """, tuple(audit_params))
        auditoria = []
        for row in cur.fetchall():
            detalhes = row[6]
            if isinstance(detalhes, str) and detalhes:
                try:
                    detalhes = json.loads(detalhes)
                except Exception:
                    pass
            auditoria.append({
                'id': row[0],
                'acao': row[1],
                'status_anterior': row[2],
                'status_novo': row[3],
                'usuario_nome': row[4],
                'usuario_username': row[5],
                'detalhes': detalhes,
                'criado_em': row[7].isoformat() if hasattr(row[7], 'isoformat') else row[7]
            })
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'success': True, 'auditoria': auditoria})
    except Exception as e:
        print('Erro ao listar auditoria do agendamento:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/agendamentos/<int:agendamento_id>', methods=['PUT'])
def atualizar_agendamento(agendamento_id):
    data = request.json or {}
    request_user, auth_error = get_authenticated_user()
    if auth_error:
        return auth_error

    profissional_provided = any(key in data for key in ('profissional', 'professional', 'professionalId', 'profissional_id', 'professional_id', 'profissionalId'))
    paciente_provided = any(key in data for key in ('paciente', 'patient', 'clientName', 'paciente_id', 'patient_id', 'patientId', 'clientId'))
    profissional = data.get('profissional') or data.get('professional') or data.get('professionalId')
    profissional_id = data.get('profissional_id') or data.get('professional_id') or data.get('profissionalId')
    paciente = data.get('paciente') or data.get('patient') or data.get('clientName')
    paciente_id = data.get('paciente_id') or data.get('patient_id') or data.get('patientId') or data.get('clientId')
    tipo = data.get('tipo_atendimento')
    data_field = data.get('data')
    hora_inicio = data.get('hora_inicio')
    hora_fim = data.get('hora_fim')
    quantidade_sessoes = data.get('quantidade_sessoes')
    recurrence_scope = normalize_recurrence_scope(
        data.get('recurrence_scope') or
        data.get('recurrenceScope') or
        data.get('apply_recurrence_scope') or
        data.get('applyRecurrenceScope')
    )
    sala_id_provided = any(key in data for key in ('sala_id', 'salaId', 'roomId', 'sala'))
    sala_id = normalize_room_id(data.get('sala_id') or data.get('salaId') or data.get('roomId') or data.get('sala'))
    status_provided = 'status' in data
    novo_status = normalize_appointment_status(data.get('status'), default=None) if status_provided else None
    if status_provided and novo_status is None:
        return jsonify({'success': False, 'error': 'Status de agendamento invalido'}), 400
    ultima_acao = data.get('ultima_acao')
    release_lock = bool(data.get('release_lock'))
    changes_schedule_data = profissional_provided or paciente_provided or any(value is not None for value in (
        tipo, data_field, hora_inicio, hora_fim, quantidade_sessoes
    )) or sala_id_provided
    if changes_schedule_data and normalize_level(request_user.get('level')) not in ('admin', 'editor'):
        return jsonify({'success': False, 'error': 'Acesso negado para alterar dados do agendamento'}), 403

    try:
        authenticated_user = None
        should_require_auth = novo_status is not None or release_lock
        should_try_auth = should_require_auth or bool(request.headers.get('Authorization'))
        try:
            should_try_auth = should_try_auth or bool(session.get('current_user'))
        except Exception:
            pass
        if should_try_auth:
            authenticated_user, auth_error = get_authenticated_user()
            if auth_error and should_require_auth:
                return auth_error
            if auth_error:
                authenticated_user = None

        conn = get_connection()
        cur = conn.cursor()

        # Verificar quais colunas existem na tabela
        appointment_cols = ensure_agendamento_link_schema(cur, get_table_columns_cached(cur, 'agendamentos'))
        ensure_agendamento_recurrence_columns(cur, appointment_cols)
        ensure_agendamento_lock_columns(cur, appointment_cols)
        ensure_agendamento_auditoria_table(cur)

        audit_select_fields = [
            'profissional', 'paciente', 'tipo_atendimento', 'data',
            'hora_inicio', 'hora_fim', 'status', 'cancelado_por_username'
        ]
        if 'profissional_id' in appointment_cols:
            audit_select_fields.insert(1, 'profissional_id')
        if 'paciente_id' in appointment_cols:
            audit_select_fields.insert(3 if 'profissional_id' in appointment_cols else 2, 'paciente_id')
        if 'quantidade_sessoes' in appointment_cols:
            audit_select_fields.insert(6, 'quantidade_sessoes')
        if 'sala_id' in appointment_cols:
            audit_select_fields.insert(6, 'sala_id')
        if 'recorrencia_grupo_id' in appointment_cols:
            audit_select_fields.append('recorrencia_grupo_id')
        cur.execute(
            f"SELECT {', '.join(audit_select_fields)} FROM agendamentos WHERE id = %s",
            (agendamento_id,)
        )
        current_row = cur.fetchone()
        if not current_row:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Agendamento nao encontrado'}), 404

        current_data = dict(zip(audit_select_fields, current_row))
        current_status = current_data.get('status')
        cancelado_por_username = current_data.get('cancelado_por_username')
        current_username = authenticated_user['username'] if authenticated_user else None

        if novo_status is not None and not user_can_update_appointment_status(cur, request_user, novo_status, current_data):
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Sem permissao para alterar este status do agendamento'}), 403

        resolved_profissional_id = None
        resolved_paciente_id = None
        if profissional_provided:
            professional_ref, professional_error = resolve_professional_reference(cur, profissional, profissional_id)
            if professional_error:
                cur.close()
                conn.close()
                return jsonify({'success': False, 'error': professional_error}), 400
            profissional = str(professional_ref['id'])
            resolved_profissional_id = professional_ref['id']
        if paciente_provided:
            patient_ref, patient_error = resolve_patient_reference(cur, paciente, paciente_id)
            if patient_error:
                cur.close()
                conn.close()
                return jsonify({'success': False, 'error': patient_error}), 400
            paciente = patient_ref['nome']
            resolved_paciente_id = patient_ref['id']
        if sala_id_provided:
            room_ref, room_error = resolve_room_reference(cur, sala_id)
            if room_error:
                cur.close()
                conn.close()
                return jsonify({'success': False, 'error': room_error}), 400
            sala_id = room_ref['id']

        if release_lock:
            if current_status != 'cancelado_profissional':
                cur.close()
                conn.close()
                return jsonify({'success': False, 'error': 'Somente agendamentos cancelados pelo profissional podem ser liberados'}), 400
            if cancelado_por_username and cancelado_por_username != current_username:
                cur.close()
                conn.close()
                return jsonify({'success': False, 'error': 'Somente o usuario que cancelou pode liberar este agendamento'}), 403

        effective_recurrence_scope = recurrence_scope if (
            changes_schedule_data and novo_status is None and not release_lock
        ) else 'single'
        recurrence_group_id = current_data.get('recorrencia_grupo_id')
        target_rows = [{'id': agendamento_id, **current_data}]
        if effective_recurrence_scope != 'single':
            if recurrence_group_id:
                target_select_fields = ['id'] + audit_select_fields
                target_params = [recurrence_group_id]
                target_where = 'recorrencia_grupo_id = %s'
                if effective_recurrence_scope == 'weekday':
                    target_where += ' AND EXTRACT(DOW FROM data::date) = EXTRACT(DOW FROM %s::date)'
                    target_params.append(current_data.get('data'))
                cur.execute(
                    f"""
                    SELECT {', '.join(target_select_fields)}
                    FROM agendamentos
                    WHERE {target_where}
                    ORDER BY data ASC, hora_inicio ASC, id ASC
                    """,
                    tuple(target_params)
                )
                target_rows = [dict(zip(target_select_fields, row)) for row in cur.fetchall()]
            if not recurrence_group_id or not target_rows:
                effective_recurrence_scope = 'single'
                target_rows = [{'id': agendamento_id, **current_data}]

        target_ids = sorted({int(item['id']) for item in target_rows})
        apply_data_field = data_field is not None and effective_recurrence_scope == 'single'
        skipped_fields = []
        if data_field is not None and effective_recurrence_scope != 'single':
            skipped_fields.append('data')

        if changes_schedule_data:
            for target_row in target_rows:
                effective_paciente_id = resolved_paciente_id if paciente_provided else target_row.get('paciente_id')
                effective_sala_id = sala_id if sala_id_provided else target_row.get('sala_id')
                effective_data = data_field if apply_data_field else target_row.get('data')
                effective_hora_inicio = hora_inicio if hora_inicio is not None else target_row.get('hora_inicio')
                effective_hora_fim = hora_fim if hora_fim is not None else (target_row.get('hora_fim') or effective_hora_inicio)
                conflict = find_patient_room_conflict(
                    cur,
                    effective_paciente_id,
                    effective_sala_id,
                    effective_data,
                    effective_hora_inicio,
                    effective_hora_fim,
                    exclude_agendamento_id=target_row.get('id'),
                    appointment_cols=appointment_cols
                )
                if conflict:
                    cur.close()
                    conn.close()
                    return jsonify({
                        'success': False,
                        'error': build_patient_room_conflict_error(conflict),
                        'conflict': {
                            'id': conflict.get('id'),
                            'paciente': conflict.get('paciente'),
                            'sala_id': conflict.get('sala_id'),
                            'sala_nome': conflict.get('sala_nome'),
                            'data': serialize_audit_value(conflict.get('data')),
                            'hora_inicio': format_conflict_time(conflict.get('hora_inicio')),
                            'hora_fim': format_conflict_time(conflict.get('hora_fim'))
                        }
                    }), 409

        fields = []
        values = []
        applied_field_pairs = {}
        if profissional_provided:
            fields.append('profissional = %s')
            values.append(profissional)
            applied_field_pairs['profissional'] = profissional
            if 'profissional_id' in appointment_cols:
                fields.append('profissional_id = %s')
                values.append(resolved_profissional_id)
                applied_field_pairs['profissional_id'] = resolved_profissional_id
        if paciente_provided:
            fields.append('paciente = %s')
            values.append(paciente)
            applied_field_pairs['paciente'] = paciente
            if 'paciente_id' in appointment_cols:
                fields.append('paciente_id = %s')
                values.append(resolved_paciente_id)
                applied_field_pairs['paciente_id'] = resolved_paciente_id
        if tipo is not None:
            fields.append('tipo_atendimento = %s')
            values.append(tipo)
            applied_field_pairs['tipo_atendimento'] = tipo
        if apply_data_field:
            fields.append('data = %s')
            values.append(data_field)
            applied_field_pairs['data'] = data_field
        if hora_inicio is not None:
            fields.append('hora_inicio = %s')
            values.append(hora_inicio)
            applied_field_pairs['hora_inicio'] = hora_inicio
        if hora_fim is not None:
            fields.append('hora_fim = %s')
            values.append(hora_fim)
            applied_field_pairs['hora_fim'] = hora_fim
        if quantidade_sessoes is not None:
            fields.append('quantidade_sessoes = %s')
            values.append(quantidade_sessoes)
            applied_field_pairs['quantidade_sessoes'] = quantidade_sessoes
        if sala_id_provided:
            fields.append('sala_id = %s')
            values.append(sala_id)
            applied_field_pairs['sala_id'] = sala_id

        if release_lock:
            fields.append('cancelado_por_username = %s')
            values.append(None)
            if ultima_acao:
                fields.append('ultima_acao = %s')
                values.append(ultima_acao)
            elif 'ultima_acao' in appointment_cols and authenticated_user:
                fields.append('ultima_acao = %s')
                values.append(authenticated_user['name'])
            if 'atualizado_em' in appointment_cols:
                fields.append('atualizado_em = NOW()')
        
        # Tratamento do status
        if novo_status is not None:
            is_locked_by_other = (
                current_status == 'cancelado_profissional'
                and cancelado_por_username
                and cancelado_por_username != current_username
            )
            allowed_after_reception_cancel = novo_status in {'em_analise', 'online'}
            if is_locked_by_other and novo_status != current_status and not allowed_after_reception_cancel:
                cur.close()
                conn.close()
                return jsonify({
                    'success': False,
                    'error': f'Somente o usuario {cancelado_por_username} pode alterar este cancelamento'
                }), 403

            fields.append('status = %s')
            values.append(novo_status)
            
            # Atualizar ultima_acao se provided
            if ultima_acao:
                fields.append('ultima_acao = %s')
                values.append(ultima_acao)
            elif 'ultima_acao' in appointment_cols:
                # Se não fornecido, usar "Sistema" como padrão
                fields.append('ultima_acao = %s')
                values.append('Sistema')
            
            # Atualizar atualizado_em se a coluna existir
            if 'atualizado_em' in appointment_cols:
                fields.append('atualizado_em = NOW()')

            if novo_status == 'cancelado_profissional':
                fields.append('cancelado_por_username = %s')
                values.append(current_username)
            elif 'cancelado_por_username' in appointment_cols and current_status == 'cancelado_profissional':
                fields.append('cancelado_por_username = %s')
                values.append(None)

        if len(fields) == 0:
            return jsonify({'success': False, 'error': 'Nada para atualizar'}), 400

        id_placeholders = ', '.join(['%s'] * len(target_ids))
        sql = f"UPDATE agendamentos SET {', '.join(fields)} WHERE id IN ({id_placeholders})"
        cur.execute(sql, tuple(values + target_ids))
        updated_count = cur.rowcount

        def build_row_audit_changes(row_data):
            row_changes = {}
            for field_name, new_value in applied_field_pairs.items():
                if (new_value is None and field_name != 'sala_id') or field_name not in row_data:
                    continue
                old_value = serialize_audit_value(row_data.get(field_name))
                new_serialized = serialize_audit_value(new_value)
                if str(old_value or '') != str(new_serialized or ''):
                    row_changes[field_name] = {
                        'antes': old_value,
                        'depois': new_serialized
                    }
            return row_changes

        audit_user = build_audit_user(data, authenticated_user, ultima_acao)
        for target_row in target_rows:
            row_id = target_row.get('id')
            row_status = target_row.get('status')
            row_changes = build_row_audit_changes(target_row)
            if novo_status is not None:
                insert_agendamento_audit(
                    cur,
                    row_id,
                    'status_alterado',
                    audit_user,
                    status_anterior=row_status,
                    status_novo=novo_status,
                    detalhes={'alteracoes': row_changes} if row_changes else None
                )
            elif release_lock:
                insert_agendamento_audit(
                    cur,
                    row_id,
                    'bloqueio_liberado',
                    audit_user,
                    status_anterior=row_status,
                    status_novo=row_status,
                    detalhes={'cancelado_por_username_anterior': target_row.get('cancelado_por_username')}
                )
            elif row_changes:
                details = {'alteracoes': row_changes}
                if effective_recurrence_scope != 'single':
                    details['recorrencia_escopo'] = effective_recurrence_scope
                    details['recorrencia_grupo_id'] = recurrence_group_id
                insert_agendamento_audit(
                    cur,
                    row_id,
                    'editado',
                    audit_user,
                    status_anterior=row_status,
                    status_novo=row_status,
                    detalhes=details
                )
        conn.commit()
        invalidate_agendamentos_list_cache()
        cur.close()
        conn.close()

        return jsonify({
            'success': True,
            'updated': updated_count,
            'updated_ids': target_ids,
            'recurrence_scope': effective_recurrence_scope,
            'skipped_fields': skipped_fields
        })
    except Exception as e:
        print('Erro ao atualizar agendamento:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/agendamentos/bulk-status', methods=['PUT'])
def bulk_update_agendamento_status():
    admin_check = require_admin()
    if admin_check:
        return admin_check

    data = request.json or {}
    appointment_ids = data.get('appointmentIds') or data.get('appointment_ids') or data.get('ids')
    status = normalize_appointment_status(data.get('status'), default=None)
    ultima_acao = data.get('ultima_acao')

    if not status:
        return jsonify({'success': False, 'error': 'Status de agendamento invalido'}), 400

    if not appointment_ids or not isinstance(appointment_ids, list):
        return jsonify({'success': False, 'error': 'appointmentIds deve ser uma lista de IDs'}), 400
    if not status:
        return jsonify({'success': False, 'error': 'Status é obrigatório'}), 400

    try:
        authenticated_user, auth_error = get_authenticated_user()
        if auth_error:
            return auth_error

        appointment_ids = [int(apt_id) for apt_id in appointment_ids if str(apt_id).strip()]
        if not appointment_ids:
            return jsonify({'success': False, 'error': 'Nenhum ID de agendamento válido fornecido'}), 400

        conn = get_connection()
        cur = conn.cursor()
        
        # Verificar se as colunas existem
        appointment_cols = get_table_columns_cached(cur, 'agendamentos')
        ensure_agendamento_lock_columns(cur, appointment_cols)

        id_placeholders = ', '.join(['%s'] * len(appointment_ids))
        cur.execute(
            f"SELECT id, status, cancelado_por_username FROM agendamentos WHERE id IN ({id_placeholders})",
            tuple(appointment_ids)
        )
        current_rows = cur.fetchall()
        current_by_id = {row[0]: {'status': row[1], 'cancelado_por_username': row[2]} for row in current_rows}
        missing_ids = [apt_id for apt_id in appointment_ids if apt_id not in current_by_id]
        if missing_ids:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': f'Agendamentos nao encontrados: {missing_ids}'}), 404

        current_username = authenticated_user['username']
        locked_by_other = []
        for apt_id in appointment_ids:
            current_row = current_by_id[apt_id]
            if (
                current_row['status'] == 'cancelado_profissional'
                and current_row['cancelado_por_username']
                and current_row['cancelado_por_username'] != current_username
                and status != current_row['status']
            ):
                locked_by_other.append({
                    'id': apt_id,
                    'cancelado_por_username': current_row['cancelado_por_username']
                })

        if locked_by_other:
            cur.close()
            conn.close()
            return jsonify({
                'success': False,
                'error': 'Existem agendamentos bloqueados por outro usuario',
                'locked_appointments': locked_by_other
            }), 403

        # Construir query dinamicamente
        update_fields = ['status = %s']
        update_values = [status]
        
        # Adicionar ultima_acao se existir no banco
        if ultima_acao:
            update_fields.append('ultima_acao = %s')
            update_values.append(ultima_acao)
        elif 'ultima_acao' in appointment_cols:
            update_fields.append('ultima_acao = %s')
            update_values.append('Sistema')
        
        # Adicionar atualizado_em se existir no banco
        if 'atualizado_em' in appointment_cols:
            update_fields.append('atualizado_em = NOW()')

        if status == 'cancelado_profissional':
            update_fields.append('cancelado_por_username = %s')
            update_values.append(current_username)
        elif 'cancelado_por_username' in appointment_cols:
            update_fields.append('cancelado_por_username = %s')
            update_values.append(None)

        query_params = tuple(update_values + appointment_ids)

        sql = (
            f"UPDATE agendamentos SET {', '.join(update_fields)} "
            f"WHERE id IN ({id_placeholders})"
        )
        cur.execute(sql, query_params)
        updated_count = cur.rowcount
        audit_user = build_audit_user(data, authenticated_user, ultima_acao)
        for apt_id in appointment_ids:
            current_row = current_by_id.get(apt_id, {})
            insert_agendamento_audit(
                cur,
                apt_id,
                'status_alterado_lote',
                audit_user,
                status_anterior=current_row.get('status'),
                status_novo=status,
                detalhes={'alteracao_em_lote': True}
            )
        conn.commit()
        invalidate_agendamentos_list_cache()
        cur.close()
        conn.close()

        return jsonify({'success': True, 'updated': updated_count})
    except Exception as e:
        print('Erro ao atualizar status em massa de agendamentos:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/agendamentos/<int:agendamento_id>', methods=['DELETE'])
def deletar_agendamento(agendamento_id):
    # Require admin credentials to delete an appointment
    admin_check = require_admin()
    if admin_check:
        return admin_check

    try:
        data = request.get_json(silent=True) or {}
        delete_repetitions = str(
            data.get('delete_repetitions') or
            data.get('deleteRepetitions') or
            data.get('delete_series') or
            ''
        ).strip().lower() in ('1', 'true', 'yes', 'sim', 's')
        delete_scope = normalize_recurrence_scope(
            data.get('delete_scope') or
            data.get('deleteScope') or
            data.get('recurrence_scope') or
            data.get('recurrenceScope')
        )
        if delete_repetitions and delete_scope == 'single':
            delete_scope = 'all'
        authenticated_user, _auth_error = get_authenticated_user()
        conn = get_connection()
        cur = conn.cursor()

        appointment_cols = get_table_columns_cached(cur, 'agendamentos')
        ensure_agendamento_recurrence_columns(cur, appointment_cols)

        select_fields = [
            'id',
            'profissional',
            'paciente',
            'tipo_atendimento',
            'data',
            'hora_inicio',
            'hora_fim',
            'status' if 'status' in appointment_cols else 'NULL AS status'
        ]
        field_names = [
            'id',
            'profissional',
            'paciente',
            'tipo_atendimento',
            'data',
            'hora_inicio',
            'hora_fim',
            'status'
        ]
        if 'recorrencia_grupo_id' in appointment_cols:
            select_fields.append('recorrencia_grupo_id')
            field_names.append('recorrencia_grupo_id')

        cur.execute(
            f"SELECT {', '.join(select_fields)} FROM agendamentos WHERE id = %s",
            (agendamento_id,)
        )
        row = cur.fetchone()
        if not row:
            conn.commit()
            cur.close()
            conn.close()
            return jsonify({'success': True, 'deleted': 0, 'deleted_ids': []})

        row_data = dict(zip(field_names, row))
        rows_to_delete = [row_data]
        recurrence_group_id = row_data.get('recorrencia_grupo_id')

        if delete_scope != 'single' and recurrence_group_id:
            target_params = [recurrence_group_id]
            target_where = 'recorrencia_grupo_id = %s'
            if delete_scope == 'weekday':
                target_where += ' AND EXTRACT(DOW FROM data::date) = EXTRACT(DOW FROM %s::date)'
                target_params.append(row_data.get('data'))
            cur.execute(
                f"""
                SELECT {', '.join(select_fields)}
                FROM agendamentos
                WHERE {target_where}
                ORDER BY data ASC, hora_inicio ASC, id ASC
                """,
                tuple(target_params)
            )
            rows_to_delete = [dict(zip(field_names, item)) for item in cur.fetchall()]
            if not rows_to_delete:
                delete_scope = 'single'
                rows_to_delete = [row_data]
        elif not recurrence_group_id:
            delete_scope = 'single'

        target_ids = sorted({int(item['id']) for item in rows_to_delete})
        audit_user = build_audit_user(data, authenticated_user)
        for item in rows_to_delete:
            insert_agendamento_audit(
                cur,
                item['id'],
                'excluido',
                audit_user,
                status_anterior=item.get('status'),
                status_novo=None,
                detalhes={
                    'profissional': item.get('profissional'),
                    'paciente': item.get('paciente'),
                    'tipo_atendimento': item.get('tipo_atendimento'),
                    'data': serialize_audit_value(item.get('data')),
                    'hora_inicio': serialize_audit_value(item.get('hora_inicio')),
                    'hora_fim': serialize_audit_value(item.get('hora_fim')),
                    'recorrencia_grupo_id': recurrence_group_id,
                    'exclusao_recorrencia': bool(delete_scope != 'single' and recurrence_group_id),
                    'recorrencia_escopo': delete_scope
                }
            )
        id_placeholders = ', '.join(['%s'] * len(target_ids))
        cur.execute(
            f'DELETE FROM agendamentos WHERE id IN ({id_placeholders})',
            tuple(target_ids)
        )
        deleted_count = cur.rowcount
        conn.commit()
        invalidate_agendamentos_list_cache()
        cur.close()
        conn.close()
        return jsonify({
            'success': True,
            'deleted': deleted_count,
            'deleted_ids': target_ids,
            'deleted_repetitions': bool(delete_scope != 'single' and recurrence_group_id and len(target_ids) > 1),
            'delete_scope': delete_scope
        })
    except Exception as e:
        print('Erro ao deletar agendamento:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/agendamentos/bulk-delete', methods=['DELETE'])
def deletar_agendamentos_em_lote():
    admin_check = require_admin()
    if admin_check:
        return admin_check

    data = request.json or {}
    appointment_ids = data.get('appointmentIds') or data.get('appointment_ids') or data.get('ids')

    if not appointment_ids or not isinstance(appointment_ids, list):
        return jsonify({'success': False, 'error': 'appointmentIds deve ser uma lista de IDs'}), 400

    try:
        authenticated_user, auth_error = get_authenticated_user()
        if auth_error:
            return auth_error

        appointment_ids = [int(apt_id) for apt_id in appointment_ids if str(apt_id).strip()]
        if not appointment_ids:
            return jsonify({'success': False, 'error': 'Nenhum ID de agendamento valido fornecido'}), 400

        conn = get_connection()
        cur = conn.cursor()

        appointment_cols = get_table_columns_cached(cur, 'agendamentos')
        ensure_agendamento_lock_columns(cur, appointment_cols)

        id_placeholders = ', '.join(['%s'] * len(appointment_ids))
        select_fields = [
            'id',
            'profissional',
            'paciente',
            'tipo_atendimento',
            'data',
            'hora_inicio',
            'hora_fim',
            'status' if 'status' in appointment_cols else 'NULL AS status',
            'cancelado_por_username'
        ]
        field_names = [
            'id',
            'profissional',
            'paciente',
            'tipo_atendimento',
            'data',
            'hora_inicio',
            'hora_fim',
            'status',
            'cancelado_por_username'
        ]
        if 'profissional_id' in appointment_cols:
            select_fields.insert(2, 'profissional_id')
            field_names.insert(2, 'profissional_id')
        if 'paciente_id' in appointment_cols:
            select_fields.insert(4 if 'profissional_id' in appointment_cols else 3, 'paciente_id')
            field_names.insert(4 if 'profissional_id' in appointment_cols else 3, 'paciente_id')
        if 'sala_id' in appointment_cols:
            select_fields.insert(-2, 'sala_id')
            field_names.insert(-2, 'sala_id')
        if 'recorrencia_grupo_id' in appointment_cols:
            select_fields.append('recorrencia_grupo_id')
            field_names.append('recorrencia_grupo_id')

        cur.execute(
            f"SELECT {', '.join(select_fields)} FROM agendamentos WHERE id IN ({id_placeholders})",
            tuple(appointment_ids)
        )
        current_rows = cur.fetchall()
        current_by_id = {row[0]: dict(zip(field_names, row)) for row in current_rows}
        missing_ids = [apt_id for apt_id in appointment_ids if apt_id not in current_by_id]
        if missing_ids:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': f'Agendamentos nao encontrados: {missing_ids}'}), 404

        audit_user = build_audit_user(data, authenticated_user)
        for apt_id in appointment_ids:
            item = current_by_id.get(apt_id, {})
            insert_agendamento_audit(
                cur,
                apt_id,
                'excluido_lote',
                audit_user,
                status_anterior=item.get('status'),
                status_novo=None,
                detalhes={
                    'profissional': item.get('profissional'),
                    'profissional_id': item.get('profissional_id'),
                    'paciente': item.get('paciente'),
                    'paciente_id': item.get('paciente_id'),
                    'tipo_atendimento': item.get('tipo_atendimento'),
                    'data': serialize_audit_value(item.get('data')),
                    'hora_inicio': serialize_audit_value(item.get('hora_inicio')),
                    'hora_fim': serialize_audit_value(item.get('hora_fim')),
                    'sala_id': item.get('sala_id'),
                    'recorrencia_grupo_id': item.get('recorrencia_grupo_id'),
                    'exclusao_em_lote': True
                }
            )

        cur.execute(
            f"DELETE FROM agendamentos WHERE id IN ({id_placeholders})",
            tuple(appointment_ids)
        )
        deleted_count = cur.rowcount
        conn.commit()
        invalidate_agendamentos_list_cache()
        cur.close()
        conn.close()

        return jsonify({'success': True, 'deleted': deleted_count})
    except Exception as e:
        print('Erro ao deletar agendamentos em massa:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/remarques', methods=['GET'])
def listar_remarques():
    authenticated_user, auth_error = get_authenticated_user()
    if auth_error:
        return auth_error

    conn = None
    cur = None
    try:
        force_refresh = str(request.args.get('force') or '').lower() in ('1', 'true', 'yes', 'sim')
        cache_key = (str(authenticated_user.get('username') or '').lower(),)
        if not force_refresh:
            cached_payload = get_remarque_list_cache(cache_key)
            if cached_payload is not None:
                conn = get_connection()
                cur = conn.cursor()
                ensure_app_config_table(cur)
                cached_payload['can_authorize'] = user_can_authorize_remarque(cur, authenticated_user)
                cached_payload['can_manage_config'] = user_can_manage_remarque_config(cur, authenticated_user)
                cached_payload['requests_enabled'] = get_remarque_requests_enabled(cur)
                cur.close()
                conn.close()
                return jsonify(cached_payload)

        conn = get_connection()
        cur = conn.cursor()
        ensure_remarque_table_cached(cur)
        ensure_app_config_table(cur)
        conn.commit()
        can_authorize = user_can_authorize_remarque(cur, authenticated_user)
        can_manage_config = user_can_manage_remarque_config(cur, authenticated_user)
        requests_enabled = get_remarque_requests_enabled(cur)

        cur.execute("""
            SELECT r.id, r.agendamento_id, r.profissional_id, r.original_data, r.original_hora_inicio, r.original_hora_fim,
                   r.nova_data, r.nova_hora_inicio, r.nova_hora_fim, r.inverter_horarios, r.conflito_agendamento_id,
                   r.conflito_nova_data, r.conflito_nova_hora_inicio, r.conflito_nova_hora_fim, r.conflito_realocacoes,
                   r.observacao, r.status, r.solicitado_por, r.solicitado_por_username, r.solicitado_em,
                   r.autorizado_por, r.autorizado_em, r.rejeitado_por, r.rejeitado_em,
                   r.motivo_reprovacao, r.decidido_por_setor,
                   a.paciente AS paciente_nome,
                   paciente_info.convenio AS paciente_convenio
            FROM remarque_solicitacoes r
            LEFT JOIN agendamentos a ON a.id = r.agendamento_id
            LEFT JOIN LATERAL (
                SELECT p.convenio
                FROM pacientes p
                WHERE CAST(a.paciente AS TEXT) = CAST(p.id AS TEXT)
                   OR lower(trim(CAST(a.paciente AS TEXT))) = lower(trim(CAST(p.nome AS TEXT)))
                ORDER BY p.criado_em DESC NULLS LAST
                LIMIT 1
            ) paciente_info ON TRUE
            ORDER BY r.solicitado_em DESC
            LIMIT 500
        """)
        rows = cur.fetchall()
        keys = [
            'id', 'agendamento_id', 'profissional_id', 'original_data', 'original_hora_inicio', 'original_hora_fim',
            'nova_data', 'nova_hora_inicio', 'nova_hora_fim', 'inverter_horarios', 'conflito_agendamento_id',
            'conflito_nova_data', 'conflito_nova_hora_inicio', 'conflito_nova_hora_fim', 'conflito_realocacoes',
            'observacao', 'status', 'solicitado_por', 'solicitado_por_username', 'solicitado_em',
            'autorizado_por', 'autorizado_em', 'rejeitado_por', 'rejeitado_em',
            'motivo_reprovacao', 'decidido_por_setor', 'paciente_nome', 'paciente_convenio'
        ]
        remarques = []
        for row in rows:
            item = dict(zip(keys, row))
            for date_key in ('original_data', 'nova_data', 'conflito_nova_data'):
                if item.get(date_key) is not None and hasattr(item[date_key], 'isoformat'):
                    item[date_key] = item[date_key].isoformat()
            for dt_key in ('solicitado_em', 'autorizado_em', 'rejeitado_em'):
                if item.get(dt_key) is not None and hasattr(item[dt_key], 'isoformat'):
                    item[dt_key] = item[dt_key].isoformat()
            remarques.append(item)

        cur.close()
        conn.close()
        payload = {
            'success': True,
            'remarques': remarques,
            'can_authorize': can_authorize,
            'can_manage_config': can_manage_config,
            'requests_enabled': requests_enabled
        }
        set_remarque_list_cache(cache_key, payload)
        return jsonify(payload)
    except Exception as e:
        print('Erro ao listar remarques:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/remarques/config', methods=['GET', 'PUT'])
def remarque_config():
    authenticated_user, auth_error = get_authenticated_user()
    if auth_error:
        return auth_error

    conn = None
    cur = None
    try:
        conn = get_connection()
        cur = conn.cursor()
        ensure_app_config_table(cur)
        can_manage = user_can_manage_remarque_config(cur, authenticated_user)

        if request.method == 'PUT':
            if not can_manage:
                cur.close()
                conn.close()
                return jsonify({'success': False, 'error': 'Somente ATAC, FINANCEIRO ou CEO podem alterar esta configuracao'}), 403

            data = request.json or {}
            enabled_value = data.get('enabled')
            if enabled_value is None:
                enabled_value = data.get('requests_enabled')
            enabled = parse_bool_config(enabled_value, True)
            actor = authenticated_user.get('name') or authenticated_user.get('username')
            set_app_config_value(cur, 'remarque_solicitacoes_ativas', 'true' if enabled else 'false', actor)
            conn.commit()
            invalidate_remarque_list_cache()
            cur.close()
            conn.close()
            return jsonify({
                'success': True,
                'enabled': enabled,
                'can_manage': can_manage,
                'updated_by': actor
            })

        enabled = get_remarque_requests_enabled(cur)
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({
            'success': True,
            'enabled': enabled,
            'can_manage': can_manage
        })
    except Exception as e:
        print('Erro ao carregar configuracao de remarque:', e)
        try:
            if conn:
                conn.rollback()
        except:
            pass
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/remarques/check-pendente', methods=['GET'])
def check_remarque_pendente():
    authenticated_user, auth_error = get_authenticated_user()
    if auth_error:
        return auth_error

    agendamento_id = request.args.get('agendamento_id') or request.args.get('appointmentId')
    if not agendamento_id:
        return jsonify({'success': False, 'error': 'agendamento_id e obrigatorio'}), 400

    try:
        conn = get_connection()
        cur = conn.cursor()
        ensure_remarque_table_cached(cur)
        cur.execute("""
            SELECT id
            FROM remarque_solicitacoes
            WHERE agendamento_id = %s
              AND status = 'pendente'
            LIMIT 1
        """, (int(agendamento_id),))
        row = cur.fetchone()
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'success': True, 'has_pending': bool(row), 'id': row[0] if row else None})
    except Exception as e:
        print('Erro ao checar remarque pendente:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/remarques', methods=['POST'])
def criar_remarque():
    authenticated_user, auth_error = get_authenticated_user()
    if auth_error:
        return auth_error

    data = request.json or {}
    agendamento_id = data.get('appointmentId') or data.get('agendamento_id')
    nova_data = data.get('newDate') or data.get('nova_data')
    nova_hora_inicio = data.get('newTime') or data.get('nova_hora_inicio')
    nova_hora_fim = data.get('newEndTime') or data.get('nova_hora_fim')
    inverter_horarios = bool(data.get('invertTimes') or data.get('inverter_horarios'))
    conflito_agendamento_id = data.get('conflictAppointmentId') or data.get('conflito_agendamento_id')
    conflito_nova_data = data.get('conflictNewDate') or data.get('conflito_nova_data')
    conflito_nova_hora_inicio = data.get('conflictNewTime') or data.get('conflito_nova_hora_inicio')
    conflito_nova_hora_fim = data.get('conflictNewEndTime') or data.get('conflito_nova_hora_fim')
    conflito_realocacoes = data.get('conflictRelocations') or data.get('conflito_realocacoes') or []
    observacao = data.get('reason') or data.get('observacao')

    if not agendamento_id or not nova_data or not nova_hora_inicio or not nova_hora_fim:
        return jsonify({'success': False, 'error': 'Agendamento, nova data e horarios sao obrigatorios'}), 400

    try:
        nova_data_obj = datetime.strptime(str(nova_data).split('T')[0], '%Y-%m-%d').date()
    except Exception:
        return jsonify({'success': False, 'error': 'Nova data invalida'}), 400

    if nova_data_obj < date.today():
        return jsonify({'success': False, 'error': 'Nao e possivel solicitar remarque para uma data passada'}), 400

    if isinstance(conflito_realocacoes, str):
        try:
            conflito_realocacoes = json.loads(conflito_realocacoes)
        except Exception:
            conflito_realocacoes = []
    if conflito_agendamento_id and not inverter_horarios and not conflito_realocacoes:
        conflito_realocacoes = [{
            'appointmentId': str(conflito_agendamento_id),
            'newDate': conflito_nova_data,
            'newTime': conflito_nova_hora_inicio,
            'newEndTime': conflito_nova_hora_fim
        }]
    if conflito_realocacoes and len(conflito_realocacoes) > 3:
        return jsonify({'success': False, 'error': 'Limite de 3 realocacoes por remarque'}), 400
    if conflito_realocacoes:
        first_relocation = conflito_realocacoes[0]
        conflito_nova_data = first_relocation.get('newDate') or first_relocation.get('nova_data') or conflito_nova_data
        conflito_nova_hora_inicio = first_relocation.get('newTime') or first_relocation.get('nova_hora_inicio') or conflito_nova_hora_inicio
        conflito_nova_hora_fim = first_relocation.get('newEndTime') or first_relocation.get('nova_hora_fim') or conflito_nova_hora_fim

    conflito_nova_data_obj = None
    if conflito_agendamento_id and not inverter_horarios:
        if not conflito_nova_data or not conflito_nova_hora_inicio or not conflito_nova_hora_fim:
            return jsonify({'success': False, 'error': 'Informe para onde vai o agendamento conflitante'}), 400
        try:
            conflito_nova_data_obj = datetime.strptime(str(conflito_nova_data).split('T')[0], '%Y-%m-%d').date()
        except Exception:
            return jsonify({'success': False, 'error': 'Nova data do conflito invalida'}), 400
        if conflito_nova_data_obj < date.today():
            return jsonify({'success': False, 'error': 'Nao e possivel realocar conflito para uma data passada'}), 400
        conflito_inicio_min = time_to_minutes(conflito_nova_hora_inicio)
        conflito_fim_min = time_to_minutes(conflito_nova_hora_fim)
        if conflito_inicio_min is None or conflito_fim_min is None or conflito_fim_min <= conflito_inicio_min:
            return jsonify({'success': False, 'error': 'Horario final do conflito deve ser maior que o inicial'}), 400
        if conflito_nova_data_obj == nova_data_obj and times_overlap(
            conflito_nova_hora_inicio, conflito_nova_hora_fim, nova_hora_inicio, nova_hora_fim
        ):
            return jsonify({'success': False, 'error': 'Realocacao do conflito nao pode ocupar o mesmo horario do remarque'}), 400

    try:
        conn = get_connection()
        cur = conn.cursor()
        ensure_remarque_table_cached(cur)
        ensure_app_config_table(cur)
        if not get_remarque_requests_enabled(cur):
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Solicitacoes de remarque estao desativadas no momento'}), 403

        cur.execute(
            'SELECT id, profissional, data, hora_inicio, hora_fim FROM agendamentos WHERE id = %s',
            (int(agendamento_id),)
        )
        agendamento = cur.fetchone()
        if not agendamento:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Agendamento nao encontrado'}), 404

        profissional_id = str(agendamento[1])
        cur.execute("""
            SELECT id
            FROM remarque_solicitacoes
            WHERE agendamento_id = %s
              AND status = 'pendente'
            LIMIT 1
        """, (int(agendamento_id),))
        existing_remarque = cur.fetchone()
        if existing_remarque:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'A solicitacao de remarque deste agendamento ja foi solicitada e esta pendente'}), 409

        if conflito_agendamento_id:
            cur.execute(
                'SELECT id, profissional FROM agendamentos WHERE id = %s',
                (int(conflito_agendamento_id),)
            )
            conflito = cur.fetchone()
            if not conflito:
                cur.close()
                conn.close()
                return jsonify({'success': False, 'error': 'Agendamento conflitante nao encontrado'}), 404
            if str(conflito[1]) != profissional_id:
                cur.close()
                conn.close()
                return jsonify({'success': False, 'error': 'Agendamento conflitante pertence a outro profissional'}), 400
            if not inverter_horarios:
                relocation_ids = []
                relocation_destinations = [{
                    'date': nova_data_obj,
                    'start': nova_hora_inicio,
                    'end': nova_hora_fim,
                    'label': 'remarque principal'
                }]
                normalized_relocations = []
                for idx, relocation in enumerate(conflito_realocacoes):
                    relocation_id = relocation.get('appointmentId') or relocation.get('agendamento_id')
                    relocation_date = relocation.get('newDate') or relocation.get('nova_data')
                    relocation_start = relocation.get('newTime') or relocation.get('nova_hora_inicio')
                    relocation_end = relocation.get('newEndTime') or relocation.get('nova_hora_fim')
                    if not relocation_id or not relocation_date or not relocation_start or not relocation_end:
                        cur.close()
                        conn.close()
                        return jsonify({'success': False, 'error': 'Realocacao de conflito incompleta'}), 400
                    try:
                        relocation_date_obj = datetime.strptime(str(relocation_date).split('T')[0], '%Y-%m-%d').date()
                    except Exception:
                        cur.close()
                        conn.close()
                        return jsonify({'success': False, 'error': 'Data de realocacao de conflito invalida'}), 400
                    relocation_start_min = time_to_minutes(relocation_start)
                    relocation_end_min = time_to_minutes(relocation_end)
                    if relocation_date_obj < date.today():
                        cur.close()
                        conn.close()
                        return jsonify({'success': False, 'error': 'Nao e possivel realocar conflito para uma data passada'}), 400
                    if relocation_start_min is None or relocation_end_min is None or relocation_end_min <= relocation_start_min:
                        cur.close()
                        conn.close()
                        return jsonify({'success': False, 'error': 'Horario final do conflito deve ser maior que o inicial'}), 400
                    for destination in relocation_destinations:
                        if destination['date'] == relocation_date_obj and times_overlap(relocation_start, relocation_end, destination['start'], destination['end']):
                            cur.close()
                            conn.close()
                            return jsonify({'success': False, 'error': 'Realocacoes da mesma solicitacao estao em conflito'}), 400
                    relocation_ids.append(int(relocation_id))
                    relocation_destinations.append({
                        'date': relocation_date_obj,
                        'start': relocation_start,
                        'end': relocation_end,
                        'label': f'agendamento {relocation_id}'
                    })
                    normalized_relocations.append({
                        'appointmentId': str(relocation_id),
                        'newDate': relocation_date_obj.isoformat(),
                        'newTime': relocation_start,
                        'newEndTime': relocation_end
                    })
                if int(conflito_agendamento_id) not in relocation_ids:
                    cur.close()
                    conn.close()
                    return jsonify({'success': False, 'error': 'Primeira realocacao deve ser do agendamento conflitante'}), 400
                relocation_placeholders = ','.join(['%s'] * len(relocation_ids))
                cur.execute(
                    f"SELECT id, profissional FROM agendamentos WHERE id IN ({relocation_placeholders})",
                    relocation_ids
                )
                relocation_rows = cur.fetchall()
                if len(relocation_rows) != len(set(relocation_ids)):
                    cur.close()
                    conn.close()
                    return jsonify({'success': False, 'error': 'Um dos agendamentos conflitantes nao foi encontrado'}), 404
                if any(str(row[1]) != profissional_id for row in relocation_rows):
                    cur.close()
                    conn.close()
                    return jsonify({'success': False, 'error': 'Todos os conflitos devem ser do mesmo profissional'}), 400

                excluded_ids = [int(agendamento_id)] + relocation_ids
                cur.execute("""
                    SELECT id, hora_inicio, hora_fim
                    FROM agendamentos
                    WHERE profissional = %s
                      AND data = %s
                      AND id <> ALL(%s)
                """, (
                    profissional_id, conflito_nova_data_obj,
                    excluded_ids
                ))
                for outro_id, outro_inicio, outro_fim in cur.fetchall():
                    if times_overlap(conflito_nova_hora_inicio, conflito_nova_hora_fim, outro_inicio, outro_fim):
                        cur.close()
                        conn.close()
                        return jsonify({'success': False, 'error': f'Realocacao do conflito ocupa horario do agendamento {outro_id}'}), 400
                for relocation in normalized_relocations[1:]:
                    relocation_date_obj = datetime.strptime(relocation['newDate'], '%Y-%m-%d').date()
                    cur.execute("""
                        SELECT id, hora_inicio, hora_fim
                        FROM agendamentos
                        WHERE profissional = %s
                          AND data = %s
                          AND id <> ALL(%s)
                    """, (
                        profissional_id, relocation_date_obj, excluded_ids
                    ))
                    for outro_id, outro_inicio, outro_fim in cur.fetchall():
                        if times_overlap(relocation['newTime'], relocation['newEndTime'], outro_inicio, outro_fim):
                            cur.close()
                            conn.close()
                            return jsonify({'success': False, 'error': f'Realocacao do conflito ocupa horario do agendamento {outro_id}'}), 400
                conflito_realocacoes = normalized_relocations

        cur.execute("""
            INSERT INTO remarque_solicitacoes (
                agendamento_id, profissional_id, original_data, original_hora_inicio, original_hora_fim,
                nova_data, nova_hora_inicio, nova_hora_fim, inverter_horarios, conflito_agendamento_id,
                conflito_nova_data, conflito_nova_hora_inicio, conflito_nova_hora_fim, conflito_realocacoes,
                observacao, solicitado_por, solicitado_por_username
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s::jsonb,%s,%s,%s)
            RETURNING id
        """, (
            int(agendamento_id), profissional_id, agendamento[2], agendamento[3], agendamento[4],
            nova_data, nova_hora_inicio, nova_hora_fim, inverter_horarios,
            int(conflito_agendamento_id) if conflito_agendamento_id else None,
            conflito_nova_data_obj if conflito_nova_data_obj else None,
            conflito_nova_hora_inicio if conflito_agendamento_id and not inverter_horarios else None,
            conflito_nova_hora_fim if conflito_agendamento_id and not inverter_horarios else None,
            json.dumps(conflito_realocacoes) if conflito_realocacoes and not inverter_horarios else None,
            observacao, authenticated_user.get('name') or authenticated_user.get('username'), authenticated_user.get('username')
        ))
        row = cur.fetchone()
        conn.commit()
        invalidate_remarque_list_cache()
        cur.close()
        conn.close()
        return jsonify({'success': True, 'id': row[0]})
    except Exception as e:
        print('Erro ao criar remarque:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/remarques/<int:remarque_id>/<action>', methods=['PUT'])
def decidir_remarque(remarque_id, action):
    authenticated_user, auth_error = get_authenticated_user()
    if auth_error:
        return auth_error
    if action not in ('approve', 'reject'):
        return jsonify({'success': False, 'error': 'Acao invalida'}), 400

    try:
        conn = get_connection()
        cur = conn.cursor()
        ensure_remarque_table_cached(cur)
        if not user_can_authorize_remarque(cur, authenticated_user):
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Usuario sem permissao para autorizar remarque'}), 403

        cur.execute("""
            SELECT agendamento_id, original_data, original_hora_inicio, original_hora_fim,
                   nova_data, nova_hora_inicio, nova_hora_fim, inverter_horarios, conflito_agendamento_id, status,
                   conflito_nova_data, conflito_nova_hora_inicio, conflito_nova_hora_fim, conflito_realocacoes
            FROM remarque_solicitacoes
            WHERE id = %s
        """, (remarque_id,))
        remarque = cur.fetchone()
        if not remarque:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Solicitacao nao encontrada'}), 404
        if remarque[9] != 'pendente':
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Solicitacao ja decidida'}), 400

        actor = authenticated_user.get('name') or authenticated_user.get('username')
        actor_sector = get_remarque_authorizer_sector(cur, authenticated_user)
        if action == 'reject':
            data = request.json or {}
            rejection_reason = (data.get('reason') or data.get('motivo_reprovacao') or data.get('rejectionReason') or '').strip()
            if not rejection_reason:
                cur.close()
                conn.close()
                return jsonify({'success': False, 'error': 'Motivo da reprovacao e obrigatorio'}), 400
            cur.execute(
                """
                UPDATE remarque_solicitacoes
                SET status = 'reprovado',
                    rejeitado_por = %s,
                    rejeitado_em = NOW(),
                    motivo_reprovacao = %s,
                    decidido_por_setor = %s
                WHERE id = %s
                """,
                (actor, rejection_reason, actor_sector, remarque_id)
            )
            conn.commit()
            invalidate_remarque_list_cache()
            cur.close()
            conn.close()
            return jsonify({'success': True, 'decidido_por_setor': actor_sector, 'rejeitado_por': actor, 'motivo_reprovacao': rejection_reason})

        (
            agendamento_id, original_data, original_inicio, original_fim,
            nova_data, nova_inicio, nova_fim, inverter, conflito_id, _status,
            conflito_nova_data, conflito_nova_inicio, conflito_nova_fim, conflito_realocacoes
        ) = remarque
        cur.execute('SELECT profissional FROM agendamentos WHERE id = %s', (agendamento_id,))
        agendamento_row = cur.fetchone()
        if not agendamento_row:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Agendamento nao encontrado'}), 404
        profissional_id = str(agendamento_row[0])
        if isinstance(conflito_realocacoes, str):
            try:
                conflito_realocacoes = json.loads(conflito_realocacoes)
            except Exception:
                conflito_realocacoes = []
        if not conflito_realocacoes and conflito_id and not inverter:
            conflito_realocacoes = [{
                'appointmentId': str(conflito_id),
                'newDate': conflito_nova_data.isoformat() if hasattr(conflito_nova_data, 'isoformat') else conflito_nova_data,
                'newTime': conflito_nova_inicio,
                'newEndTime': conflito_nova_fim
            }]
        relocation_ids = []
        for relocation in conflito_realocacoes or []:
            relocation_id = relocation.get('appointmentId') or relocation.get('agendamento_id')
            if relocation_id:
                relocation_ids.append(int(relocation_id))
        excluded_ids = [int(agendamento_id)]
        if conflito_id:
            excluded_ids.append(int(conflito_id))
        for relocation_id in relocation_ids:
            if relocation_id not in excluded_ids:
                excluded_ids.append(relocation_id)
        excluded_placeholders = ','.join(['%s'] * len(excluded_ids))
        cur.execute(
            f"""
            SELECT id, hora_inicio, hora_fim
            FROM agendamentos
            WHERE profissional = %s
              AND data = %s
              AND id NOT IN ({excluded_placeholders})
            """,
            [profissional_id, nova_data] + excluded_ids
        )
        for outro_id, outro_inicio, outro_fim in cur.fetchall():
            if times_overlap(nova_inicio, nova_fim, outro_inicio, outro_fim):
                cur.close()
                conn.close()
                return jsonify({'success': False, 'error': f'Horario do remarque conflita com agendamento {outro_id}'}), 400

        destino_conflito = None
        if inverter and conflito_id:
            destino_conflito = (original_data, original_inicio, original_fim)
        elif conflito_id:
            destino_conflito = (conflito_nova_data, conflito_nova_inicio, conflito_nova_fim)
        if destino_conflito:
            destino_data, destino_inicio, destino_fim = destino_conflito
            if not destino_data or not destino_inicio or not destino_fim:
                cur.close()
                conn.close()
                return jsonify({'success': False, 'error': 'Realocacao do conflito nao informada'}), 400
            cur.execute(
                f"""
                SELECT id, hora_inicio, hora_fim
                FROM agendamentos
                WHERE profissional = %s
                  AND data = %s
                  AND id NOT IN ({excluded_placeholders})
                """,
                [profissional_id, destino_data] + excluded_ids
            )
            for outro_id, outro_inicio, outro_fim in cur.fetchall():
                if times_overlap(destino_inicio, destino_fim, outro_inicio, outro_fim):
                    cur.close()
                    conn.close()
                    return jsonify({'success': False, 'error': f'Realocacao do conflito ocupa horario do agendamento {outro_id}'}), 400

        move_destinations = [{
            'date': nova_data,
            'start': nova_inicio,
            'end': nova_fim,
            'label': 'remarque principal'
        }]
        if not inverter and conflito_realocacoes:
            for relocation in conflito_realocacoes:
                relocation_id = relocation.get('appointmentId') or relocation.get('agendamento_id')
                relocation_date = relocation.get('newDate') or relocation.get('nova_data')
                relocation_start = relocation.get('newTime') or relocation.get('nova_hora_inicio')
                relocation_end = relocation.get('newEndTime') or relocation.get('nova_hora_fim')
                if not relocation_id or not relocation_date or not relocation_start or not relocation_end:
                    cur.close()
                    conn.close()
                    return jsonify({'success': False, 'error': 'Realocacao de conflito incompleta'}), 400
                try:
                    relocation_date_obj = datetime.strptime(str(relocation_date).split('T')[0], '%Y-%m-%d').date()
                except Exception:
                    cur.close()
                    conn.close()
                    return jsonify({'success': False, 'error': 'Data de realocacao de conflito invalida'}), 400
                for destination in move_destinations:
                    if destination['date'] == relocation_date_obj and times_overlap(relocation_start, relocation_end, destination['start'], destination['end']):
                        cur.close()
                        conn.close()
                        return jsonify({'success': False, 'error': 'Realocacoes da mesma solicitacao estao em conflito'}), 400
                cur.execute(
                    f"""
                    SELECT id, hora_inicio, hora_fim
                    FROM agendamentos
                    WHERE profissional = %s
                      AND data = %s
                      AND id NOT IN ({excluded_placeholders})
                    """,
                    [profissional_id, relocation_date_obj] + excluded_ids
                )
                for outro_id, outro_inicio, outro_fim in cur.fetchall():
                    if times_overlap(relocation_start, relocation_end, outro_inicio, outro_fim):
                        cur.close()
                        conn.close()
                        return jsonify({'success': False, 'error': f'Realocacao do conflito ocupa horario do agendamento {outro_id}'}), 400
                move_destinations.append({
                    'date': relocation_date_obj,
                    'start': relocation_start,
                    'end': relocation_end,
                    'label': f'agendamento {relocation_id}'
                })

        cur.execute(
            'UPDATE agendamentos SET data = %s, hora_inicio = %s, hora_fim = %s WHERE id = %s',
            (nova_data, nova_inicio, nova_fim, agendamento_id)
        )
        if inverter and conflito_id:
            cur.execute(
                'UPDATE agendamentos SET data = %s, hora_inicio = %s, hora_fim = %s WHERE id = %s',
                (original_data, original_inicio, original_fim, conflito_id)
            )
        elif conflito_realocacoes:
            for relocation in conflito_realocacoes:
                relocation_id = relocation.get('appointmentId') or relocation.get('agendamento_id')
                relocation_date = relocation.get('newDate') or relocation.get('nova_data')
                relocation_start = relocation.get('newTime') or relocation.get('nova_hora_inicio')
                relocation_end = relocation.get('newEndTime') or relocation.get('nova_hora_fim')
                cur.execute(
                    'UPDATE agendamentos SET data = %s, hora_inicio = %s, hora_fim = %s WHERE id = %s',
                    (relocation_date, relocation_start, relocation_end, int(relocation_id))
                )
        elif conflito_id:
            cur.execute(
                'UPDATE agendamentos SET data = %s, hora_inicio = %s, hora_fim = %s WHERE id = %s',
                (conflito_nova_data, conflito_nova_inicio, conflito_nova_fim, conflito_id)
            )
        cur.execute(
            """
            UPDATE remarque_solicitacoes
            SET status = 'aprovado',
                autorizado_por = %s,
                autorizado_em = NOW(),
                decidido_por_setor = %s
            WHERE id = %s
            """,
            (actor, actor_sector, remarque_id)
        )
        conn.commit()
        invalidate_agendamentos_list_cache()
        invalidate_remarque_list_cache()
        cur.close()
        conn.close()
        return jsonify({'success': True, 'decidido_por_setor': actor_sector, 'autorizado_por': actor})
    except Exception as e:
        print('Erro ao decidir remarque:', e)
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'error': str(e)}), 500


# =======================
# INICIAR FLASK
# =======================
def start_flask():
    threading.Thread(target=ensure_performance_indexes_background, daemon=True).start()
    print_server_access_urls()
    app.run(host=APP_HOST, port=APP_PORT)


def get_lan_ip():
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
            sock.connect(("8.8.8.8", 80))
            return sock.getsockname()[0]
    except Exception:
        return None


def print_server_access_urls():
    print("")
    print("Servidor web iniciado.")
    print(f"Neste computador: http://127.0.0.1:{APP_PORT}/mobile")
    lan_ip = get_lan_ip()
    if APP_HOST in ("0.0.0.0", "::") and lan_ip and not lan_ip.startswith("127."):
        print(f"Celular na mesma rede: http://{lan_ip}:{APP_PORT}/mobile")
    print("")

# =======================
# INICIAR APP
# =======================



# =======================
# 2. CONFIGURAÇÕES
# =======================
if getattr(sys, 'frozen', False):
    base_path = os.path.dirname(sys.executable)
else:
    base_path = os.path.abspath(".")


pasta_dados = os.path.join("E:\\", "SistemaDados")
pasta_dados = os.path.join(base_path, "dados")



# =======================
# 3. API (ponte JS ↔ Python)
# =======================
class Api:
    def ping(self):
        return {"status": "ok"}

    def limpar_cache(self):
        result = {"success": True, "desktop": True, "cookies_cleared": False}
        try:
            if webview is None:
                result["desktop"] = False
                return result

            window = webview.active_window()
            if window and hasattr(window, "clear_cookies"):
                window.clear_cookies()
                result["cookies_cleared"] = True
            return result
        except Exception as e:
            print(f"[API] Erro ao limpar cache do WebView: {e}")
            return {"success": False, "error": str(e)}

    def salvar_arquivo(self, data):
        """
        Salva um arquivo usando a janela de diálogo do sistema.
        data deve conter: {filename: str, content: str (base64)}
        """
        import base64
        import tempfile
        import os
        try:
            filename = data.get('filename', 'arquivo.xlsx')
            content_base64 = data.get('content', '')
            
            # Decodificar base64 para bytes
            if content_base64:
                content = base64.b64decode(content_base64)
            else:
                content = b''
            
            if not content:
                return {"success": False, "error": "Conteúdo vazio"}
            
            # Criar arquivo temporário
            safe_filename = os.path.basename(filename or 'arquivo.xlsx')
            base_name, extension = os.path.splitext(safe_filename)
            extension = extension or '.xlsx'
            base_name = ''.join(
                char for char in base_name
                if char.isalnum() or char in (' ', '-', '_')
            ).strip() or 'arquivo'
            base_name = base_name[:80]
            temp_dir = tempfile.gettempdir()
            fd, temp_path = tempfile.mkstemp(
                prefix=f'{base_name}_',
                suffix=extension,
                dir=temp_dir
            )
            
            # Salvar em arquivo temporário
            with os.fdopen(fd, 'wb') as f:
                f.write(content)
            
            # Tentar usar o diálogo do sistema via pywebview
            try:
                import webview
                window = webview.active_window()
                if window:
                    # Usar file_dialog para selecionar onde salvar
                    # O pywebview não tem save_file direto, então usamos uma abordagem alternativa
                    # Vamos retornar o caminho e o JavaScript pode abrir o arquivo
                    return {"success": True, "temp_path": temp_path, "filename": filename}
            except Exception as e:
                print(f"[API] Erro ao usar pywebview: {e}")
            
            # Fallback: retornar o caminho do arquivo temporário
            return {"success": True, "temp_path": temp_path, "filename": filename}
            
        except Exception as e:
            print(f"[API] Erro ao salvar arquivo: {e}")
            return {"success": False, "error": str(e)}

    def abrir_arquivo(self, filepath):
        """
        Abre um arquivo com o aplicativo padrão do sistema.
        """
        import os
        import subprocess
        try:
            if os.path.exists(filepath):
                # No Windows, usar start para abrir com o aplicativo padrão
                subprocess.Popen(['start', '', filepath], shell=True)
                return {"success": True}
            return {"success": False, "error": "Arquivo não encontrado"}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def salvar_usuario(self, usuario):
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO usuarios (username, password, level, name, notes, created_by, created_at, is_active)
            VALUES (%s, %s, %s, %s, %s, %s, NOW(), %s)
        """, (
            usuario.get("username"),
            hash_password(usuario.get("senha") or usuario.get("password") or ''),
            usuario.get("level") or 'viewer',
            usuario.get("name") or usuario.get("nome"),
            usuario.get("notes", ''),
            usuario.get("created_by") or 'system',
            True
        ))

        conn.commit()
        cur.close()
        conn.close()

        return {"status": "salvo"}


# =======================
# 5. EXPORTAÇÃO DE DADOS
# =======================
@app.route('/api/export/agendamentos', methods=['GET'])
def export_agendamentos():
    """Exporta todos os agendamentos para um arquivo Excel na pasta Downloads"""
    auth_check = require_editor_or_admin()
    if auth_check:
        return auth_check

    try:
        if not HAS_OPENPYXL:
            return jsonify({
                "success": False, 
                "error": "openpyxl não está instalado. Execute: pip install openpyxl"
            }), 500

        # Verificar se é请求 de download direto (com parâmetro download=true)
        download_mode = request.args.get('download', 'false').lower() == 'true'

        conn = get_connection()
        cur = conn.cursor()
        
        # Buscar agendamentos com dados do profissional
        cur.execute("""
            SELECT 
                a.id, a.data, a.hora_inicio, a.paciente, a.tipo_atendimento,
                a.observacoes, a.criado_em, p.nome as profissional, p.especialidade
            FROM agendamentos a
            LEFT JOIN profissionais p ON a.profissional::text = p.id::text
            ORDER BY a.data, a.hora_inicio
        """)
        
        agendamentos = cur.fetchall()
        cur.close()
        conn.close()
        
        # Criar workbook Excel
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Agendamentos"
        
        # Headers
        headers = ["Data", "Horário", "Profissional", "Especialidade", "Paciente", "Tipo", "Observações"]
        ws.append(headers)
        
        # Estilos para o header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Preenchendo dados
        type_labels = {
            'clinica': 'Clínica',
            'analise': 'Análise',
            'analise_pago': 'Análise já paga',
            'discussao': 'Discussão de Caso',
            'cls': 'CLS',
            'cls_pre': 'CLS do pré atendimento',
            'supervisao': 'Supervisão',
            'treinamento': 'Treinamento',
            'reuniao_treinamento': 'Reunião/Treinamento',
            'orientacao': 'Orientação Parental',
            'bloqueado': 'Bloqueado'
        }
        
        # Gerar nome do arquivo
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f'agendamentos_export_{timestamp}.xlsx'

        for row in agendamentos:
            data_obj = row[1]
            data_str = data_obj.strftime('%d/%m/%Y') if hasattr(data_obj, 'strftime') else str(row[1])
            
            tipo_label = type_labels.get(row[4], row[4] or "")
            
            ws.append([
                data_str,
                row[2] or "",  # hora_inicio
                row[8] or "N/A",  # profissional
                row[9] or "",  # especialidade
                row[3] or "",  # paciente
                tipo_label,
                row[5] or ""  # observações
            ])
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 30
        
        # Usar o diretório de backup configurado
        filename = f"agendamentos_clinica_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = get_backup_filepath(filename)
        
        print(f"[EXPORT] Salvando arquivo em: {filepath}")
        
        wb.save(filepath)
        
        # Verificar se o arquivo foi criado
        if os.path.exists(filepath):
            print(f"[EXPORT] ✅ Arquivo criado com sucesso: {filepath}")
            return jsonify({
                "success": True,
                "message": f"Arquivo exportado com sucesso!",
                "file": filename,
                "path": filepath
            })
        else:
            print(f"[EXPORT] ❌ Erro: Arquivo não foi criado")
            return jsonify({
                "success": False,
                "error": "Arquivo não foi criado no disco"
            }), 500
    
    except Exception as e:
        print(f"[EXPORT] ❌ Erro ao exportar agendamentos: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"Erro ao exportar: {str(e)}"
        }), 500


@app.route('/api/export/download', methods=['GET'])
def export_download():
    """Endpoint para desktop app: retorna o arquivo para download direto via browser"""
    auth_check = require_editor_or_admin()
    if auth_check:
        return auth_check

    try:
        if not HAS_OPENPYXL:
            return jsonify({"success": False, "error": "openpyxl não está instalado"}), 500

        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT a.id, a.data, a.hora_inicio, a.paciente, a.tipo_atendimento,
                   a.observacoes, a.criado_em, p.nome as profissional, p.especialidade
            FROM agendamentos a
            LEFT JOIN profissionais p ON a.profissional::text = p.id::text
            ORDER BY a.data, a.hora_inicio
        """)
        agendamentos = cur.fetchall()
        cur.close()
        conn.close()

        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from datetime import datetime
        import io

        wb = Workbook()
        ws = wb.active
        ws.title = "Agendamentos"

        headers = ["Data", "Horário", "Profissional", "Especialidade", "Paciente", "Tipo", "Observações"]
        ws.append(headers)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        type_labels = {
            'clinica': 'Clínica', 'analise': 'Análise', 'analise_pago': 'Análise já paga',
            'discussao': 'Discussão de Caso', 'cls': 'CLS', 'cls_pre': 'CLS do pré atendimento',
            'supervisao': 'Supervisão', 'treinamento': 'Treinamento',
            'reuniao_treinamento': 'Reunião/Treinamento', 'orientacao': 'Orientação Parental',
            'bloqueado': 'Bloqueado'
        }

        for row in agendamentos:
            data_obj = row[1]
            data_str = data_obj.strftime('%d/%m/%Y') if hasattr(data_obj, 'strftime') else str(row[1])
            tipo_label = type_labels.get(row[4], row[4] or "")
            ws.append([data_str, row[2] or "", row[7] or "N/A", row[8] or "", row[3] or "", tipo_label, row[5] or ""])

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 30

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'agendamentos_{timestamp}.xlsx'

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        from flask import make_response
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        return response

    except Exception as e:
        print('Erro ao exportar download:', e)
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/relatorio/agendamentos', methods=['GET'])
def relatorio_agendamentos():
    """Exporta um relatório de agendamentos em Excel para admin/editor."""
    try:
        if not HAS_OPENPYXL:
            return jsonify({
                "success": False,
                "error": "openpyxl não está instalado. Execute: pip install openpyxl"
            }), 500

        auth_check = require_editor_or_admin()
        if auth_check:
            return auth_check

        include_type = parse_bool_config(request.args.get('include_type'), True)
        include_sessions = parse_bool_config(request.args.get('include_sessions'), True)

        conn = get_connection()
        cur = conn.cursor()

        appointment_cols = get_table_columns_cached(cur, 'agendamentos')
        patient_cols = get_table_columns_cached(cur, 'pacientes')
        professional_cols = get_table_columns_cached(cur, 'profissionais')

        select_fields = [
            'a.id',
            'a.data',
            'a.hora_inicio',
            'a.hora_fim',
            'a.paciente',
            'a.tipo_atendimento'
        ]

        if 'status' in appointment_cols:
            select_fields.append('a.status')

        if 'label' in appointment_cols:
            select_fields.append('a.label AS relatorio_label')
        elif 'rotulo' in appointment_cols:
            select_fields.append('a.rotulo AS relatorio_label')
        elif 'cnn' in appointment_cols:
            select_fields.append('a.cnn AS relatorio_label')

        if 'criado_por' in appointment_cols:
            select_fields.append('a.criado_por AS relatorio_created_by')
        elif 'created_by' in appointment_cols:
            select_fields.append('a.created_by AS relatorio_created_by')

        if 'observacoes' in appointment_cols:
            select_fields.append('a.observacoes AS relatorio_observations')
        elif 'observations' in appointment_cols:
            select_fields.append('a.observations AS relatorio_observations')

        if 'fixo' in appointment_cols:
            select_fields.append('a.fixo AS relatorio_fixo')
        elif 'esporadico' in appointment_cols:
            select_fields.append('a.esporadico AS relatorio_fixo')
        elif 'recorrencia' in appointment_cols:
            select_fields.append('a.recorrencia AS relatorio_fixo')

        if 'agenda' in appointment_cols:
            select_fields.append('a.agenda AS relatorio_agenda')
        elif 'cnn' in appointment_cols and 'label' not in appointment_cols:
            select_fields.append('a.cnn AS relatorio_agenda')

        if 'procedimento' in appointment_cols:
            select_fields.append('a.procedimento AS relatorio_procedimento')
        elif 'procedure' in appointment_cols:
            select_fields.append('a.procedure AS relatorio_procedimento')

        if 'quantidade_sessoes' in appointment_cols:
            select_fields.append('a.quantidade_sessoes')

        remark_keys = [
            ('remarque_profissional', 'relatorio_remarque_profissional'),
            ('remarque_dia', 'relatorio_remarque_dia'),
            ('remarque_hora', 'relatorio_remarque_hora'),
            ('remarque_tipo', 'relatorio_remarque_tipo'),
            ('reagendado_profissional', 'relatorio_remarque_profissional'),
            ('reagendado_dia', 'relatorio_remarque_dia'),
            ('reagendado_hora', 'relatorio_remarque_hora'),
            ('reagendado_tipo', 'relatorio_remarque_tipo')
        ]
        for column_name, alias_name in remark_keys:
            if column_name in appointment_cols and alias_name not in select_fields:
                select_fields.append(f'a.{column_name} AS {alias_name}')

        if 'telefone' in patient_cols:
            select_fields.append('p.telefone AS paciente_telefone')
        if 'phone' in patient_cols:
            select_fields.append('p.phone AS paciente_telefone')
        if 'data_nascimento' in patient_cols:
            select_fields.append('p.data_nascimento AS paciente_data_nascimento')
        if 'convenio' in patient_cols:
            select_fields.append('p.convenio AS paciente_convenio')
        if 'plano' in patient_cols:
            select_fields.append('p.plano AS paciente_plano')
        if 'plano_convenio' in patient_cols:
            select_fields.append('p.plano_convenio AS paciente_plano')

        if 'especialidade' in professional_cols:
            professional_field = 'p_prof.especialidade AS profissional_especialidade'
        else:
            professional_field = 'NULL AS profissional_especialidade'

        start_date = normalize_date_for_db(request.args.get('start_date'))
        end_date = normalize_date_for_db(request.args.get('end_date'))
        query_where = []
        params = []
        professional_filter_id = normalize_optional_int(
            request.args.get('professional_id') or request.args.get('profissional_id')
        )
        type_filters = []
        for value in request.args.getlist('type') + request.args.getlist('tipo_atendimento') + request.args.getlist('tipo'):
            text = str(value or '').strip()
            if text:
                type_filters.append(text)
        raw_types = request.args.get('types') or request.args.get('tipos')
        if raw_types:
            for value in str(raw_types).split(','):
                text = value.strip()
                if text:
                    type_filters.append(text)
        type_filters = list(dict.fromkeys(type_filters))

        if start_date:
            query_where.append("a.data::date >= %s")
            params.append(start_date)
        if end_date:
            query_where.append("a.data::date <= %s")
            params.append(end_date)
        if professional_filter_id is not None:
            if 'profissional_id' in appointment_cols:
                query_where.append("(a.profissional_id = %s OR (a.profissional_id IS NULL AND a.profissional::text = %s))")
                params.extend([professional_filter_id, str(professional_filter_id)])
            else:
                query_where.append("a.profissional::text = %s")
                params.append(str(professional_filter_id))
        if type_filters:
            placeholders = ', '.join(['%s'] * len(type_filters))
            query_where.append(f"a.tipo_atendimento IN ({placeholders})")
            params.extend(type_filters)

        professional_join = "LEFT JOIN profissionais p_prof ON a.profissional::text = p_prof.id::text"
        if 'profissional_id' in appointment_cols:
            professional_join = "LEFT JOIN profissionais p_prof ON (a.profissional_id = p_prof.id OR (a.profissional_id IS NULL AND a.profissional::text = p_prof.id::text))"

        query = f"""
            SELECT {', '.join(select_fields)},
                   p.nome AS paciente_nome,
                   p_prof.nome AS profissional,
                   {professional_field}
            FROM agendamentos a
            LEFT JOIN pacientes p ON (a.paciente::text = p.id::text OR a.paciente = p.nome)
            {professional_join}
            {('WHERE ' + ' AND '.join(query_where)) if query_where else ''}
            ORDER BY a.data, a.hora_inicio
        """

        cur.execute(query, tuple(params))
        rows = cur.fetchall()
        column_names = [desc[0] for desc in cur.description]
        data_rows = []
        for row in rows:
            item = dict(zip(column_names, row))
            data_rows.append(item)

        remarque_report_by_appointment = {}
        appointment_ids = [item.get('id') for item in data_rows if item.get('id') is not None]
        if appointment_ids:
            try:
                ensure_remarque_table_cached(cur)
                id_placeholders = ','.join(['%s'] * len(appointment_ids))
                cur.execute(f"""
                    SELECT DISTINCT ON (agendamento_id)
                           agendamento_id,
                           original_data,
                           original_hora_inicio,
                           original_hora_fim,
                           nova_data,
                           nova_hora_inicio,
                           nova_hora_fim
                    FROM remarque_solicitacoes
                    WHERE status = 'aprovado'
                      AND agendamento_id IN ({id_placeholders})
                    ORDER BY agendamento_id, autorizado_em DESC NULLS LAST, id DESC
                """, tuple(appointment_ids))
                for remarque_row in cur.fetchall():
                    remarque_report_by_appointment[str(remarque_row[0])] = {
                        'original_data': remarque_row[1],
                        'original_hora_inicio': remarque_row[2],
                        'original_hora_fim': remarque_row[3],
                        'nova_data': remarque_row[4],
                        'nova_hora_inicio': remarque_row[5],
                        'nova_hora_fim': remarque_row[6]
                    }
                conn.commit()
            except Exception as remarque_error:
                print('Erro ao buscar informacoes de remarque para relatorio:', remarque_error)

        cur.close()
        conn.close()

        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório Agendamentos"

        headers = [
            'Cod',
            'Situação',
            'Tipo de Atendimento',
            'Data',
            'Horário inicial',
            'Horário final',
            'Quantidade de sessões',
            'Criado por',
            'Paciente',
            'Telefone',
            'Data de nascimento',
            'Profissional',
            'Setor / Especialidade',
            'Convênio',
            'Procedimento',
            'Observações',
            'Fixo ou esporádico',
            'Informações de remarque'
        ]
        if not include_type:
            headers.pop(2)
        if not include_sessions:
            session_header_index = 6 if include_type else 5
            headers.pop(session_header_index)
        ws.append(headers)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        def format_date(value):
            if not value:
                return ''
            if hasattr(value, 'strftime'):
                return value.strftime('%d/%m/%Y')
            if isinstance(value, str):
                try:
                    parsed = datetime.fromisoformat(value)
                    return parsed.strftime('%d/%m/%Y')
                except Exception:
                    try:
                        parsed = datetime.strptime(value, '%d/%m/%Y')
                        return parsed.strftime('%d/%m/%Y')
                    except Exception:
                        return value
            return str(value)

        def format_report_time(value):
            if not value:
                return ''
            if hasattr(value, 'strftime'):
                return value.strftime('%H:%M')
            text = str(value)
            parts = text.split(':')
            if len(parts) >= 2:
                return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}"
            return text

        def format_session_count(start_time, end_time):
            if not start_time or not end_time:
                return ''

            def parse_time(value):
                if isinstance(value, str) and ':' in value:
                    for fmt in ('%H:%M', '%H:%M:%S'):
                        try:
                            return datetime.strptime(value, fmt)
                        except Exception:
                            continue
                    return None
                if hasattr(value, 'hour') and hasattr(value, 'minute'):
                    return datetime(2000, 1, 1, value.hour, value.minute, getattr(value, 'second', 0) or 0)
                return None

            try:
                start_dt = parse_time(start_time)
                end_dt = parse_time(end_time)
                if start_dt and end_dt:
                    diff = int((end_dt - start_dt).total_seconds() // 60)
                    return int(diff / 30) if diff >= 0 else ''
            except Exception:
                pass
            return ''

        status_map = {
            'agendado': 'Agendado',
            'confirmado': 'Confirmado',
            'pré-atendimento': 'Pré-atendimento',
            'pre-atendimento': 'Pré-atendimento',
            'pre_atendimento': 'Pré-atendimento',
            'pré_atendimento': 'Pré-atendimento',
            'pago pré': 'Pago pré',
            'pago_pre': 'Pago pré',
            'em espera': 'Em espera',
            'em_espera': 'Em espera',
            'realizado': 'Realizado',
            'finalizado': 'Finalizado',
            'análise': 'Análise (faltou/consumo)',
            'analise': 'Análise (faltou/consumo)',
            'em atendimento': 'Em atendimento / orientação',
            'em_atendimento': 'Em atendimento / orientação',
            'orientacao': 'Em atendimento / orientação',
            'chegou': 'Chegou',
            'faltou': 'Faltou',
            'nao_compareceu': 'Faltou',
            'nao compareceu': 'Faltou',
            'cancelado_profissional': 'Cancelado pelo profissional',
            'cancelado profissional': 'Cancelado pelo profissional',
            'cancelado': 'Cancelado pelo profissional'
        }

        for item in data_rows:
            status_raw = (item.get('status') or '').strip().lower()
            status_text = status_map.get(status_raw, item.get('status') or 'Agendado')

            label_text = item.get('tipo_atendimento') or item.get('relatorio_label') or ''
            created_by_text = item.get('relatorio_created_by') or ''
            paciente_nome = item.get('paciente_nome') or item.get('paciente') or ''
            telefone_text = item.get('paciente_telefone') or ''
            nascimento_text = format_date(item.get('paciente_data_nascimento'))
            profissional_text = item.get('profissional') or ''
            especialidade_text = item.get('profissional_especialidade') or ''
            convenio_text = item.get('paciente_convenio') or ''
            plano_text = item.get('paciente_plano') or ''
            procedimento_text = item.get('relatorio_procedimento') or ''
            observacoes_text = item.get('relatorio_observations') or ''
            fixo_text = item.get('relatorio_fixo')
            if isinstance(fixo_text, bool):
                fixo_text = 'Fixo' if fixo_text else 'Esporádico'
            elif isinstance(fixo_text, str):
                lower = fixo_text.strip().lower()
                if lower in ('true', 'sim', 's', '1', 'fixo'):
                    fixo_text = 'Fixo'
                elif lower in ('false', 'não', 'nao', '0', 'esporádico', 'esporadico', 'espora'):
                    fixo_text = 'Esporádico'
                else:
                    fixo_text = fixo_text
            else:
                fixo_text = fixo_text or ''

            remarque_parts = []
            for key in ('relatorio_remarque_profissional', 'relatorio_remarque_dia', 'relatorio_remarque_hora', 'relatorio_remarque_tipo'):
                if item.get(key):
                    label = key.replace('relatorio_remarque_', '').replace('_', ' ').capitalize()
                    remarque_parts.append(f"{label}: {item.get(key)}")
            remarque_text = ' | '.join(remarque_parts)
            approved_remarque = remarque_report_by_appointment.get(str(item.get('id')))
            if approved_remarque:
                remarque_text = (
                    f"Agendamento remarcado de "
                    f"{format_date(approved_remarque.get('original_data'))} "
                    f"{format_report_time(approved_remarque.get('original_hora_inicio'))} - "
                    f"{format_report_time(approved_remarque.get('original_hora_fim'))} "
                    f"para {format_date(approved_remarque.get('nova_data'))} "
                    f"{format_report_time(approved_remarque.get('nova_hora_inicio'))} - "
                    f"{format_report_time(approved_remarque.get('nova_hora_fim'))}"
                )

            session_count = item.get('quantidade_sessoes') if item.get('quantidade_sessoes') is not None else format_session_count(item.get('hora_inicio'), item.get('hora_fim'))
            if session_count == '' or session_count is None:
                session_count = ''

            data_value = item.get('data')
            # Corrigir: garantir que só a data local seja usada, sem deslocamento de timezone
            if isinstance(data_value, str):
                # Extrair apenas a parte da data (YYYY-MM-DD)
                date_part = data_value.split('T')[0].split(' ')[0]
                try:
                    data_value = datetime.strptime(date_part, '%Y-%m-%d')
                except Exception:
                    try:
                        data_value = datetime.strptime(date_part, '%d/%m/%Y')
                    except Exception:
                        data_value = None
            elif hasattr(data_value, 'strftime'):
                # já é datetime/date
                pass
            else:
                data_value = None

            row_values = [
                item.get('id'),
                status_text,
                label_text,
                format_date(data_value),
                item.get('hora_inicio') or '',
                item.get('hora_fim') or '',
                session_count,
                created_by_text,
                paciente_nome,
                telefone_text,
                nascimento_text,
                profissional_text,
                especialidade_text,
                convenio_text,
                procedimento_text,
                observacoes_text,
                fixo_text,
                remarque_text
            ]
            if not include_type:
                row_values.pop(2)
            if not include_sessions:
                session_index = 6 if include_type else 5
                row_values.pop(session_index)
            ws.append(row_values)

        col_widths = [10, 18, 20, 14, 14, 14, 12, 20, 25, 18, 15, 25, 22, 18, 18, 22, 30, 40]
        if not include_type:
            col_widths.pop(2)
        if not include_sessions:
            session_width_index = 6 if include_type else 5
            col_widths.pop(session_width_index)
        for index, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width

        try:
            logo_path = get_file_path('logo.png')
            if os.path.exists(logo_path):
                from importlib import import_module, util

                if util.find_spec('PIL') is not None:
                    from openpyxl.drawing.image import Image as XLImage

                    pil_image = import_module('PIL.Image')
                    logo = pil_image.open(logo_path).convert('RGBA')
                    max_width = 520
                    if logo.width > max_width:
                        ratio = max_width / float(logo.width)
                        logo = logo.resize((max_width, int(logo.height * ratio)))

                    alpha = logo.getchannel('A')
                    logo.putalpha(alpha.point(lambda value: int(value * 0.12)))

                    watermark_buffer = io.BytesIO()
                    logo.save(watermark_buffer, format='PNG')
                    watermark_buffer.seek(0)

                    watermark = XLImage(watermark_buffer)
                    watermark.anchor = 'E6'
                    ws.add_image(watermark)
                    wb._hpt_watermark_buffer = watermark_buffer
        except Exception as watermark_error:
            print('Aviso: nao foi possivel aplicar marca d agua no relatorio:', watermark_error)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        file_content = buffer.getvalue()

        # Verificar se é uma requisição do pywebview (via API)
        use_api = request.args.get('use_api', 'false').lower() == 'true'
        
        if use_api:
            # Retornar JSON com conteúdo base64 para usar com pywebview API
            import base64
            user_suffix = (request.args.get('user', '') or 'usuario')
            # Remover caracteres especiais do nome do usuário
            user_suffix = ''.join(c for c in user_suffix if c.isalnum() or c in '_-')
            if not user_suffix:
                user_suffix = 'usuario'
            start_date_str = request.args.get('start_date', '')
            end_date_str = request.args.get('end_date', '')
            filename = f'relatorio_agendamentos_{user_suffix}_{start_date_str}_a_{end_date_str}.xlsx'
            
            return jsonify({
                "success": True,
                "filename": filename,
                "content": base64.b64encode(file_content).decode('utf-8')
            })

        response = send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='relatorio_agendamentos.xlsx'
        )
        return response

    except Exception as e:
        print(f"[RELATORIO] ❌ Erro ao exportar relatório de agendamentos: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"Erro ao gerar relatório: {str(e)}"
        }), 500


@app.route('/api/export/modelo', methods=['GET'])
def export_modelo():
    auth_check = require_authenticated()
    if auth_check:
        return auth_check

    """Exporta um modelo de planilha para importação"""
    try:
        if not HAS_OPENPYXL:
            return jsonify({
                "success": False,
                "error": "openpyxl não está instalado"
            }), 500
        
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Modelo"
        
        # Dados de exemplo
        headers = ["Data", "Horário", "Profissional ID", "Paciente ID", "Sala ID", "Tipo", "Observações"]
        ws.append(headers)
        
        # Estilos para o header
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Exemplo
        ws.append([
            "01/01/2024",
            "09:00",
            "1",
            "1",
            "1",
            "Clínica",
            "Observações opcionais"
        ])
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 30
        
        # Usar o diretório de backup configurado
        filename = f"modelo_agendamentos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = get_backup_filepath(filename)
        
        print(f"[MODELO] Salvando arquivo em: {filepath}")
        
        wb.save(filepath)
        
        # Verificar se o arquivo foi criado
        if os.path.exists(filepath):
            print(f"[MODELO] ✅ Modelo criado com sucesso: {filepath}")
            return jsonify({
                "success": True,
                "message": "Modelo exportado com sucesso!",
                "file": filename,
                "path": filepath
            })
        else:
            print(f"[MODELO] ❌ Erro: Arquivo não foi criado")
            return jsonify({
                "success": False,
                "error": "Arquivo não foi criado no disco"
            }), 500
    
    except Exception as e:
        print(f"[MODELO] ❌ Erro ao exportar modelo: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"Erro ao exportar: {str(e)}"
        }), 500


# =======================
# 6. TESTES E VALIDAÇÃO
# =======================
@app.route('/api/test/backup', methods=['GET'])
def test_backup():
    auth_check = require_admin()
    if auth_check:
        return auth_check

    """Testa se é possível criar arquivos no diretório de backup"""
    try:
        # Tentar criar um arquivo de teste
        test_filename = f"test_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        test_filepath = get_backup_filepath(test_filename)
        
        # Criar arquivo de teste
        with open(test_filepath, 'w', encoding='utf-8') as f:
            f.write(f"Teste de salvamento\n")
            f.write(f"Timestamp: {datetime.now().isoformat()}\n")
            f.write(f"Caminho: {test_filepath}\n")
        
        # Verificar se foi criado
        if os.path.exists(test_filepath):
            # Ler arquivo de teste
            with open(test_filepath, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Limpar arquivo de teste
            os.remove(test_filepath)
            
            return jsonify({
                "success": True,
                "message": "Diretório de backup está funcionando corretamente!",
                "backup_dir": BACKUP_DIR,
                "test_file": test_filename,
                "test_content": content
            })
        else:
            return jsonify({
                "success": False,
                "error": "Arquivo de teste não foi criado",
                "backup_dir": BACKUP_DIR
            }), 500
    
    except Exception as e:
        print(f"[TEST] ❌ Erro ao testar backup: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"Erro ao testar: {str(e)}",
            "backup_dir": BACKUP_DIR
        }), 500


@app.route('/api/test/list-backups', methods=['GET'])
def test_list_backups():
    auth_check = require_admin()
    if auth_check:
        return auth_check

    """Lista todos os arquivos no diretório de backup"""
    try:
        if not os.path.exists(BACKUP_DIR):
            return jsonify({
                "success": False,
                "error": f"Diretório não existe: {BACKUP_DIR}"
            }), 404
        
        files = []
        for filename in os.listdir(BACKUP_DIR):
            filepath = os.path.join(BACKUP_DIR, filename)
            if os.path.isfile(filepath):
                stat = os.stat(filepath)
                files.append({
                    "name": filename,
                    "size": stat.st_size,
                    "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    "path": filepath
                })
        
        files.sort(key=lambda x: x["modified"], reverse=True)
        
        return jsonify({
            "success": True,
            "backup_dir": BACKUP_DIR,
            "file_count": len(files),
            "files": files
        })
    
    except Exception as e:
        print(f"[TEST] ❌ Erro ao listar backups: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"Erro ao listar: {str(e)}"
        }), 500


# =======================
# SINCRONIZAÇÃO COM SUPABASE
# =======================
@app.route('/api/sync/supabase', methods=['POST'])
def sync_with_supabase():
    """
    Sincroniza dados com Supabase:
    - Profissionais
    - Agendamentos
    - Usuários
    
    Apenas administradores podem chamar esta rota.
    """
    admin_check = require_admin()
    if admin_check:
        return admin_check

    try:
        # Verificar autenticação
        auth_header = request.headers.get('Authorization', '') or 'Bearer session:authenticated'
        if not auth_header.startswith('Bearer '):
            return jsonify({
                'success': False,
                'error': 'Autenticação necessária'
            }), 401
        
        # Extrair credenciais do header
        credentials = auth_header.replace('Bearer ', '')
        if ':' not in credentials:
            return jsonify({
                'success': False,
                'error': 'Formato de autenticação inválido'
            }), 401
        
        username, password = credentials.split(':', 1)
        
        # Obter dados do request
        data = request.json
        
        professionals = data.get('professionals', [])
        appointments = data.get('appointments', [])
        users_data = data.get('users', {})
        
        sync_summary = {
            'professionals_synced': 0,
            'appointments_synced': 0,
            'users_synced': 0,
            'timestamp': datetime.now().isoformat()
        }
        
        print(f"[SYNC] 📊 Iniciando sincronização com Supabase...")
        print(f"[SYNC] 👨‍⚕️ Profissionais: {len(professionals)}")
        print(f"[SYNC] 📅 Agendamentos: {len(appointments)}")
        print(f"[SYNC] 👥 Usuários: {len(users_data)}")
        
        # Sincronizar profissionais
        if professionals:
            for prof in professionals:
                try:
                    conn = get_connection()
                    cur = conn.cursor()
                    
                    # Tentar inserir ou atualizar
                    nome = prof.get('nome') or prof.get('name', '')
                    especialidade = prof.get('especialidade') or prof.get('specialty', '')
                    prof_id = prof.get('id')
                    
                    # Inserir se não existir
                    cur.execute("""
                        INSERT INTO profissionais (nome, especialidade)
                        SELECT %s, %s
                        WHERE NOT EXISTS (
                            SELECT 1 FROM profissionais WHERE LOWER(nome) = LOWER(%s)
                        )
                        RETURNING id
                    """, (nome, especialidade, nome))
                    
                    conn.commit()
                    sync_summary['professionals_synced'] += 1
                    
                    cur.close()
                    conn.close()
                except Exception as e:
                    print(f"[SYNC] ⚠️ Erro ao sincronizar profissional {nome}: {str(e)}")
                    try:
                        conn.rollback()
                        cur.close()
                        conn.close()
                    except:
                        pass
        
        # Sincronizar agendamentos
        if appointments:
            for apt in appointments:
                try:
                    conn = get_connection()
                    cur = conn.cursor()
                    
                    appointment_cols = ensure_agendamento_link_schema(cur, get_table_columns_cached(cur, 'agendamentos'))

                    profissional = apt.get('profissional') or apt.get('professional', '')
                    profissional_id = apt.get('profissional_id') or apt.get('professional_id') or apt.get('professionalId')
                    paciente = apt.get('paciente') or apt.get('patient', '')
                    paciente_id = apt.get('paciente_id') or apt.get('patient_id') or apt.get('patientId') or apt.get('clientId')
                    sala_id = normalize_room_id(apt.get('sala_id') or apt.get('salaId') or apt.get('roomId') or apt.get('sala'))
                    tipo_atendimento = apt.get('tipo_atendimento') or apt.get('service_type', '')
                    data_apt = apt.get('data') or apt.get('date', '')
                    hora_inicio = apt.get('hora_inicio') or apt.get('start_time', '')
                    hora_fim = apt.get('hora_fim') or apt.get('end_time', '')

                    professional_ref, professional_error = resolve_professional_reference(cur, profissional, profissional_id)
                    patient_ref, patient_error = resolve_patient_reference(cur, paciente, paciente_id)
                    room_ref, room_error = resolve_room_reference(cur, sala_id)
                    if professional_error or patient_error or room_error:
                        raise ValueError(professional_error or patient_error or room_error)

                    profissional = str(professional_ref['id'])
                    profissional_id = professional_ref['id']
                    paciente = patient_ref['nome']
                    paciente_id = patient_ref['id']
                    sala_id = room_ref['id']

                    conflict = find_patient_room_conflict(
                        cur,
                        paciente_id,
                        sala_id,
                        data_apt,
                        hora_inicio,
                        hora_fim or hora_inicio,
                        appointment_cols=appointment_cols
                    )
                    if conflict:
                        raise ValueError(build_patient_room_conflict_error(conflict))
                    
                    # Inserir se não existir (duplicidade por data+hora+profissional+paciente)
                    cur.execute("""
                        INSERT INTO agendamentos (
                            profissional, profissional_id, paciente, paciente_id,
                            tipo_atendimento, data, hora_inicio, hora_fim, sala_id, criado_em
                        )
                        SELECT %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW()
                        WHERE NOT EXISTS (
                            SELECT 1 FROM agendamentos
                            WHERE data = %s
                            AND hora_inicio = %s
                            AND profissional_id = %s
                            AND paciente_id = %s
                        )
                        RETURNING id
                    """, (
                        profissional, profissional_id, paciente, paciente_id,
                        tipo_atendimento, data_apt, hora_inicio, hora_fim, sala_id,
                        data_apt, hora_inicio, profissional_id, paciente_id
                    ))
                    
                    conn.commit()
                    invalidate_agendamentos_list_cache()
                    sync_summary['appointments_synced'] += 1
                    
                    cur.close()
                    conn.close()
                except Exception as e:
                    print(f"[SYNC] ⚠️ Erro ao sincronizar agendamento: {str(e)}")
                    try:
                        conn.rollback()
                        cur.close()
                        conn.close()
                    except:
                        pass
        
        # Sincronizar usuários
        if users_data:
            for username_key, user_info in users_data.items():
                try:
                    conn = get_connection()
                    cur = conn.cursor()
                    
                    user_name = user_info.get('name', '')
                    user_password = user_info.get('password', '')
                    if not user_password:
                        print(f"[SYNC] Usuário {username_key} ignorado sem senha local para sincronização segura")
                        continue
                    if not is_password_hash(user_password):
                        user_password = hash_password(user_password)
                    user_level = user_info.get('level', 'viewer')
                    
                    # Inserir se não existir
                    cur.execute("""
                        INSERT INTO usuarios (username, password, name, level, is_active, created_at)
                        SELECT %s, %s, %s, %s, TRUE, NOW()
                        WHERE NOT EXISTS (
                            SELECT 1 FROM usuarios WHERE LOWER(username) = LOWER(%s)
                        )
                        RETURNING id
                    """, (username_key, user_password, user_name, user_level, username_key))
                    
                    conn.commit()
                    sync_summary['users_synced'] += 1
                    
                    cur.close()
                    conn.close()
                except Exception as e:
                    print(f"[SYNC] ⚠️ Erro ao sincronizar usuário {username_key}: {str(e)}")
                    try:
                        conn.rollback()
                        cur.close()
                        conn.close()
                    except:
                        pass
        
        print(f"[SYNC] ✅ Sincronização concluída!")
        print(f"[SYNC] 📊 Resumo: {sync_summary}")
        
        return jsonify({
            'success': True,
            'message': 'Dados sincronizados com sucesso com Supabase',
            'summary': sync_summary
        })
    
    except Exception as e:
        print(f"[SYNC] ❌ Erro na sincronização: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            'success': False,
            'error': f'Erro na sincronização: {str(e)}'
        }), 500


# =======================
# 4. START DO SISTEMA
# =======================
SPLASH_DURATION_SECONDS = 5.0
MAIN_URL = f"http://127.0.0.1:{APP_PORT}/"
MAIN_TITLE = "Clínica Leticia Segretti v1.1"
MAIN_WINDOW_SIZE = (1200, 800)


def criar_splash(js_api=None):
    splash_path = get_file_path('splash.png')
    splash_image_src = Path(splash_path).resolve().as_posix() if os.path.isfile(splash_path) else ''

    def _get_center_position(width, height):
        try:
            if sys.platform == 'win32':
                import ctypes
                user32 = ctypes.windll.user32
                screen_width = user32.GetSystemMetrics(0)
                screen_height = user32.GetSystemMetrics(1)
                return max(0, int((screen_width - width) / 2)), max(0, int((screen_height - height) / 2))
        except Exception:
            pass
        return None, None

    splash_width = 780
    splash_height = 520
    x, y = _get_center_position(splash_width, splash_height)

    splash_data_url = ''
    if splash_image_src:
        try:
            with open(get_file_path('splash.png'), 'rb') as f:
                encoded = base64.b64encode(f.read()).decode('ascii')
                splash_data_url = f'data:image/jpeg;base64,{encoded}'
        except Exception:
            splash_data_url = f'file:///{splash_image_src}'

    splash_html = f"""
    <html>
    
      <head>
        <meta charset="utf-8" />
        <style>
          body {{
            margin: 0;
            background: #ffffff;
            color: #ffffff;
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif;
            display: flex;
            align-items: center;
            justify-content: center;
            min-height: 100vh;
            opacity: 0;
            animation: fade-in-out 3s ease-in-out forwards;
          }}
          @keyframes fade-in-out {{
            0% {{ opacity: 0; }}
            20% {{ opacity: 1; }}
            80% {{ opacity: 1; }}
            100% {{ opacity: 0; }}
          }}
          .container {{
            text-align: center;
            max-width: 100%;
          }}
          .logo {{
            max-width: 35vw;
            max-height: 35vh;
            border-radius: 0px;
            box-shadow: none;
          }}
          .title {{
            margin: 24px auto 0;
            font-size: 1.35rem;
            letter-spacing: 0.12em;
            text-transform: uppercase;
            opacity: 0.92;
          }}
          .subtitle {{
            margin-top: 8px;
            color: #c2c2c2;
            font-size: 0.95rem;
          }}
        </style>
      </head>
      <body>
        <div class="container">
          {f'<img class="logo" src="{splash_data_url}" alt="Splash">' if splash_data_url else '<div style="padding:24px;background:#141414;border-radius:18px;max-width:85vw;">Carregando Clínica Leticia Segretti...</div>'}

        </div>
      </body>
    </html>
    """

    window_args = {
        'title': 'Clínica Leticia Segretti',
        'html': splash_html,
        'width': splash_width,
        'height': splash_height,
        'resizable': False,
        'frameless': True,
        'easy_drag': True,
        'fullscreen': False,
        'min_size': (600, 400),
        'confirm_close': False,
        'background_color': '#0d0d0d',
        'on_top': True
    }
    if js_api is not None:
        window_args['js_api'] = js_api
    if x is not None and y is not None:
        window_args['x'] = x
        window_args['y'] = y

    return webview.create_window(**window_args)


def iniciar_app():
    if webview is None:
        raise RuntimeError("pywebview nao esta instalado. Instale as dependencias desktop com requirements.txt.")

    threading.Thread(target=start_flask, daemon=True).start()
    api = Api()

    splash = criar_splash()
    try:
        main_window = webview.create_window(
            MAIN_TITLE,
            MAIN_URL,
            js_api=api,
            width=MAIN_WINDOW_SIZE[0],
            height=MAIN_WINDOW_SIZE[1],
            resizable=True,
            maximized=True,
            min_size=(1024, 700),
            confirm_close=True,
            background_color='#ffffff',
            hidden=True
        )
    except Exception:
        main_window = None

    def fechar_splash_e_mostrar_main():
        nonlocal main_window, splash
        if splash is not None:
            try:
                splash.destroy()
            except Exception:
                pass
            try:
                webview.destroy_window(splash)
            except Exception:
                pass
            try:
                splash.hide()
            except Exception:
                pass
            splash = None

        if main_window is not None:
            try:
                main_window.maximize()
            except Exception:
                pass
            try:
                main_window.show()
            except Exception:
                pass

    def on_loaded():
        threading.Timer(SPLASH_DURATION_SECONDS, fechar_splash_e_mostrar_main).start()

    webview.start(on_loaded)


if __name__ == "__main__":
    server_args = any(arg.lower() in ("--server", "--servidor", "server", "servidor") for arg in sys.argv[1:])
    running_on_web_host = bool(os.environ.get("RENDER") or os.environ.get("PORT"))
    if server_args or running_on_web_host:
        start_flask()
    else:
        iniciar_app()
