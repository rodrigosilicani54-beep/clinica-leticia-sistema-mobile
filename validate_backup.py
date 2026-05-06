#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de validação do sistema de backup
Verifica se todos os caminhos de salvamento estão configurados corretamente
"""

import os
import sys
from pathlib import Path
from datetime import datetime

# Cores para output
class Colors:
    GREEN = '\033[92m'
    RED = '\033[91m'
    YELLOW = '\033[93m'
    BLUE = '\033[94m'
    END = '\033[0m'

def print_success(msg):
    print(f"{Colors.GREEN}✅ {msg}{Colors.END}")

def print_error(msg):
    print(f"{Colors.RED}❌ {msg}{Colors.END}")

def print_warning(msg):
    print(f"{Colors.YELLOW}⚠️  {msg}{Colors.END}")

def print_info(msg):
    print(f"{Colors.BLUE}ℹ️  {msg}{Colors.END}")

def validate_backup_system():
    """Valida o sistema de backup completo"""
    
    print("\n" + "="*60)
    print("VALIDAÇÃO DO SISTEMA DE BACKUP")
    print("="*60 + "\n")
    
    # 1. Validar caminho configurado
    print("1️⃣  VALIDANDO CAMINHO CONFIGURADO")
    print("-" * 60)
    
    expected_backup_dir = r"C:\SISTEMA\backup"
    
    if os.path.exists(expected_backup_dir):
        print_success(f"Diretório existe: {expected_backup_dir}")
    else:
        print_warning(f"Diretório não existe, criando: {expected_backup_dir}")
        try:
            os.makedirs(expected_backup_dir, exist_ok=True)
            print_success(f"Diretório criado com sucesso!")
        except Exception as e:
            print_error(f"Erro ao criar diretório: {str(e)}")
            return False
    
    # 2. Validar permissões
    print("\n2️⃣  VALIDANDO PERMISSÕES")
    print("-" * 60)
    
    if os.access(expected_backup_dir, os.R_OK):
        print_success("Permissão de leitura: ✓")
    else:
        print_error("Permissão de leitura: ✗")
        return False
    
    if os.access(expected_backup_dir, os.W_OK):
        print_success("Permissão de escrita: ✓")
    else:
        print_error("Permissão de escrita: ✗")
        return False
    
    if os.access(expected_backup_dir, os.X_OK):
        print_success("Permissão de execução: ✓")
    else:
        print_error("Permissão de execução: ✗")
        return False
    
    # 3. Testar criação de arquivo
    print("\n3️⃣  TESTANDO CRIAÇÃO DE ARQUIVO")
    print("-" * 60)
    
    test_filename = f"test_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    test_filepath = os.path.join(expected_backup_dir, test_filename)
    
    try:
        with open(test_filepath, 'w', encoding='utf-8') as f:
            f.write(f"Teste de salvamento\n")
            f.write(f"Timestamp: {datetime.now().isoformat()}\n")
            f.write(f"Sistema: backup validation\n")
        
        print_success(f"Arquivo criado: {test_filename}")
        
        if os.path.exists(test_filepath):
            print_success(f"Arquivo verificado no disco")
            
            # Ler arquivo
            with open(test_filepath, 'r', encoding='utf-8') as f:
                content = f.read()
            print_success(f"Arquivo lido com sucesso ({len(content)} bytes)")
            
            # Limpar arquivo
            os.remove(test_filepath)
            print_success(f"Arquivo de teste removido")
        else:
            print_error(f"Arquivo não foi encontrado após criação")
            return False
            
    except Exception as e:
        print_error(f"Erro ao criar/testar arquivo: {str(e)}")
        return False
    
    # 4. Testar criação de arquivo Excel (se openpyxl estiver disponível)
    print("\n4️⃣  TESTANDO CRIAÇÃO DE ARQUIVO EXCEL")
    print("-" * 60)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font
        
        test_excel = f"test_excel_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        test_excel_path = os.path.join(expected_backup_dir, test_excel)
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Teste"
        
        # Adicionar dados
        ws['A1'] = "Teste de Excel"
        ws['A1'].font = Font(bold=True)
        ws['B1'] = datetime.now().isoformat()
        
        # Salvar
        wb.save(test_excel_path)
        print_success(f"Arquivo Excel criado: {test_excel}")
        
        if os.path.exists(test_excel_path):
            file_size = os.path.getsize(test_excel_path)
            print_success(f"Arquivo Excel verificado ({file_size} bytes)")
            
            # Limpar
            os.remove(test_excel_path)
            print_success(f"Arquivo Excel de teste removido")
        else:
            print_error(f"Arquivo Excel não foi encontrado após criação")
            return False
            
    except ImportError:
        print_warning("openpyxl não instalado - pulando teste de Excel")
    except Exception as e:
        print_error(f"Erro ao criar/testar arquivo Excel: {str(e)}")
        return False
    
    # 5. Listar arquivos existentes
    print("\n5️⃣  LISTANDO ARQUIVOS DE BACKUP")
    print("-" * 60)
    
    try:
        files = os.listdir(expected_backup_dir)
        if files:
            print_info(f"Total de arquivos: {len(files)}")
            for filename in sorted(files)[-10:]:  # Últimos 10
                filepath = os.path.join(expected_backup_dir, filename)
                if os.path.isfile(filepath):
                    size = os.path.getsize(filepath)
                    modified = datetime.fromtimestamp(os.path.getmtime(filepath))
                    print(f"  • {filename} ({size} bytes, modificado: {modified.strftime('%d/%m/%Y %H:%M:%S')})")
        else:
            print_info("Nenhum arquivo no diretório")
    except Exception as e:
        print_error(f"Erro ao listar arquivos: {str(e)}")
        return False
    
    # Resumo final
    print("\n" + "="*60)
    print_success("VALIDAÇÃO CONCLUÍDA COM SUCESSO!")
    print("="*60)
    print(f"\n✅ Diretório de backup: {expected_backup_dir}")
    print(f"✅ Todas as funcionalidades validadas\n")
    
    return True

if __name__ == "__main__":
    success = validate_backup_system()
    sys.exit(0 if success else 1)
